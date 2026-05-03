#!/usr/bin/env python3
"""Build a four-page data dictionary for one cohort.

Reads:
    packs/cohorts/<cohort>.yaml        (provider, disease, schema, etc.)
    packs/categories.yaml              (table -> Category map)
    packs/table_descriptions.yaml      (table -> purpose text)
    packs/column_descriptions.yaml     (table.column -> description)
    packs/pii.yaml                     (PII column allowlist + regex)
    packs/variables/<disease>.yaml     (Page-4 clinical variables)

Walks the cohort's Postgres schema (via introspect_cohort.introspect)
and emits:
    Output/<schema>_dictionary.xlsx
    Output/<schema>_dictionary.html

Four sheets / sections (matches century/Data dictionary.pdf):
    Summary  — provider, disease, patients, years_of_data, tables, cols
    Tables   — one row per warehouse table (row_count, column_count, purpose)
    Columns  — one row per physical column (existing inventory w/ enrichment)
    Variables — one row per clinical concept, driven by the disease pack

Usage:
    python build_dictionary.py --cohort mtc_aat
    python build_dictionary.py --cohort mtc_alzheimers
    python build_dictionary.py --cohort mtc_aat --audience sales
    python build_dictionary.py --cohort mtc_aat --dry-run    # no DB required

See README.md for the canonical model and the nine-PR shipping plan.
This script implements PR 1-4 worth of the plan for the MTC cohorts
end-to-end; PR 5-9 (audience filters, PDF renderer, validation, batch
runner, combined views) layer on top of the same CohortModel.
"""
from __future__ import annotations

import argparse
import dataclasses
import datetime as _dt
import hashlib
import json
import re
import subprocess
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

# This module lives in dictionary_v2/, so add the repo root to sys.path
# before importing the existing introspection backbone from there.
# Guarded so repeated imports don't pile up duplicate path entries.
_REPO_ROOT_FOR_IMPORTS = str(Path(__file__).resolve().parent.parent)
if _REPO_ROOT_FOR_IMPORTS not in sys.path:
    sys.path.insert(0, _REPO_ROOT_FOR_IMPORTS)

# Re-use the existing introspection backbone — don't duplicate the
# schema-walking code.
from introspect_cohort import (
    ColumnInfo,
    TableInfo,
    _classify_metric_kind,
    _compile_continuous,
    _compile_date_range,
    _compile_top_values,
    _format_value_distribution,
    _require_psycopg,
    _require_yaml,
    _safe_null_count,
    build_conn_kwargs,
    fetch_person_count,
    load_dotenv,
    LIST_COLUMNS_SQL,
    LIST_TABLES_SQL,
    ROW_COUNT_SQL_TEMPLATE,
)


# --------------------------------------------------------------------------- #
# Paths / version
# --------------------------------------------------------------------------- #

REPO_ROOT = Path(__file__).resolve().parent.parent
PACKS_DIR = REPO_ROOT / "packs"
OUTPUT_DIR = REPO_ROOT / "Output"
INTROSPECT_VERSION = "0.5.0-dictionary"

# Auto-load .env the same way introspect_cohort does.
load_dotenv(REPO_ROOT / ".env")


# --------------------------------------------------------------------------- #
# Dataclasses — the canonical CohortModel
# --------------------------------------------------------------------------- #


@dataclass
class DateCoverage:
    min_date: str | None = None
    max_date: str | None = None
    years_of_data: float | None = None
    contributing_columns: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return dataclasses.asdict(self)


@dataclass
class CohortSummary:
    patient_count: int | None
    table_count: int
    column_count: int
    date_coverage: DateCoverage


@dataclass
class TableRow:
    """One row on Page 2 (Tables).

    `description` / `inclusion_criteria` / `data_source` / `source_table`
    are the Flatiron-style metadata the workbook leads with. `purpose`
    is retained as a backward-compatible alias for `description` so any
    older consumer of the model.to_dict() output keeps working; new code
    should read `description`.
    """
    table_name: str
    category: str
    row_count: int
    column_count: int
    patient_count_in_table: int | None
    purpose: str            # Back-compat alias; mirrors `description`.
    description: str = ""
    inclusion_criteria: str = ""
    data_source: str = ""
    source_table: str = ""


@dataclass
class ColumnRow:
    """One row on Page 3 (Columns). Extends introspect_cohort.ColumnInfo with
    enrichment (category, description, pii flag, patient-level completeness)."""
    category: str
    table: str
    column: str
    description: str
    data_type: str
    values: str
    distribution: str
    median_iqr: str
    completeness_pct: float
    patient_pct: float | None
    extraction_type: str
    pii: bool
    notes: str
    nullable: str = ""      # "Yes" / "No"
    example: str = ""       # Single representative value.
    coding_schema: str = "" # Closed value set, when known.
    data_source: str = ""   # Normalized / Derived / Abstracted / NLP / Enhanced.


@dataclass
class VariableRow:
    """One row on Page 4 (Variables). Driven by packs/variables/<disease>.yaml.

    Column order matches the earlier Century workbook plus the
    Flatiron-style additions:

        Category | Variable | Description | Inclusion Criteria | Table |
        Column(s) | Criteria | Field Type | Example | Coding Schema |
        Values | Distribution | Median (IQR) | Completeness |
        Implemented | % Patient | Data Source | Notes

    `criteria` (raw SQL / configured matcher) renders for the
    technical and customer audiences. Sales / pharma see only
    `inclusion_criteria` (prose). Customer keeps both columns side
    by side so reviewers can map prose to the underlying rule.
    """
    category: str
    variable: str
    description: str
    table: str
    column: str
    criteria: str
    values: str
    distribution: str
    median_iqr: str
    completeness_pct: float | None
    implemented: str        # "Yes" / "No"
    patient_pct: float | None
    extraction_type: str
    notes: str
    pii: bool = False       # Audience filter drops these for sales/pharma.
    inclusion_criteria: str = ""
    field_type: str = ""
    example: str = ""
    coding_schema: str = ""
    data_source: str = ""
    # Structured top-10 observed labels for the variable's column.
    # Populated alongside `values` (comma-joined) so the sales
    # Value Sets cell can render labels verbatim — even when a
    # label itself contains a comma (e.g. an OMOP concept name).
    # Splitting `values` on commas would silently fragment such
    # labels into bogus cell entries.
    top_value_labels: list[str] = field(default_factory=list)
    # Tempus-style sales-spec metadata. `proposal` is the
    # Standard / Custom tag, authored in the variables YAML.
    # Defaults to empty so un-curated rows still build. (Value Sets
    # is data-driven — derived from observed top-N values, no
    # curation field.)
    proposal: str = ""


@dataclass
class CohortModel:
    cohort: str
    provider: str
    disease: str
    schema_name: str
    variant: str
    display_name: str
    description: str
    status: str
    generated_at: str
    git_sha: str
    introspect_version: str
    schema_snapshot_digest: str
    summary: CohortSummary
    tables: list[TableRow]
    columns: list[ColumnRow]
    variables: list[VariableRow]
    # Cohort-level freshness / governance metadata (Commit B). All
    # optional. Read from the cohort YAML and surfaced on the
    # stakeholder cover sheet when present — empty values render
    # nothing (no blank rows). Fields:
    #   data_cutoff_date    — latest event date the ETL pulled
    #                         (ISO date, e.g. "2026-04-15").
    #   last_etl_run        — when the cohort was refreshed
    #                         (ISO date or datetime).
    #   known_limitations   — free-form caveats a reviewer should
    #                         see before evaluating the cohort.
    #   sign_off            — {reviewer, date, notes?} dict naming
    #                         the SME who approved the dictionary.
    data_cutoff_date: str = ""
    last_etl_run: str = ""
    known_limitations: list[str] = field(default_factory=list)
    sign_off: dict[str, str] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        def _conv(v: Any) -> Any:
            if dataclasses.is_dataclass(v):
                return _conv(dataclasses.asdict(v))
            if isinstance(v, dict):
                return {k: _conv(val) for k, val in v.items()}
            if isinstance(v, list):
                return [_conv(x) for x in v]
            return v
        return _conv(dataclasses.asdict(self))


# --------------------------------------------------------------------------- #
# Pack loaders
# --------------------------------------------------------------------------- #


def _yaml_load(path: Path) -> dict[str, Any]:
    yaml = _require_yaml()
    if not path.is_file():
        return {}
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def load_cohort_pack(cohort: str) -> dict[str, Any]:
    path = PACKS_DIR / "cohorts" / f"{cohort}.yaml"
    if not path.is_file():
        raise FileNotFoundError(
            f"cohort pack missing: {path}. Available: "
            + ", ".join(sorted(p.stem for p in (PACKS_DIR / "cohorts").glob("*.yaml")))
        )
    data = _yaml_load(path)
    required = ("provider", "disease", "schema_name", "cohort_name")
    missing = [k for k in required if not data.get(k)]
    if missing:
        raise ValueError(f"{path} missing required fields: {', '.join(missing)}")
    return data


def load_variables_pack(disease_slug: str) -> list[dict[str, Any]]:
    """Load packs/variables/<slug>.yaml and resolve any `include:` list.

    Override semantics: this pack's local rows replace any inherited
    row with the same (category, variable) key. New local rows that
    don't match anything inherited are appended at the end. This lets
    a cohort pack carry a per-cohort `match:` block that *replaces*
    the shared `<disease>_common.yaml`'s fuzzy ILIKE definition for
    that cohort only — without polluting the shared pack and without
    producing duplicate rows in the resolved variable list.

    Inheritance order matches the include declaration: leftmost
    include resolved first, rightmost include can override siblings,
    and the local pack's `variables:` overrides everything inherited.
    """
    if not disease_slug:
        return []
    path = PACKS_DIR / "variables" / f"{disease_slug}.yaml"
    if not path.is_file():
        sys.stderr.write(f"[warn] variables pack not found: {path} -> Page 4 will be empty\n")
        return []
    data = _yaml_load(path)

    # Resolve includes first. Each include's rows can themselves
    # already be the result of overrides at deeper levels of the chain.
    result: list[dict[str, Any]] = []
    for inc in data.get("include") or []:
        result.extend(load_variables_pack(inc))

    # Apply this pack's local rows as overrides on top of the
    # inherited rows. Same (category, variable) → in-place replace
    # (preserving position so audience layouts stay stable). New
    # row → append.
    def _key(r: dict[str, Any]) -> tuple[str, str]:
        return (
            (r.get("category") or "").strip(),
            (r.get("variable") or "").strip(),
        )

    index_by_key: dict[tuple[str, str], int] = {
        _key(r): i for i, r in enumerate(result)
    }
    for local_row in (data.get("variables") or []):
        key = _key(local_row)
        if key in index_by_key:
            result[index_by_key[key]] = local_row
        else:
            index_by_key[key] = len(result)
            result.append(local_row)
    return result


def load_categories_map() -> dict[str, str]:
    """Invert packs/categories.yaml into table_name -> Category."""
    raw = _yaml_load(PACKS_DIR / "categories.yaml").get("categories", {}) or {}
    out: dict[str, str] = {}
    for category, payload in raw.items():
        for table in (payload or {}).get("tables", []) or []:
            out[table] = category
    return out


# Keys the renderer reads from each table_descriptions.yaml entry.
# Listed here so the loader can normalise them to stripped strings
# while still preserving any unknown key the YAML carries (the loader
# contract — see docstring on load_table_descriptions).
_TABLE_DESCRIPTION_KNOWN_KEYS = (
    "description", "inclusion_criteria", "data_source", "source_table",
)


def load_table_descriptions() -> dict[str, dict[str, str]]:
    """Load packs/table_descriptions.yaml.

    Accepts two YAML shapes for backward compatibility:

      - Legacy:  `tables: { foo: "<one-line description>" }`
      - Current: `tables: { foo: { description: "...",
                                    inclusion_criteria: "...",
                                    data_source: "Normalized",
                                    source_table: "OMOP CONDITION_OCCURRENCE" } }`

    Always returns a `{table_name: {field: value}}` map. The four
    known keys above are normalised (missing → empty, leading /
    trailing whitespace stripped). Any additional keys present in the
    YAML are preserved unchanged so future additions to
    table_descriptions.yaml don't require a code change here.
    """
    raw = _yaml_load(PACKS_DIR / "table_descriptions.yaml").get("tables", {}) or {}
    out: dict[str, dict[str, Any]] = {}
    for table, payload in raw.items():
        if isinstance(payload, str):
            out[table] = {k: "" for k in _TABLE_DESCRIPTION_KNOWN_KEYS}
            out[table]["description"] = payload.strip()
        elif isinstance(payload, dict):
            entry: dict[str, Any] = dict(payload)
            for k in _TABLE_DESCRIPTION_KNOWN_KEYS:
                entry[k] = (entry.get(k) or "").strip() if isinstance(entry.get(k), str) else (entry.get(k) or "")
            out[table] = entry
        else:
            out[table] = {k: "" for k in _TABLE_DESCRIPTION_KNOWN_KEYS}
    return out


def load_column_descriptions() -> dict[str, str]:
    return _yaml_load(PACKS_DIR / "column_descriptions.yaml").get("columns", {}) or {}


def load_pii_pack() -> tuple[set[tuple[str, str]], list[re.Pattern[str]]]:
    raw = _yaml_load(PACKS_DIR / "pii.yaml")
    pairs: set[tuple[str, str]] = set()
    for table, cols in (raw.get("pii_columns") or {}).items():
        for c in cols or []:
            pairs.add((table, c))
    patterns = [re.compile(p) for p in (raw.get("pii_name_patterns") or [])]
    return pairs, patterns


def is_pii(table: str, column: str,
           pii_pairs: set[tuple[str, str]],
           pii_patterns: list[re.Pattern[str]]) -> bool:
    if (table, column) in pii_pairs:
        return True
    return any(p.search(column) for p in pii_patterns)


# --------------------------------------------------------------------------- #
# Enrichment heuristics / excludes
# --------------------------------------------------------------------------- #


_SURROGATE_KEY_SUFFIXES = ("_id", "_concept_id")
_SURROGATE_KEY_EXACT = {"person_id", "visit_occurrence_id", "visit_detail_id"}


def is_surrogate_key(column: str) -> bool:
    """IDs we don't want to summarize as continuous measurements."""
    if column in _SURROGATE_KEY_EXACT:
        return True
    return any(column.endswith(suf) for suf in _SURROGATE_KEY_SUFFIXES)


_DATE_COVERAGE_CANDIDATES = [
    ("visit_occurrence", "visit_start_date"),
    ("condition_occurrence", "condition_start_date"),
    ("drug_exposure", "drug_exposure_start_date"),
    ("measurement", "measurement_date"),
    ("observation", "observation_date"),
    ("procedure_occurrence", "procedure_date"),
]


# --------------------------------------------------------------------------- #
# Data Source mapping (Flatiron-style typology)
#
# Three-way classification (Normalized / Derived / Abstracted / NLP /
# Enhanced) is more informative for customers than the binary
# Structured / Unstructured we used to render. Mapping rules:
#
#   - explicit `data_source:` on the pack row wins;
#   - else the table name selects Enhanced / Derived for known
#     curated / derived abstraction tables;
#   - else extraction_type maps as below.
#
# Curated / derived tables are listed explicitly so a generic
# `extraction_type: Structured` row pointing at a curated abstraction
# still surfaces as Enhanced. Adding a new curated table is a one-line
# entry here.
# --------------------------------------------------------------------------- #


_ENHANCED_TABLES: set[str] = {
    # Cohort-curated abstraction tables seen in the live dumps.
    "eosinophil_standardized",
    "standard_profile_data_model",
    "dv_tokenized_profile_data",
    "infusion",
}

_DERIVED_TABLES: set[str] = {
    # Tables computed from upstream OMOP rather than ingested directly.
    "cohort_patients",
}

_EXTRACTION_TO_DATA_SOURCE: dict[str, str] = {
    "structured":   "Normalized",
    "unstructured": "NLP",
    "abstracted":   "Abstracted",
}


def derive_data_source(extraction_type: str, table: str = "",
                       explicit: str = "") -> str:
    """Map a row's extraction_type + table to a Flatiron-style classification.

    `explicit` is the pack row's `data_source:` key when set; it wins
    unconditionally so authors can override on a per-row basis.
    """
    if explicit:
        return explicit.strip()
    if table in _ENHANCED_TABLES:
        return "Enhanced"
    if table in _DERIVED_TABLES:
        return "Derived"
    return _EXTRACTION_TO_DATA_SOURCE.get(
        (extraction_type or "").strip().lower(), "Normalized"
    )


# Per-table fallback prose for variable rows that have NO `criteria:`
# at all (a row that points at a column unfiltered — typically a
# demographic, visit, or generic-diagnosis umbrella row). Mirrors the
# Flatiron convention that every Inclusion Criteria cell is non-blank.
# Unknown tables fall through to a safe generic default in the helper.
_TABLE_NO_CRITERIA_DEFAULTS: dict[str, str] = {
    "person":
        "One record is included for each patient in the cohort.",
    "location":
        "One record is included for each location referenced by a patient in the cohort.",
    "payer_plan_period":
        "One record is included for each insurance coverage window for each patient in the cohort.",
    "visit_occurrence":
        "One record is included for each visit recorded for each patient in the cohort.",
    "visit_detail":
        "One record is included for each visit-detail segment for each patient in the cohort.",
    "condition_occurrence":
        "One record is included for each diagnosis recorded for each patient in the cohort.",
    "drug_exposure":
        "One record is included for each medication exposure recorded for each patient in the cohort.",
    "procedure_occurrence":
        "One record is included for each procedure recorded for each patient in the cohort.",
    "measurement":
        "One record is included for each measurement recorded for each patient in the cohort.",
    "observation":
        "One record is included for each observation recorded for each patient in the cohort.",
    "note":
        "One record is included for each clinical note recorded for each patient in the cohort.",
    "note_nlp":
        "One record is included for each NLP-extracted concept identified within a clinical note.",
    "document":
        "One record is included for each document attached to a patient encounter.",
    "death":
        "One record is included for each patient with a recorded death.",
    "infusion":
        "One record is included for each infusion episode recorded for each patient in the cohort.",
}


def derive_inclusion_criteria(criteria_sql: str, explicit: str = "",
                              table: str = "") -> str:
    """Resolve a row's customer-visible Inclusion Criteria sentence.

    Resolution order:
      1. `explicit` — pack-author prose, returned unchanged.
      2. No criteria at all → table-keyed default (e.g. "One record
         is included for each diagnosis recorded for each patient in
         the cohort.") so the rendered Inclusion Criteria cell is
         never blank for a generic umbrella row.
      3. Any `criteria:` present → empty string. There is no
         auto-translation from SQL — every row with a criteria
         clause must carry an explicit `inclusion_criteria:`. The
         older "concept matches 'X'" friendly fallback was removed
         because it produced QA-style copy ("Records where the
         observation concept matches 'language'.") rather than
         Flatiron-style clinical prose. The validator enforces the
         contract.
    """
    if explicit:
        return explicit.strip()
    sql = (criteria_sql or "").strip()
    if not sql:
        return _TABLE_NO_CRITERIA_DEFAULTS.get(
            (table or "").strip(),
            "Records are included for each patient in the cohort.",
        )
    return ""


def _example_from_column_info(ci: "ColumnInfo") -> str:
    """Single representative value for the Example column.

    Prefers categorical top-1 (matches what most customers want to
    see). Falls back to the Min from a continuous / date numeric
    summary. Returns empty when the column has no observed values.
    """
    if ci.top_values:
        first = ci.top_values[0][0]
        return first if first else ""
    summary = ci.numeric_summary or ""
    m = re.search(r"Min:\s*([^,;]+)", summary)
    if m:
        return m.group(1).strip()
    return ""


# --------------------------------------------------------------------------- #
# Introspection — extends introspect_cohort.introspect() with PR-1..PR-4 logic
# --------------------------------------------------------------------------- #


def _patient_completeness_sql(schema: str, table: str, column: str) -> str:
    return (
        f'SELECT COUNT(DISTINCT "person_id") '
        f'FROM "{schema}"."{table}" '
        f'WHERE "{column}" IS NOT NULL;'
    )


def _distinct_patients_in_table_sql(schema: str, table: str) -> str:
    return f'SELECT COUNT(DISTINCT "person_id") FROM "{schema}"."{table}";'


def introspect_cohort(
    conn,
    schema: str,
    sample_values_default: int = 5,
    sample_values_concept: int = 20,
) -> tuple[list[ColumnInfo], list[TableInfo], dict[str, int | None], set[str]]:
    """Walk the cohort schema with PR-1 / PR-3 rules baked in.

    PR 1: skip surrogate keys from continuous summary, bump sample depth
          for *_concept_name columns, collapse empty tables.
    PR 3: patient-level completeness for tables with person_id.

    Returns (columns, tables, patient_count_per_table, tables_with_person_id).
    The last set is a separate signal from the patient_count dict because
    "no person_id column" and "query failed" would otherwise both show up
    as None.
    """
    psycopg = _require_psycopg()  # noqa: F841 — import-time validation only
    columns_out: list[ColumnInfo] = []
    tables_out: list[TableInfo] = []
    patients_per_table: dict[str, int | None] = {}
    tables_with_person_id: set[str] = set()

    with conn.cursor() as cur:
        cur.execute(LIST_TABLES_SQL, (schema,))
        table_rows = cur.fetchall()
    tables = [t for (t, _) in table_rows]

    for table in tables:
        with conn.cursor() as cur:
            cur.execute(ROW_COUNT_SQL_TEMPLATE.format(schema=schema, table=table))
            row_count = cur.fetchone()[0]
            cur.execute(LIST_COLUMNS_SQL, (schema, table))
            raw_columns = cur.fetchall()

        tables_out.append(TableInfo(
            name=table, row_count=row_count, column_count=len(raw_columns)
        ))

        # patient count in this table (if person_id present)
        col_names = {c[0] for c in raw_columns}
        has_person_id = "person_id" in col_names
        if has_person_id:
            tables_with_person_id.add(table)
        if has_person_id and row_count > 0:
            try:
                with conn.cursor() as cur:
                    cur.execute(_distinct_patients_in_table_sql(schema, table))
                    patients_per_table[table] = int(cur.fetchone()[0])
            except Exception:
                conn.rollback()
                patients_per_table[table] = None
        else:
            patients_per_table[table] = None

        # Empty-table collapse: list columns but skip per-column summaries.
        if row_count == 0:
            for column_name, data_type, is_nullable_str, _ml, _p in raw_columns:
                columns_out.append(ColumnInfo(
                    schema=schema, table=table, column=column_name,
                    data_type=data_type,
                    is_nullable=(is_nullable_str == "YES"),
                    row_count=0, null_count=0, completeness_pct=0.0,
                ))
            continue

        for column_name, data_type, is_nullable_str, _ml, _p in raw_columns:
            is_nullable = is_nullable_str == "YES"
            kind = _classify_metric_kind(data_type)
            null_count = _safe_null_count(conn, schema, table, column_name)

            value_distribution = ""
            numeric_summary = ""
            median_iqr = ""
            top_values: list[tuple[str, int]] = []

            # PR 1: exclude surrogate keys from continuous summary
            if kind == "continuous" and not is_surrogate_key(column_name):
                numeric_summary, median_iqr = _compile_continuous(
                    conn, schema, table, column_name
                )
            elif kind == "date":
                numeric_summary = _compile_date_range(
                    conn, schema, table, column_name
                )
            elif kind == "categorical":
                # PR 1: deeper sample for *_concept_name columns
                limit = sample_values_concept if column_name.endswith("_concept_name") \
                    else sample_values_default
                top_values = _compile_top_values(
                    conn, schema, table, column_name, limit=limit,
                )
                value_distribution = _format_value_distribution(
                    top_values, null_count, row_count
                )

            completeness = (1 - null_count / row_count) * 100 if row_count else 0.0
            columns_out.append(ColumnInfo(
                schema=schema, table=table, column=column_name,
                data_type=data_type, is_nullable=is_nullable,
                row_count=row_count, null_count=null_count,
                completeness_pct=completeness,
                value_distribution=value_distribution,
                numeric_summary=numeric_summary,
                median_iqr=median_iqr,
                top_values=top_values,
            ))

    return columns_out, tables_out, patients_per_table, tables_with_person_id


def compute_patient_completeness(
    conn, schema: str, column: ColumnInfo, total_patients: int | None,
) -> float | None:
    """% Patient = distinct person_ids with col IS NOT NULL / total patients."""
    if total_patients is None or total_patients <= 0:
        return None
    if column.row_count == 0:
        return 0.0
    try:
        with conn.cursor() as cur:
            cur.execute(
                _patient_completeness_sql(schema, column.table, column.column)
            )
            n = int(cur.fetchone()[0])
    except Exception:
        conn.rollback()
        return None
    return 100.0 * n / total_patients


# --------------------------------------------------------------------------- #
# Date coverage rollup (PR 4)
# --------------------------------------------------------------------------- #


_DATE_RE = re.compile(r"Min: (\S+), Max: (\S+)")


def compute_date_coverage(columns: list[ColumnInfo]) -> DateCoverage:
    """Scan the per-column date-range summaries and roll up the earliest /
    latest date across the approved clinical-date candidates."""
    min_seen: str | None = None
    max_seen: str | None = None
    contributing: list[str] = []

    by_key = {(c.table, c.column): c for c in columns}
    for table, col in _DATE_COVERAGE_CANDIDATES:
        info = by_key.get((table, col))
        if info is None or not info.numeric_summary:
            continue
        m = _DATE_RE.search(info.numeric_summary)
        if not m:
            continue
        mn, mx = m.group(1), m.group(2)
        contributing.append(f"{table}.{col}")
        if min_seen is None or mn < min_seen:
            min_seen = mn
        if max_seen is None or mx > max_seen:
            max_seen = mx

    years = None
    if min_seen and max_seen:
        try:
            d_min = _dt.date.fromisoformat(min_seen[:10])
            d_max = _dt.date.fromisoformat(max_seen[:10])
            years = round((d_max - d_min).days / 365.25, 2)
        except ValueError:
            years = None

    return DateCoverage(
        min_date=min_seen,
        max_date=max_seen,
        years_of_data=years,
        contributing_columns=contributing,
    )


# --------------------------------------------------------------------------- #
# Variable (Page 4) resolution
# --------------------------------------------------------------------------- #


def _format_top_values_from_rows(
    rows: list[tuple[str, int]], total: int,
) -> tuple[str, str, list[str]]:
    """(values_cell, distribution_cell, labels) from SQL top-N rows.

    Returns three projections of the same data:
      - values_cell:  comma-joined top-10 labels (legacy display in
                      the technical / customer Variables sheets).
      - distribution_cell:  semicolon-joined `label: count (pct%)`.
      - labels:       the structured top-10 label list, preserved
                      verbatim. Sales Value Sets reads this so a
                      label that itself contains a comma (e.g. an
                      OMOP concept name with a comma) renders as
                      one cell entry, not two.
    """
    if not rows or total <= 0:
        return "", "", []
    labels = [r[0] if r[0] else "(null)" for r in rows[:10]]
    values = ", ".join(labels)
    distribution = "; ".join(
        f"{r[0] if r[0] else '(null)'}: {r[1]} ({100.0 * r[1] / total:.1f}%)"
        for r in rows
    )
    return values, distribution, labels


_TEXT_COLUMN_PATTERNS = (
    re.compile(r"_text$"),
    re.compile(r"_note$"),
    re.compile(r"^note_text$"),
)


def _is_freetext_column(column: str) -> bool:
    return any(p.search(column) for p in _TEXT_COLUMN_PATTERNS)


# --------------------------------------------------------------------------- #
# Structured `match:` config for variable Criteria. Reviewer asked for
# the Criteria column to live in config and be exact matches to the
# column, not fuzzy ILIKE. Variable YAML carries:
#
#   match:
#     column: drug_concept_name
#     values:
#       - aspirin 81 MG Oral Tablet
#       - aspirin 325 MG Oral Tablet
#
# which compiles to: "drug_concept_name" IN ('aspirin 81 MG Oral Tablet', ...)
#
# Values can also live in a separate file when the list is long. The
# path is resolved relative to packs/, so omit the `packs/` prefix:
#
#   match:
#     column: drug_concept_name
#     values_file: value_sets/aspirin.yaml   # → packs/value_sets/aspirin.yaml
#
# The build never derives matches from observed data — see
# dictionary_v2/discover_exact_matches.py for the offline discovery
# workflow that proposes additions to a `match:` list.
# --------------------------------------------------------------------------- #


def _sql_quote(value: str) -> str:
    """Escape a string for inclusion in a single-quoted SQL literal."""
    return value.replace("'", "''")


# Allowed values for the sales spec's `proposal:` field. Empty is
# also OK (un-curated row); anything else is a typo.
SALES_PROPOSAL_VALUES = ("Standard", "Custom")


def _load_match_values_file(rel_path: str) -> list[str]:
    """Load `values:` list from a YAML file referenced by `values_file:`.

    Path is resolved relative to PACKS_DIR. Accepts either a top-level
    list or a `values:` key for forward-compatibility with richer
    value-set metadata.
    """
    path = (PACKS_DIR / rel_path).resolve()
    data = _yaml_load(path)
    if isinstance(data, list):
        return [str(v) for v in data]
    if isinstance(data, dict):
        return [str(v) for v in (data.get("values") or [])]
    return []


def compile_match_block(match: dict[str, Any] | None) -> str:
    """Return `"<column>" IN (...)` SQL from a structured match block,
    or "" if the block is missing/empty.

    Two source forms, mutually exclusive:

    - `concept_ids:` (preferred for clinical accuracy) — list of
      OMOP concept-ID integers. Compiles to a bare-integer
      IN list with NO database vocabulary lookup at build time:
          "drug_concept_id" IN (40221901, 793143, 35606214)
      Concept names are display metadata only — render them via
      the variable's `column:` (typically the matching
      `*_concept_name`) on the Observed Values cell. Keeping the
      filter integer-only means the matcher is deterministic even
      if a vocabulary lookup is unavailable.

    - `values:` / `values_file:` — list of string labels. Compiles
      to a quoted IN list:
          "drug_concept_name" IN ('Lecanemab', 'Donanemab')
      Both inline and file values are unioned and deduplicated
      while preserving first-seen order.

    If both `concept_ids` and `values` are populated on the same
    block, `concept_ids` wins (concept-ID matching is canonical;
    string matching is a fallback). The validator flags the
    combination so packs don't drift into ambiguity.
    """
    if not isinstance(match, dict):
        return ""
    column = (match.get("column") or "").strip()
    if not column:
        return ""

    # Concept-ID branch — bare integers, no name lookup.
    raw_ids = match.get("concept_ids") or []
    if raw_ids:
        ids: list[int] = []
        seen_ids: set[int] = set()
        for v in raw_ids:
            try:
                i = int(v)
            except (TypeError, ValueError):
                continue
            if i not in seen_ids:
                seen_ids.add(i)
                ids.append(i)
        if ids:
            return f'"{column}" IN ({", ".join(str(i) for i in ids)})'
        return ""

    # String values branch (legacy + values_file).
    values: list[str] = []
    seen: set[str] = set()
    for v in (match.get("values") or []):
        s = str(v)
        if s not in seen:
            seen.add(s)
            values.append(s)
    values_file = (match.get("values_file") or "").strip()
    if values_file:
        for v in _load_match_values_file(values_file):
            if v not in seen:
                seen.add(v)
                values.append(v)

    if not values:
        return ""
    quoted = ", ".join(f"'{_sql_quote(v)}'" for v in values)
    return f'"{column}" IN ({quoted})'


def resolve_variables(
    conn, schema: str,
    variables_pack: list[dict[str, Any]],
    total_patients: int | None,
    pii_pairs: set[tuple[str, str]] | None = None,
    pii_patterns: list[re.Pattern[str]] | None = None,
    tables_with_person_id: set[str] | None = None,
    column_types: dict[tuple[str, str], str] | None = None,
) -> list[VariableRow]:
    """For each entry in the disease pack, query the cohort and populate
    Values / Distribution / Median (IQR) / Completeness / Implemented /
    % Patient.

    Pack fields consumed:
        table           (required)
        column          (required — drives the "Column(s)" display cell)
        expression      (optional — SQL expression used in place of
                         a bare `"<column>"` reference; lets the pack
                         say `LEFT("zip", 3)` without losing the `zip`
                         label for display)
        criteria        (optional — SQL WHERE fragment)
        extraction_type (Structured / Abstracted / Unstructured)
        category / variable / description / notes

    Skips:
      - GROUP BY "{column}" for `extraction_type: Unstructured` rows or
        column names matching `*_text` / `*_note` — unique-per-row
        values + expensive query on note tables.
      - COUNT(DISTINCT person_id) for tables not in
        `tables_with_person_id` (location and other dimension tables
        don't carry a patient FK, so the query would fail + roll back
        every column).
    """
    pii_pairs = pii_pairs or set()
    pii_patterns = pii_patterns or []
    tables_with_person_id = tables_with_person_id or set()
    column_types = column_types or {}

    out: list[VariableRow] = []
    for v in variables_pack:
        table = v.get("table") or ""
        column = v.get("column") or ""
        expression = (v.get("expression") or "").strip() or f'"{column}"'
        criteria = (v.get("criteria") or "").strip()
        category = v.get("category") or ""
        # Structured `match:` block compiles to strict `column IN (...)`
        # and overrides any free-form `criteria:` for both display and
        # filtering. The data dictionary thus shows config-owned exact
        # matches instead of fuzzy ILIKE, per reviewer feedback.
        # Variables without a `match:` block keep their hand-written
        # criteria untouched — the build never derives criteria from
        # observed data, since that would let observations redefine
        # the clinical variable. See dictionary_v2/discover_exact_matches.py
        # for the offline discovery workflow.
        match_block = v.get("match")
        match_sql = compile_match_block(match_block) if match_block else ""
        if match_sql:
            criteria = match_sql
        variable_name = v.get("variable") or column
        description = v.get("description") or ""
        extraction = v.get("extraction_type") or "Structured"
        notes = v.get("notes") or ""

        values_cell = ""
        distribution_cell = ""
        top_value_labels: list[str] = []
        median_iqr_cell = ""
        completeness_pct: float | None = None
        implemented = "No"
        patient_pct: float | None = None

        # Two WHERE forms:
        #   `where_criteria` scopes to rows matching the variable's
        #                    criteria (null or not) — Completeness
        #                    denominator.
        #   `where_nonnull`  adds the is-not-null guard — the rows
        #                    that actually contribute data.
        where_criteria = f"({criteria})" if criteria else "TRUE"
        where_nonnull = f"{expression} IS NOT NULL"
        if criteria:
            where_nonnull += f" AND ({criteria})"

        total_with_criteria = 0
        try:
            with conn.cursor() as cur:
                cur.execute(
                    f'SELECT COUNT(*) FROM "{schema}"."{table}" '
                    f'WHERE {where_criteria};'
                )
                total_with_criteria = int(cur.fetchone()[0])
        except Exception as exc:
            sys.stderr.write(
                f"[warn] {category}/{variable_name}: criteria count failed "
                f"({table}.{column}): {exc}\n"
            )
            conn.rollback()

        total_nonnull = 0
        try:
            with conn.cursor() as cur:
                cur.execute(
                    f'SELECT COUNT(*) FROM "{schema}"."{table}" '
                    f'WHERE {where_nonnull};'
                )
                total_nonnull = int(cur.fetchone()[0])
        except Exception as exc:
            sys.stderr.write(
                f"[warn] {category}/{variable_name}: nonnull count failed "
                f"({table}.{column}): {exc}\n"
            )
            conn.rollback()

        if total_with_criteria > 0:
            completeness_pct = 100.0 * total_nonnull / total_with_criteria

        skip_top_values = (
            extraction.lower() == "unstructured" or _is_freetext_column(column)
        )

        # Type-classify so date columns get Min/Max instead of top-N of
        # the most common exact dates, which the reviewer called out as
        # unhelpful and inconsistent with the column-inventory page.
        # Expression-backed rows stay categorical because the expression
        # output type may not match the raw column's data_type.
        raw_type = column_types.get((table, column), "")
        raw_kind = _classify_metric_kind(raw_type) if raw_type else "categorical"
        has_expression = v.get("expression") is not None
        treat_as_date = raw_kind == "date" and not has_expression

        if total_nonnull > 0:
            implemented = "Yes"

            if skip_top_values:
                distribution_cell = (
                    f"{total_nonnull:,} rows; values not aggregated (free text)"
                )
            elif treat_as_date:
                # Date/timestamp columns: min/max range is what reviewers
                # want; the column-inventory page uses the same format.
                distribution_cell = (
                    _compile_date_range_filtered(
                        conn, schema, table, column, where_nonnull
                    )
                    or f"{total_nonnull:,} rows"
                )
            else:
                try:
                    with conn.cursor() as cur:
                        cur.execute(
                            f'SELECT {expression}::text AS v, COUNT(*) AS n '
                            f'FROM "{schema}"."{table}" WHERE {where_nonnull} '
                            f'GROUP BY {expression} ORDER BY n DESC LIMIT 10;'
                        )
                        rows = [(str(r[0]), int(r[1])) for r in cur.fetchall()]
                    values_cell, distribution_cell, top_value_labels = (
                        _format_top_values_from_rows(rows, total_nonnull)
                    )
                except Exception as exc:
                    sys.stderr.write(
                        f"[warn] {category}/{variable_name}: top-values query "
                        f"failed: {exc}\n"
                    )
                    conn.rollback()

            # Median (IQR) for numeric-typed underlying columns. Skip when:
            #   - an expression is used (expression's output type may
            #     not match the raw column's data_type — e.g. LEFT(zip,3)
            #     is text even though zip might be numeric), or
            #   - the column is a surrogate key / concept id (numeric
            #     by type but not a measurement — a median of
            #     drug_concept_id values is meaningless).
            if (
                raw_kind == "continuous"
                and not has_expression
                and not is_surrogate_key(column)
            ):
                median_iqr_cell = _compile_continuous_filtered(
                    conn, schema, table, column, where_nonnull
                )

            # % Patient — only if the table actually has person_id.
            if (
                table in tables_with_person_id
                and total_patients and total_patients > 0
            ):
                try:
                    with conn.cursor() as cur:
                        cur.execute(
                            f'SELECT COUNT(DISTINCT "person_id") '
                            f'FROM "{schema}"."{table}" WHERE {where_nonnull};'
                        )
                        n = int(cur.fetchone()[0])
                    patient_pct = 100.0 * n / total_patients
                except Exception:
                    conn.rollback()

        # Flatiron-style additions. All optional pack keys; safe defaults
        # so older packs without them keep rendering unchanged.
        explicit_inclusion = (v.get("inclusion_criteria") or "").strip()
        explicit_data_source = (v.get("data_source") or "").strip()
        coding_schema = (v.get("coding_schema") or "").strip()

        # Field Type comes from the underlying column's data_type when
        # known; an `expression:`-backed row falls back to the pack's
        # column type because the expression's output type isn't in
        # information_schema.
        field_type = column_types.get((table, column), "")

        # Single representative example value. Prefer the structured
        # top_value_labels list — labels can contain commas (OMOP
        # concept names like "Cancer, malignant"), so splitting
        # values_cell would fragment them. Fall back to parsing the
        # distribution cell's "Min: ..." prefix when categorical
        # sampling didn't fire (date / continuous).
        example = ""
        if top_value_labels:
            example = top_value_labels[0]
        elif distribution_cell:
            m = re.search(r"Min:\s*([^,;]+)", distribution_cell)
            if m:
                example = m.group(1).strip()

        out.append(VariableRow(
            category=category,
            variable=variable_name,
            description=description,
            table=table,
            column=column,
            criteria=criteria,
            values=values_cell,
            distribution=distribution_cell,
            top_value_labels=top_value_labels,
            median_iqr=median_iqr_cell,
            completeness_pct=completeness_pct,
            implemented=implemented,
            patient_pct=patient_pct,
            extraction_type=extraction,
            notes=notes,
            pii=is_pii(table, column, pii_pairs, pii_patterns),
            inclusion_criteria=derive_inclusion_criteria(
                criteria, explicit_inclusion, table,
            ),
            field_type=field_type,
            example=example,
            coding_schema=coding_schema,
            data_source=derive_data_source(extraction, table, explicit_data_source),
            proposal=(v.get("proposal") or "").strip(),
        ))
    return out


def _compile_continuous_filtered(
    conn, schema: str, table: str, column: str, where: str
) -> str:
    """`Median (IQR)` cell for a numeric column, scoped to rows matching
    the variable's `where` clause. Returns empty string on failure or
    when the subset has no rows."""
    sql = f"""
    SELECT
      PERCENTILE_CONT(0.25) WITHIN GROUP (ORDER BY "{column}")::text,
      PERCENTILE_CONT(0.5)  WITHIN GROUP (ORDER BY "{column}")::text,
      PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY "{column}")::text
    FROM "{schema}"."{table}"
    WHERE {where};
    """
    try:
        with conn.cursor() as cur:
            cur.execute(sql)
            row = cur.fetchone()
    except Exception:
        conn.rollback()
        return ""
    if not row or row[1] is None:
        return ""
    q1, median, q3 = row
    return f"Median: {median} (IQR: {q1}-{q3})"


def _compile_date_range_filtered(
    conn, schema: str, table: str, column: str, where: str
) -> str:
    """`Min: X, Max: Y` for a date/timestamp column, scoped to the
    variable's `where` clause. Matches the format used by the column
    inventory page so date variables read the same way on both sheets."""
    sql = f'''
    SELECT MIN("{column}")::text, MAX("{column}")::text
    FROM "{schema}"."{table}"
    WHERE {where};
    '''
    try:
        with conn.cursor() as cur:
            cur.execute(sql)
            row = cur.fetchone()
    except Exception:
        conn.rollback()
        return ""
    if not row or row[0] is None:
        return ""
    return f"Min: {row[0]}, Max: {row[1]}"


# --------------------------------------------------------------------------- #
# Model builder — orchestrates introspection + pack enrichment
# --------------------------------------------------------------------------- #


def _git_sha() -> str:
    try:
        return subprocess.check_output(
            ["git", "-C", str(REPO_ROOT), "rev-parse", "--short", "HEAD"],
            stderr=subprocess.DEVNULL,
        ).decode("utf-8").strip()
    except Exception:
        return "unknown"


def _schema_snapshot_digest(columns: list[ColumnInfo]) -> str:
    """Stable hash of (table, column, data_type) tuples so we can diff for
    drift in a later PR."""
    h = hashlib.sha256()
    for c in sorted(columns, key=lambda x: (x.table, x.column)):
        h.update(f"{c.table}.{c.column}:{c.data_type}\n".encode("utf-8"))
    return f"sha256:{h.hexdigest()}"


def build_model(
    cohort: str,
    conn,
    dry_run: bool = False,
) -> CohortModel:
    pack = load_cohort_pack(cohort)
    categories_map = load_categories_map()
    table_descriptions = load_table_descriptions()
    column_descriptions = load_column_descriptions()
    pii_pairs, pii_patterns = load_pii_pack()
    variables_pack = load_variables_pack(pack.get("variables_pack", ""))

    schema = pack["schema_name"]

    if dry_run:
        # Synthesise empty inventory so the renderer path is still exercisable
        columns_raw: list[ColumnInfo] = []
        tables_raw: list[TableInfo] = []
        patients_per_table: dict[str, int | None] = {}
        tables_with_person_id: set[str] = set()
        total_patients: int | None = None
        variables_rows = []
        for v in variables_pack:
            extraction = v.get("extraction_type", "Structured")
            table = v.get("table", "")
            criteria = (v.get("criteria") or "").strip()
            # Mirror the live path: a structured `match:` block compiles
            # to strict `column IN (...)` and overrides any free-form
            # `criteria:`. Without this, dry-run previews show stale
            # fuzzy criteria while live builds emit strict IN clauses.
            match_sql = compile_match_block(v.get("match"))
            if match_sql:
                criteria = match_sql
            variables_rows.append(VariableRow(
                category=v.get("category", ""),
                variable=v.get("variable", v.get("column", "")),
                description=v.get("description", ""),
                table=table,
                column=v.get("column", ""),
                criteria=criteria,
                values="", distribution="",
                median_iqr="",
                completeness_pct=None,
                implemented="No",
                patient_pct=None,
                extraction_type=extraction,
                notes=v.get("notes", ""),
                pii=is_pii(
                    table, v.get("column", ""), pii_pairs, pii_patterns,
                ),
                inclusion_criteria=derive_inclusion_criteria(
                    criteria, (v.get("inclusion_criteria") or "").strip(), table,
                ),
                field_type="",
                example="",
                coding_schema=(v.get("coding_schema") or "").strip(),
                data_source=derive_data_source(
                    extraction, table, (v.get("data_source") or "").strip(),
                ),
                proposal=(v.get("proposal") or "").strip(),
            ))
    else:
        total_patients = fetch_person_count(conn, schema)
        columns_raw, tables_raw, patients_per_table, tables_with_person_id = \
            introspect_cohort(conn, schema)
        column_types = {(c.table, c.column): c.data_type for c in columns_raw}
        variables_rows = resolve_variables(
            conn, schema, variables_pack, total_patients,
            pii_pairs=pii_pairs, pii_patterns=pii_patterns,
            tables_with_person_id=tables_with_person_id,
            column_types=column_types,
        )

    # Page 2 — Tables
    cohort_category_overrides = (pack.get("category_rules") or {})
    # invert cohort-level override: table -> Category
    override_map: dict[str, str] = {}
    for cat, payload in cohort_category_overrides.items():
        for t in (payload or {}).get("tables", []) or []:
            override_map[t] = cat

    def _category_for(table: str) -> str:
        return override_map.get(table) or categories_map.get(table) or "Other"

    def _table_meta(name: str) -> dict[str, str]:
        return table_descriptions.get(name) or {
            "description": "", "inclusion_criteria": "",
            "data_source": "", "source_table": "",
        }

    table_rows = [
        TableRow(
            table_name=t.name,
            category=_category_for(t.name),
            row_count=t.row_count,
            column_count=t.column_count,
            patient_count_in_table=patients_per_table.get(t.name),
            # `purpose` mirrors `description` for back-compat with
            # consumers that read model.to_dict().
            purpose=_table_meta(t.name)["description"],
            description=_table_meta(t.name)["description"],
            inclusion_criteria=_table_meta(t.name)["inclusion_criteria"],
            data_source=_table_meta(t.name)["data_source"],
            source_table=_table_meta(t.name)["source_table"],
        )
        for t in tables_raw
    ]

    # Page 3 — Columns
    column_rows: list[ColumnRow] = []
    for ci in columns_raw:
        kind = _classify_metric_kind(ci.data_type)
        extraction = "Unstructured" if kind == "unstructured" else "Structured"
        values_cell = ", ".join(v for v, _ in ci.top_values[:10])
        distribution_cell = ci.value_distribution or ci.numeric_summary
        pii = is_pii(ci.table, ci.column, pii_pairs, pii_patterns)
        if pii:
            extraction = "PII"
        patient_pct: float | None = None
        if not dry_run and ci.table in tables_with_person_id:
            # Denominator is the cohort's total patient count, NOT the
            # per-table distinct-patient count — otherwise sparse tables
            # always show ~100% and the column hides real cohort coverage.
            # Tables without a person_id column (e.g. location) stay None
            # — running the query there would just fail + roll back.
            patient_pct = compute_patient_completeness(
                conn, schema, ci, total_patients
            )
        column_rows.append(ColumnRow(
            category=_category_for(ci.table),
            table=ci.table,
            column=ci.column,
            description=column_descriptions.get(f"{ci.table}.{ci.column}", ""),
            data_type=ci.data_type,
            values=values_cell,
            distribution=distribution_cell,
            median_iqr=ci.median_iqr,
            completeness_pct=ci.completeness_pct,
            patient_pct=patient_pct,
            extraction_type=extraction,
            pii=pii,
            notes="",
            nullable="Yes" if ci.is_nullable else "No",
            example=_example_from_column_info(ci),
            coding_schema="",
            data_source=derive_data_source(extraction, ci.table),
        ))

    # Page 1 — Summary
    date_coverage = compute_date_coverage(columns_raw)
    summary = CohortSummary(
        patient_count=total_patients,
        table_count=len(tables_raw),
        column_count=len(columns_raw),
        date_coverage=date_coverage,
    )

    # Sort the three data sheets by (category, secondary key) before
    # they hit the model. Reviewers expect a-z ordering by category
    # first and then variable / table / column within each category;
    # the YAML packs are authored in clinical groupings rather than
    # alphabetically. Sorting here means every audience and every
    # output format (xlsx / html / json) gets consistent ordering.
    # Sort keys are case-insensitive so "Diagnoses" and "diagnoses"
    # don't split into two clusters.
    def _ci(s: Any) -> str:
        return (s or "").strip().lower() if isinstance(s, str) else ""

    table_rows = sorted(
        table_rows, key=lambda t: (_ci(t.category), _ci(t.table_name)),
    )
    column_rows = sorted(
        column_rows,
        key=lambda c: (_ci(c.category), _ci(c.table), _ci(c.column)),
    )
    variables_rows = sorted(
        variables_rows, key=lambda v: (_ci(v.category), _ci(v.variable)),
    )

    return CohortModel(
        cohort=pack["cohort_name"],
        provider=pack["provider"],
        disease=pack["disease"],
        schema_name=schema,
        variant=pack.get("variant", "raw"),
        display_name=pack.get("display_name") or pack["cohort_name"],
        description=(pack.get("description") or "").strip(),
        status=pack.get("status", "active"),
        generated_at=_dt.datetime.now(_dt.timezone.utc).isoformat(timespec="seconds"),
        git_sha=_git_sha(),
        introspect_version=INTROSPECT_VERSION,
        schema_snapshot_digest=_schema_snapshot_digest(columns_raw),
        summary=summary,
        tables=table_rows,
        columns=column_rows,
        variables=variables_rows,
        data_cutoff_date=str(pack.get("data_cutoff_date") or "").strip(),
        last_etl_run=str(pack.get("last_etl_run") or "").strip(),
        known_limitations=[
            str(x).strip()
            for x in (pack.get("known_limitations") or [])
            if str(x).strip()
        ],
        sign_off=(
            {k: str(v).strip() for k, v in (pack.get("sign_off") or {}).items()}
            if isinstance(pack.get("sign_off"), dict) else {}
        ),
    )


# --------------------------------------------------------------------------- #
# Audience filter (PR 5 — applied to the canonical model pre-render)
# --------------------------------------------------------------------------- #


# Audience -> {section: visible?}. Single source of truth driving both
# the model filter and the renderer omits.
AUDIENCE_VISIBILITY: dict[str, dict[str, bool]] = {
    "technical": {"summary": True, "tables": True, "columns": True, "variables": True},
    # Sales ships three sheets — Summary cover, Tables overview,
    # and the Tempus-style Variables spec. Columns (schema map at
    # the physical-column level) is dropped: a sales partner reads
    # Variables for clinical content; the column-level schema map
    # is mostly noise for that audience and can be served by the
    # technical-audience output when an engineer needs it.
    "sales":     {"summary": True, "tables": True, "columns": False, "variables": True},
    "pharma":    {"summary": True, "tables": False, "columns": False, "variables": True},
    # Customer audience (PR-B). Opt-in via `--audience customer`. Keeps
    # all four sheets visible but trims columns and filters internal
    # tables. See per-audience layouts below.
    "customer":  {"summary": True, "tables": True, "columns": True, "variables": True},
}


# Tables that are scaffolding / internal and should not appear in the
# customer dictionary. Sourced from packs/dictionary_layout.yaml
# (customer.exclude_tables) with optional per-cohort overrides under
# cohorts.<slug>.customer.{exclude_tables,extra_exclude_tables}.
#
# The hard-coded fallback below is the historical PR-B list and is
# only used if the layout YAML is missing or malformed; tests pin the
# config-driven path.
_CUSTOMER_TABLE_EXCLUDES_FALLBACK: frozenset[str] = frozenset({
    "standard_profile_data_model",
    "cohort_patients",
    "dv_tokenized_profile_data",
})

DICTIONARY_LAYOUT_PATH = PACKS_DIR / "dictionary_layout.yaml"


def _load_dictionary_layout() -> dict[str, Any]:
    """Load packs/dictionary_layout.yaml, or {} if missing."""
    return _yaml_load(DICTIONARY_LAYOUT_PATH)


def customer_table_excludes(
    cohort: str | list[str] | None = None,
) -> frozenset[str]:
    """Resolve the customer-audience table exclude list for a cohort.

    Reads packs/dictionary_layout.yaml. Per-cohort `exclude_tables`
    replaces the global list; `extra_exclude_tables` adds to it.
    Falls back to the hard-coded PR-B list if the file is missing.

    `cohort` accepts a single key or a list of candidate keys (e.g.
    [slug, cohort_name, schema_name]). The first key with a
    `cohorts.<key>.customer` entry wins, so callers don't have to
    know which name the YAML was authored against. The CLI/filename
    slug is the documented preferred key.
    """
    layout = _load_dictionary_layout()
    customer = layout.get("customer") or {}
    global_excludes = customer.get("exclude_tables")
    if global_excludes is None:
        global_excludes = list(_CUSTOMER_TABLE_EXCLUDES_FALLBACK)

    candidates: list[str] = []
    if isinstance(cohort, str):
        candidates = [cohort]
    elif isinstance(cohort, list):
        candidates = [c for c in cohort if c]

    cohorts_cfg = layout.get("cohorts") or {}
    for key in candidates:
        cohort_cfg = (cohorts_cfg.get(key) or {}).get("customer") or {}
        if not cohort_cfg:
            continue
        if "exclude_tables" in cohort_cfg:
            return frozenset(cohort_cfg["exclude_tables"] or [])
        extra = cohort_cfg.get("extra_exclude_tables") or []
        return frozenset(list(global_excludes) + list(extra))

    return frozenset(global_excludes)


# --------------------------------------------------------------------------- #
# Sheet layouts — shared between write_xlsx and write_html so the two
# renderers can't drift on column order or accessor logic. Each entry
# is (display_label, accessor) for Tables / Columns / Variables, or
# (xlsx_label, html_label, accessor) for Summary (where the renderers
# genuinely diverge on labels).
#
# JSON is a full internal/debug dump (write_json) and intentionally
# bypasses these layouts. The stakeholder-facing audiences — customer
# and sales — skip JSON entirely so partner bundles never carry the
# full CohortModel. The internal audiences (technical, pharma) keep
# JSON for debugging.
#
# Each sheet has a per-audience dispatcher: summary_layout(audience),
# tables_layout(audience), columns_layout(audience), and
# variables_layout(audience). Current state:
#   - technical / pharma share the original layouts (full sheets).
#   - customer gets its own trimmed Summary / Tables / Columns and
#     a customer-tail for Variables that drops debug fields.
#   - sales ships Summary + Tables + Variables (no Columns).
#     Tables uses the customer-trimmed layout, Variables uses the
#     Tempus-style _SALES_VARIABLES_LAYOUT, and Summary is the
#     customer-trimmed key/value rows that the cover renderer
#     replaces at write_xlsx time with a styled cover sheet.
#     Columns is hidden via AUDIENCE_VISIBILITY['sales']['columns']
#     = False — partners read Variables for clinical content, and
#     an engineer who needs the column-level schema map can pull
#     the technical-audience output instead.
# Adding a new audience is a single dict entry in each *_BY_AUDIENCE
# map plus a visibility row.
# --------------------------------------------------------------------------- #

# Summary layout. Each entry is (xlsx_label, html_label, accessor):
#   - `xlsx_label is None` skips the row in the XLSX Summary sheet.
#   - `html_label is None` skips the entry in the HTML <dl> block.
# Two renderers genuinely diverge: XLSX writes lowercase metric/value
# rows; HTML writes title-case entries with a merged "Date coverage"
# line. One tuple per entry carries both, so per-audience layouts are
# a single edit.
_TECHNICAL_SUMMARY_LAYOUT: list[tuple[str | None, str | None, Any]] = [
    ("cohort",                   "Cohort",             lambda m: m.cohort),
    ("provider",                 "Provider",           lambda m: m.provider),
    ("disease",                  "Disease",            lambda m: m.disease),
    ("display_name",             "Display name",       lambda m: m.display_name),
    ("schema_name",              "Schema",             lambda m: m.schema_name),
    ("variant",                  "Variant",            lambda m: m.variant),
    ("patient_count",            "Patient count",      lambda m: m.summary.patient_count),
    ("table_count",              "Table count",        lambda m: m.summary.table_count),
    ("column_count",             "Column count",       lambda m: m.summary.column_count),
    # XLSX exposes min/max/years as three rows; HTML collapses them
    # into a single human-readable string below.
    ("min_date",                 None,                 lambda m: m.summary.date_coverage.min_date),
    ("max_date",                 None,                 lambda m: m.summary.date_coverage.max_date),
    ("years_of_data",            None,                 lambda m: m.summary.date_coverage.years_of_data),
    (None,                       "Date coverage",
        lambda m: (
            f"{m.summary.date_coverage.min_date} → {m.summary.date_coverage.max_date}"
            f" ({m.summary.date_coverage.years_of_data} years)"
            if m.summary.date_coverage.min_date else "—"
        )),
    ("status",                   "Status",             lambda m: m.status),
    ("generated_at",             "Generated at",       lambda m: m.generated_at),
    ("git_sha",                  "Git SHA",            lambda m: m.git_sha),
    ("introspect_version",       "Introspect version", lambda m: m.introspect_version),
    ("schema_snapshot_digest",   "Schema snapshot",    lambda m: m.schema_snapshot_digest),
]

# Customer Summary drops implementation/debug-flavored fields:
#   variant, column_count, status, git_sha, introspect_version,
#   schema_snapshot_digest. Date coverage is still emitted as 3 XLSX
#   rows + 1 merged HTML line so the renderer code path stays identical.
_CUSTOMER_SUMMARY_LAYOUT: list[tuple[str | None, str | None, Any]] = [
    ("cohort",         "Cohort",        lambda m: m.cohort),
    ("provider",       "Provider",      lambda m: m.provider),
    ("disease",        "Disease",       lambda m: m.disease),
    ("display_name",   "Display name",  lambda m: m.display_name),
    ("schema_name",    "Schema",        lambda m: m.schema_name),
    ("patient_count",  "Patient count", lambda m: m.summary.patient_count),
    ("table_count",    "Table count",   lambda m: m.summary.table_count),
    ("min_date",       None,            lambda m: m.summary.date_coverage.min_date),
    ("max_date",       None,            lambda m: m.summary.date_coverage.max_date),
    ("years_of_data",  None,            lambda m: m.summary.date_coverage.years_of_data),
    (None,             "Date coverage",
        lambda m: (
            f"{m.summary.date_coverage.min_date} → {m.summary.date_coverage.max_date}"
            f" ({m.summary.date_coverage.years_of_data} years)"
            if m.summary.date_coverage.min_date else "—"
        )),
    ("generated_at",   "Generated at",  lambda m: m.generated_at),
]

_SUMMARY_LAYOUT_BY_AUDIENCE: dict[str, list[tuple[str | None, str | None, Any]]] = {
    "technical": _TECHNICAL_SUMMARY_LAYOUT,
    # Sales targets a stakeholder-facing Tempus-style spec, so the
    # Summary cover sheet uses the trimmed customer layout —
    # variant / column_count / git_sha / introspect_version /
    # schema_snapshot_digest are internal-only and don't belong in
    # a sales workbook.
    "sales":     _CUSTOMER_SUMMARY_LAYOUT,
    "pharma":    _TECHNICAL_SUMMARY_LAYOUT,
    "customer":  _CUSTOMER_SUMMARY_LAYOUT,
}


def summary_layout(audience: str) -> list[tuple[str | None, str | None, Any]]:
    return _SUMMARY_LAYOUT_BY_AUDIENCE.get(audience, _TECHNICAL_SUMMARY_LAYOUT)


def summary_xlsx_rows(model: Any, audience: str = "technical") -> list[dict[str, Any]]:
    """Materialize summary layout entries that opt in to the XLSX sheet."""
    return [
        {"metric": xl_label, "value": fn(model)}
        for xl_label, _, fn in summary_layout(audience) if xl_label is not None
    ]


def summary_html_pairs(model: Any, audience: str = "technical") -> list[tuple[str, Any]]:
    """Materialize summary layout entries that opt in to the HTML <dl>."""
    return [
        (html_label, fn(model))
        for _, html_label, fn in summary_layout(audience) if html_label is not None
    ]


_TECHNICAL_TABLES_LAYOUT: list[tuple[str, Any]] = [
    ("Table",              lambda t: t.table_name),
    ("Category",           lambda t: t.category),
    ("Description",        lambda t: t.description or t.purpose),
    ("Inclusion Criteria", lambda t: t.inclusion_criteria),
    ("Data Source",        lambda t: t.data_source),
    ("Source Table",       lambda t: t.source_table),
    ("Rows",               lambda t: t.row_count),
    ("Columns",            lambda t: t.column_count),
    ("Patients",           lambda t: t.patient_count_in_table
                                       if t.patient_count_in_table is not None else "—"),
]

# Customer Tables drops Data Source / Source Table.
_CUSTOMER_TABLES_LAYOUT: list[tuple[str, Any]] = [
    ("Table",              lambda t: t.table_name),
    ("Category",           lambda t: t.category),
    ("Description",        lambda t: t.description or t.purpose),
    ("Inclusion Criteria", lambda t: t.inclusion_criteria),
    ("Rows",               lambda t: t.row_count),
    ("Columns",            lambda t: t.column_count),
    ("Patients",           lambda t: t.patient_count_in_table
                                       if t.patient_count_in_table is not None else "—"),
]

_TABLES_LAYOUT_BY_AUDIENCE: dict[str, list[tuple[str, Any]]] = {
    "technical": _TECHNICAL_TABLES_LAYOUT,
    # Sales is stakeholder-facing, so it uses the same trimmed
    # Tables layout as customer (drops Data Source / Source Table).
    "sales":     _CUSTOMER_TABLES_LAYOUT,
    "pharma":    _TECHNICAL_TABLES_LAYOUT,
    "customer":  _CUSTOMER_TABLES_LAYOUT,
}


def tables_layout(audience: str) -> list[tuple[str, Any]]:
    return _TABLES_LAYOUT_BY_AUDIENCE.get(audience, _TECHNICAL_TABLES_LAYOUT)


_TECHNICAL_COLUMNS_LAYOUT: list[tuple[str, Any]] = [
    ("Category",      lambda c: c.category),
    ("Table(s)",      lambda c: c.table),
    ("Column",        lambda c: c.column),
    ("Description",   lambda c: c.description),
    ("Field Type",    lambda c: c.data_type),
    ("Nullable",      lambda c: c.nullable),
    ("Example",       lambda c: c.example),
    ("Coding Schema", lambda c: c.coding_schema),
    ("Values",        lambda c: c.values),
    ("Distribution",  lambda c: c.distribution),
    ("Median (IQR)",  lambda c: c.median_iqr),
    ("Completeness",  lambda c: f"{c.completeness_pct:.1f}%"),
    ("% Patient",     lambda c: _fmt_pct(c.patient_pct)),
    ("Data Source",   lambda c: c.data_source),
    ("PII",           lambda c: "Yes" if c.pii else ""),
    ("Notes",         lambda c: c.notes),
]

# Customer Columns is a clean schema-description tab — only the four
# fields the reviewer named. Statistics live in Variables instead.
_CUSTOMER_COLUMNS_LAYOUT: list[tuple[str, Any]] = [
    ("Table(s)",    lambda c: c.table),
    ("Column",      lambda c: c.column),
    ("Description", lambda c: c.description),
    ("Field Type",  lambda c: c.data_type),
]

_COLUMNS_LAYOUT_BY_AUDIENCE: dict[str, list[tuple[str, Any]]] = {
    "technical": _TECHNICAL_COLUMNS_LAYOUT,
    # Sales reuses the customer-trimmed Columns layout for the same
    # stakeholder reasoning — the schema map needs Table / Column /
    # Description / Field Type only, not full distribution stats.
    "sales":     _CUSTOMER_COLUMNS_LAYOUT,
    "pharma":    _TECHNICAL_COLUMNS_LAYOUT,
    "customer":  _CUSTOMER_COLUMNS_LAYOUT,
}


def columns_layout(audience: str) -> list[tuple[str, Any]]:
    return _COLUMNS_LAYOUT_BY_AUDIENCE.get(audience, _TECHNICAL_COLUMNS_LAYOUT)


# Variables layout is split into head / criteria / tail so the
# Criteria column slots in at the right position. Customer keeps both
# `Inclusion Criteria` (prose) and `Criteria` (configured matcher).
_VARIABLES_LAYOUT_HEAD: list[tuple[str, Any]] = [
    ("Category",           lambda v: v.category),
    ("Variable",           lambda v: v.variable),
    ("Description",        lambda v: v.description),
    ("Inclusion Criteria", lambda v: v.inclusion_criteria),
    ("Table(s)",           lambda v: v.table),
    ("Column(s)",          lambda v: v.column),
]

_VARIABLES_LAYOUT_CRITERIA: tuple[str, Any] = ("Criteria", lambda v: v.criteria)

_TECHNICAL_VARIABLES_TAIL: list[tuple[str, Any]] = [
    ("Field Type",    lambda v: v.field_type),
    ("Example",       lambda v: v.example),
    ("Coding Schema", lambda v: v.coding_schema),
    ("Values",        lambda v: v.values),
    ("Distribution",  lambda v: v.distribution),
    ("Median (IQR)",  lambda v: v.median_iqr),
    ("Completeness",  lambda v: _fmt_pct(v.completeness_pct)),
    ("Implemented",   lambda v: v.implemented),
    # Renamed from "% Patient" — same field (patient_pct), clearer
    # label that matches what the metric actually measures: the
    # fraction of cohort patients with at least one non-null row
    # for this variable. Stakeholder audiences (pharma, sales,
    # customer) drop the row-level Completeness column and use
    # only this metric; technical keeps both for internal QA.
    ("% Patients With Value", lambda v: _fmt_pct(v.patient_pct)),
    ("Data Source",   lambda v: v.data_source),
    ("Notes",         lambda v: v.notes),
]


# Shared accessor for the stakeholder "Observed Values" cell.
# Reads the structured top_value_labels list (newline-separated,
# no counts) so OMOP labels with internal commas render verbatim.
def _observed_values_cell(v: Any) -> str:
    if v.top_value_labels:
        return "\n".join(v.top_value_labels)
    return ""


# Pharma Variables tail: methodology-rich view for scientific /
# evidence reviewers. Carries the full methodology stack (Coding
# Schema, Distribution, Median (IQR), Implemented, Data Source)
# AND the strict match Criteria — pharma scientists evaluate
# definitions and want the matcher visible. Drops only the
# row-level `Completeness` column (single coverage metric is
# `% Patients With Value`, sourced from patient_pct, consistent
# with all stakeholder audiences). Renames `Values` to
# `Observed Values` for label consistency with the
# top_value_labels-backed cell.
_PHARMA_VARIABLES_TAIL: list[tuple[str, Any]] = [
    ("Field Type",      lambda v: v.field_type),
    ("Example",         lambda v: v.example),
    ("Coding Schema",   lambda v: v.coding_schema),
    ("Observed Values", _observed_values_cell),
    ("Distribution",    lambda v: v.distribution),
    ("Median (IQR)",    lambda v: v.median_iqr),
    ("Implemented",     lambda v: v.implemented),
    ("% Patients With Value", lambda v: _fmt_pct(v.patient_pct)),
    ("Data Source",     lambda v: v.data_source),
    ("Notes",           lambda v: v.notes),
]


# Customer Variables: trimmed buyer-evaluation view. Drops the
# methodology fields that pharma keeps (Coding Schema, Distribution,
# Median (IQR), Implemented, Data Source) so the customer artifact
# stays plain-language and definition-focused. Keeps Criteria
# (added by variables_layout()) for transparency about how each
# variable is matched. Single coverage metric is
# `% Patients With Value`, sourced from patient_pct.
#
# `Observed Values` reads the structured top_value_labels list so
# OMOP labels with internal commas render verbatim (e.g.
# "Cancer, malignant" stays one cell entry, not two).
_CUSTOMER_VARIABLES_TAIL: list[tuple[str, Any]] = [
    ("Field Type",      lambda v: v.field_type),
    ("Example",         lambda v: v.example),
    ("Observed Values", _observed_values_cell),
    ("% Patients With Value", lambda v: _fmt_pct(v.patient_pct)),
    ("Notes",           lambda v: v.notes),
]


# Sales layout: stand-alone Tempus-style spec sheet. Reviewer's
# reference workbook (CH-Tempus Variables) uses these 7 columns;
# Completeness is added per follow-up so the sales reader can
# also see population coverage at a glance.
#
# `Value Sets` is purely data-driven: the observed top-N value
# labels for the variable's column, newline-separated, no counts /
# percentages. Reads `v.top_value_labels` — the structured list
# captured at build time alongside the comma-joined `values` cell —
# so a label that itself contains a comma (e.g. an OMOP concept
# name like "Cancer, malignant") renders verbatim instead of
# fragmenting. Empty for free-text columns and dry-run rows where
# no top-N query ran.
#
# A data dictionary that claims a value the cohort doesn't carry
# is wrong — so there's no curation override; if it's not in the
# data, it doesn't appear.
# `Proposal` comes from the curated YAML field. Type maps directly
# to extraction_type. The Observed Values cell uses the shared
# _observed_values_cell helper defined above (also used by the
# customer Variables tail).


_SALES_VARIABLES_LAYOUT: list[tuple[str, Any]] = [
    ("Category",    lambda v: v.category),
    ("Variable",    lambda v: v.variable),
    ("Description", lambda v: v.description),
    ("Observed Values", _observed_values_cell),
    ("Notes",       lambda v: v.notes),
    ("Type",        lambda v: v.extraction_type),
    ("Proposal",    lambda v: v.proposal),
    ("% Patients With Value", lambda v: _fmt_pct(v.patient_pct)),
]


def variables_layout(audience: str) -> list[tuple[str, Any]]:
    """Variables sheet layout for the given audience.

    Audience rules:
      - technical: head + Criteria + technical tail. Full audit
                   view; carries both Completeness (row-level)
                   and % Patients With Value side by side.
      - customer:  head + Criteria + customer tail. Plain-language
                   buyer view — definitions + Observed Values +
                   coverage; methodology fields (Coding Schema,
                   Distribution, Median (IQR), Implemented, Data
                   Source) intentionally omitted.
      - pharma:    head + Criteria + pharma tail. Methodology-rich
                   evidence view — keeps Coding Schema,
                   Distribution, Median (IQR), Implemented, Data
                   Source. Drops only row-level Completeness.
                   Pharma scientists evaluate definitions, so the
                   strict match Criteria IS shown.
      - sales:     standalone Tempus-style spec sheet (Observed
                   Values + % Patients With Value), no shared head
                   and no Criteria column.
    """
    if audience == "sales":
        return list(_SALES_VARIABLES_LAYOUT)
    layout = list(_VARIABLES_LAYOUT_HEAD)
    if audience in ("technical", "customer", "pharma"):
        layout.append(_VARIABLES_LAYOUT_CRITERIA)
    if audience == "customer":
        layout.extend(_CUSTOMER_VARIABLES_TAIL)
    elif audience == "pharma":
        layout.extend(_PHARMA_VARIABLES_TAIL)
    else:
        layout.extend(_TECHNICAL_VARIABLES_TAIL)
    return layout


def _rows_from_layout(layout: list[tuple[str, Any]], items: list[Any]) -> list[list[Any]]:
    """Materialize each item into a list of cell values matching layout order."""
    return [[fn(item) for _, fn in layout] for item in items]


def _df_from_layout(pd, layout: list[tuple[str, Any]], items: list[Any]):
    """Build a DataFrame whose columns and order come from `layout`."""
    headers = [label for label, _ in layout]
    return pd.DataFrame(_rows_from_layout(layout, items), columns=headers)


def section_visible(audience: str, section: str) -> bool:
    return AUDIENCE_VISIBILITY.get(audience, AUDIENCE_VISIBILITY["technical"])[section]


def filter_for_audience(
    model: CohortModel, audience: str, cohort_slug: str | None = None,
) -> CohortModel:
    if audience == "technical":
        return model
    # Drop PII rows from BOTH columns and variables for sales / pharma /
    # customer. Variables resolver tags `pii: true` whenever (table,
    # column) hits the PII pack, so the same predicate applies to both.
    filtered_columns = [c for c in model.columns if not c.pii]
    filtered_variables = [v for v in model.variables if not v.pii]
    filtered_tables = list(model.tables)

    # Customer audience also strips internal/scaffolding tables. Affects
    # all three lists so a hidden table doesn't leave dangling column or
    # variable rows pointing at a table the reader can't see. The
    # exclude list is per-cohort-aware and sourced from
    # packs/dictionary_layout.yaml. We pass the CLI/filename slug as
    # well as model.cohort (which is the cohort_name like
    # `balboa_ckd_cohort`) and schema_name as fallback keys, so the
    # YAML can be authored against whichever name is most natural.
    if audience in ("customer", "sales"):
        candidate_keys = [k for k in (
            cohort_slug, model.cohort, model.schema_name,
        ) if k]
        excluded = customer_table_excludes(candidate_keys)
        filtered_tables   = [t for t in filtered_tables   if t.table_name not in excluded]
        filtered_columns  = [c for c in filtered_columns  if c.table      not in excluded]
        filtered_variables = [v for v in filtered_variables if v.table    not in excluded]

    # Stakeholder audiences (customer, sales) drop variables that
    # don't have any data in the cohort. `implemented == "Yes"` is
    # the same signal the technical sheet already reports — rows
    # with implemented="No" would otherwise render as "0%" /
    # "Implemented: No" and add noise to the partner artifact.
    # Internal audiences (technical, pharma) keep them so QA can
    # see what's missing.
    #
    # Dry-run models have implemented="No" for every row (no DB
    # connection is open, so the resolver can't tell what's
    # populated). Detect that via `patient_count is None` — set
    # only by the dry-run branch in build_model — and skip the
    # implemented filter so offline previews still show every row.
    is_dry_run = model.summary.patient_count is None
    if audience in ("customer", "sales") and not is_dry_run:
        filtered_variables = [
            v for v in filtered_variables if v.implemented == "Yes"
        ]

    visibility = AUDIENCE_VISIBILITY[audience]
    return dataclasses.replace(
        model,
        tables=filtered_tables if visibility["tables"] else [],
        columns=filtered_columns if visibility["columns"] else [],
        variables=filtered_variables,
    )


# --------------------------------------------------------------------------- #
# Renderers
# --------------------------------------------------------------------------- #


def _fmt_pct(value: float | None) -> str:
    if value is None:
        return "—"
    return f"{value:.1f}%"


def write_xlsx(model: CohortModel, out_path: Path,
               audience: str = "technical") -> None:
    try:
        import pandas as pd
    except ImportError as exc:
        raise SystemExit("pandas is not installed: pip install pandas openpyxl") from exc

    summary_df = pd.DataFrame(
        summary_xlsx_rows(model, audience), columns=["metric", "value"]
    )

    tables_df = _df_from_layout(pd, tables_layout(audience), model.tables)
    columns_df = _df_from_layout(pd, columns_layout(audience), model.columns)
    variables_df = _df_from_layout(
        pd, variables_layout(audience), model.variables
    )

    # Stakeholder audiences (customer, sales) get a styled cover sheet
    # with a title block, description paragraph, hero stats, date
    # coverage, and a coverage-by-category rollup. Internal audiences
    # (technical, pharma) keep the bare key/value Summary that downstream
    # tools and tests already read.
    stakeholder_cover = audience in ("customer", "sales")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        if stakeholder_cover:
            # Create the Summary sheet directly via openpyxl rather
            # than letting pandas dump key/value rows into it.
            ws = writer.book.create_sheet("Summary", 0)
            _render_stakeholder_cover(ws, model)
        else:
            summary_df.to_excel(
                writer, sheet_name="Summary", index=False, header=True,
            )
        if section_visible(audience, "tables"):
            tables_df.to_excel(writer, sheet_name="Tables", index=False)
        if section_visible(audience, "columns"):
            columns_df.to_excel(writer, sheet_name="Columns", index=False)
        if section_visible(audience, "variables"):
            variables_df.to_excel(writer, sheet_name="Variables", index=False)
        _autosize_and_wrap(writer, skip_summary=stakeholder_cover)
        if section_visible(audience, "variables"):
            for ws in writer.book.worksheets:
                _apply_completeness_gradient(ws)
    print(f"Wrote {out_path}", file=sys.stderr)


# --------------------------------------------------------------------- #
# Stakeholder-cover helpers. Used by the customer / sales audiences to
# replace the bare key/value Summary sheet with a real cover page —
# big header, cohort description, hero stats, coverage rollup. Cohort-
# agnostic: every helper reads only from the canonical CohortModel.
# --------------------------------------------------------------------- #


def _coverage_rollup(model: CohortModel) -> list[dict[str, Any]]:
    """Per-category implementation + completeness summary.

    Returns one dict per category seen in `model.variables`:
        {category, total, implemented, implemented_pct, avg_completeness}
    Categories are alphabetised; `avg_completeness` is the mean of
    completeness_pct across the *implemented* rows (un-implemented
    rows have no signal to average). Empty model → [].
    """
    by_cat: dict[str, list[VariableRow]] = {}
    for v in model.variables:
        by_cat.setdefault((v.category or "Uncategorized").strip(), []).append(v)

    out: list[dict[str, Any]] = []
    for cat in sorted(by_cat.keys(), key=str.lower):
        rows = by_cat[cat]
        implemented = [r for r in rows if r.implemented == "Yes"]
        completes = [r.completeness_pct for r in implemented
                     if r.completeness_pct is not None]
        avg = (sum(completes) / len(completes)) if completes else None
        out.append({
            "category": cat,
            "total": len(rows),
            "implemented": len(implemented),
            "implemented_pct": (100.0 * len(implemented) / len(rows))
                                if rows else 0.0,
            "avg_completeness": avg,
        })
    return out


def _hero_stats(model: CohortModel) -> dict[str, Any]:
    """The four headline numbers that go on the cover.

    - patients:        live patient count for the cohort
    - years_of_data:   span between min/max event dates
    - variables:       total variable rows in the model
    - implemented_pct: fraction of variables with `implemented == "Yes"`

    All gracefully degrade to `None` / "—" when the model is dry-run
    (no DB) or empty, so the cover renders without crashing.
    """
    total_vars = len(model.variables)
    implemented = sum(1 for v in model.variables if v.implemented == "Yes")
    impl_pct = (100.0 * implemented / total_vars) if total_vars else 0.0
    return {
        "patients": model.summary.patient_count,
        "years_of_data": model.summary.date_coverage.years_of_data,
        "variables_total": total_vars,
        "variables_implemented": implemented,
        "variables_implemented_pct": impl_pct,
    }


def _render_stakeholder_cover(ws, model: CohortModel) -> None:
    """Write a styled Summary cover into an empty openpyxl worksheet.

    Layout (top to bottom):
      1. Title row       — display_name + " — " + disease pretty name
      2. Description     — wrapped paragraph from cohort YAML
      3. Hero stats      — N patients · X years · M variables · K% implemented
      4. Date coverage   — single readable line
      5. Freshness facts — data_cutoff_date / last_etl_run / sign_off
                            (each block only when populated)
      5b. Known limitations — bulleted list (only when populated)
      6. Coverage table  — per-category implemented + avg completeness

    Cohort-agnostic: only reads model fields. No sheet-name or
    cohort-slug specialisation. Empty freshness fields render
    nothing (no blank rows / dangling section headers).
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    title_font     = Font(name="Calibri", size=22, bold=True, color="1F3864")
    subtitle_font  = Font(name="Calibri", size=11, italic=True, color="595959")
    hero_label     = Font(name="Calibri", size=10, color="595959")
    hero_value     = Font(name="Calibri", size=18, bold=True, color="1F3864")
    section_font   = Font(name="Calibri", size=12, bold=True, color="1F3864")
    rollup_header  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    body_font      = Font(name="Calibri", size=10)
    rollup_fill    = PatternFill("solid", fgColor="1F3864")
    band_fill      = PatternFill("solid", fgColor="F2F2F2")
    thin           = Side(style="thin", color="BFBFBF")
    cell_border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top       = Alignment(wrap_text=True, vertical="top")
    center         = Alignment(horizontal="center", vertical="center")
    center_wrap    = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 1. Title — `<display_name> — <Disease> cohort`
    title = (model.display_name or model.cohort or "Cohort").strip()
    if model.disease:
        title = f"{title} — {model.disease.strip()} cohort"
    ws.cell(row=1, column=1, value=title).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    # Provider · schema · generated-at one-liner under the title.
    subtitle_bits = []
    if model.provider:
        subtitle_bits.append(f"Provider: {model.provider}")
    if model.schema_name:
        subtitle_bits.append(f"Schema: {model.schema_name}")
    if model.generated_at:
        subtitle_bits.append(f"Generated: {model.generated_at}")
    sub = ws.cell(row=2, column=1, value="  ·  ".join(subtitle_bits))
    sub.font = subtitle_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    # 2. Description paragraph (from the cohort YAML).
    row = 4
    if (model.description or "").strip():
        cell = ws.cell(row=row, column=1, value=model.description.strip())
        cell.font = body_font
        cell.alignment = wrap_top
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=6)
        ws.row_dimensions[row].height = 60
        row += 2

    # 3. Hero stats block — four cells, each two rows tall (label / value).
    hero = _hero_stats(model)

    def _fmt_int(n: Any) -> str:
        if isinstance(n, int):
            return f"{n:,}"
        return "—"

    def _fmt_years(n: Any) -> str:
        if isinstance(n, (int, float)):
            return f"{n:.1f} yrs"
        return "—"

    hero_cells = [
        ("Patients",            _fmt_int(hero["patients"])),
        ("Years of follow-up",  _fmt_years(hero["years_of_data"])),
        ("Variables",
         f"{hero['variables_total']:,}"
         if hero["variables_total"] else "—"),
        ("With data",
         f"{hero['variables_implemented_pct']:.0f}% "
         f"({hero['variables_implemented']:,}/"
         f"{hero['variables_total']:,})"
         if hero["variables_total"] else "—"),
    ]
    label_row = row
    value_row = row + 1
    for i, (label, value) in enumerate(hero_cells):
        col = 1 + i * 2  # 1, 3, 5, 7 — leave a gap col between blocks
        lc = ws.cell(row=label_row, column=col, value=label)
        lc.font = hero_label
        lc.alignment = center
        ws.merge_cells(start_row=label_row, start_column=col,
                       end_row=label_row, end_column=col + 1)
        vc = ws.cell(row=value_row, column=col, value=value)
        vc.font = hero_value
        vc.alignment = center
        ws.merge_cells(start_row=value_row, start_column=col,
                       end_row=value_row, end_column=col + 1)
    ws.row_dimensions[label_row].height = 16
    ws.row_dimensions[value_row].height = 28
    row = value_row + 2

    # 4. Date coverage line.
    dc = model.summary.date_coverage
    if dc.min_date and dc.max_date:
        coverage = (
            f"Date coverage: {dc.min_date} → {dc.max_date}"
            f"   ·   {dc.years_of_data:.1f} years"
        )
        cell = ws.cell(row=row, column=1, value=coverage)
        cell.font = body_font
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=6)
        row += 2

    # 5. Freshness / governance facts (Commit B). Each block
    #    renders only when the corresponding cohort YAML field is
    #    populated — empty values produce no row, so un-curated
    #    cohorts don't show blank section headers that would make
    #    the workbook look unfinished.
    fresh_lines: list[str] = []
    if (model.data_cutoff_date or "").strip():
        fresh_lines.append(f"Data current to: {model.data_cutoff_date}")
    if (model.last_etl_run or "").strip():
        fresh_lines.append(f"Last ETL run: {model.last_etl_run}")
    so = model.sign_off or {}
    if so.get("reviewer", "").strip():
        sign_bits = [f"Reviewed by: {so['reviewer'].strip()}"]
        if so.get("date", "").strip():
            sign_bits.append(so["date"].strip())
        if so.get("notes", "").strip():
            sign_bits.append(so["notes"].strip())
        fresh_lines.append("  ·  ".join(sign_bits))
    if fresh_lines:
        cell = ws.cell(row=row, column=1,
                       value="   ·   ".join(fresh_lines))
        cell.font = subtitle_font
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=6)
        row += 2

    if model.known_limitations:
        header = ws.cell(row=row, column=1, value="Known limitations")
        header.font = section_font
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=6)
        row += 1
        for caveat in model.known_limitations:
            cell = ws.cell(row=row, column=1, value=f"•  {caveat}")
            cell.font = body_font
            cell.alignment = wrap_top
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=6)
            row += 1
        row += 1   # gap before rollup

    # 6. Coverage rollup table.
    rollup = _coverage_rollup(model)
    if rollup:
        header = ws.cell(row=row, column=1, value="Coverage by category")
        header.font = section_font
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=6)
        row += 1

        cols = ["Category", "Variables", "Implemented", "Avg completeness"]
        for j, c in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=j, value=c)
            cell.font = rollup_header
            cell.fill = rollup_fill
            cell.alignment = center_wrap
            cell.border = cell_border
        row += 1

        for i, r in enumerate(rollup):
            band = band_fill if i % 2 else None
            avg = r["avg_completeness"]
            avg_str = f"{avg:.1f}%" if avg is not None else "—"
            impl_str = (
                f"{r['implemented']} / {r['total']} "
                f"({r['implemented_pct']:.0f}%)"
            )
            values = [r["category"], r["total"], impl_str, avg_str]
            for j, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=j, value=val)
                cell.font = body_font
                cell.border = cell_border
                if band is not None:
                    cell.fill = band
                if j == 1:
                    cell.alignment = wrap_top
                else:
                    cell.alignment = center
            row += 1

    # Set column widths for the cover. Standard for the rollup table
    # plus a generous header column.
    widths = {1: 32, 2: 14, 3: 22, 4: 20, 5: 18, 6: 18}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def _apply_completeness_gradient(ws) -> None:
    """3-color scale on the Completeness column of a Variables sheet.

    Red < 30 / yellow at 70 / green at 100. Reads the literal cell
    text — `_fmt_pct` writes values like '87.5%' — and sets a
    color-scale rule on the matching numeric range.
    """
    if ws.title != "Variables" or ws.max_row < 2:
        return
    # Find the Completeness column by its header text.
    target_col = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col_idx).value == "Completeness":
            target_col = col_idx
            break
    if target_col is None:
        return

    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    letter = get_column_letter(target_col)
    rng = f"{letter}2:{letter}{ws.max_row}"
    rule = ColorScaleRule(
        start_type="num", start_value=0,   start_color="F8696B",   # red
        mid_type="num",   mid_value=70,    mid_color="FFEB84",     # yellow
        end_type="num",   end_value=100,   end_color="63BE7B",     # green
    )
    ws.conditional_formatting.add(rng, rule)


# --------------------------------------------------------------------- #
# XLSX styling helpers. Pure presentation — no cell values change, so
# the tests that read row-1 header text and sheet names keep matching.
# --------------------------------------------------------------------- #


# Tuned widths per header name. Narrow for short flag-style columns
# (Implemented, % Patient, Data Source, PII, Nullable) so they don't
# eat horizontal space; wide for free-text columns (Description,
# Inclusion Criteria, Criteria, Distribution, Coding Schema, Notes)
# that carry the most reviewer content. Any header not listed falls
# back to the old auto-size-with-cap rule.
_COLUMN_WIDTH_OVERRIDES: dict[str, int] = {
    # Narrow
    "Implemented":      12,
    "% Patient":        11,
    "Completeness":     13,
    "PII":               6,
    "Nullable":          9,
    "Extraction Type":  15,
    "Data Source":      14,
    "Data Type":        14,
    "Field Type":       14,
    "Rows":             10,
    "Columns":          10,
    "Patients":         10,
    # Medium
    "Category":         18,
    "Table":            22,
    "Table(s)":         22,
    "Column":           24,
    "Column(s)":        24,
    "Variable":         30,
    "Median (IQR)":     26,
    "Example":          22,
    "Source Table":     26,
    # Wide
    "Values":           40,
    "Distribution":     50,
    "Description":      55,
    "Criteria":         55,
    "Inclusion Criteria": 55,
    "Coding Schema":    50,
    "Notes":            45,
    "Purpose":          55,
}

# Data sheets get freeze panes + auto-filter; Summary does not (it's
# key/value, not a filterable dataset).
_DATA_SHEET_NAMES = ("Tables", "Columns", "Variables")


def _style_xlsx_header_row(ws) -> None:
    """Bold white header text on a navy fill, centered, slightly taller
    row, thin border underneath. Keeps cell *values* untouched so every
    test that reads row-1 text still matches."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_fill = PatternFill(fill_type="solid",
                              start_color="1F3A5F", end_color="1F3A5F")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    header_border = Border(bottom=Side(style="thin", color="B5BCC6"))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = header_border
    ws.row_dimensions[1].height = 28


def _autosize_and_wrap(writer, skip_summary: bool = False) -> None:
    """Per-sheet polish:
      - Tuned column widths (per-header overrides, else old auto-size).
      - Styled header row (bold white on navy).
      - Freeze top row + auto-filter on data sheets (Tables / Columns /
        Variables). Summary is key / value and doesn't need either.
      - Word-wrap on every body cell.

    `skip_summary=True` leaves the Summary sheet untouched — used
    when the stakeholder cover renderer has already styled it
    explicitly (column widths, fonts, merged cells) and the
    auto-size pass would clobber that."""
    from openpyxl.styles import Alignment
    body_align = Alignment(wrap_text=True, vertical="top")

    for ws in writer.book.worksheets:
        if skip_summary and ws.title == "Summary":
            continue
        # Column widths + wrap on body cells.
        for col_idx, col in enumerate(ws.columns, start=1):
            header_value = ws.cell(row=1, column=col_idx).value
            override = _COLUMN_WIDTH_OVERRIDES.get(
                str(header_value) if header_value is not None else ""
            )
            max_len = 0
            for cell in col:
                if cell.row > 1:
                    cell.alignment = body_align
                if cell.value is None:
                    continue
                v = str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            letter = ws.cell(row=1, column=col_idx).column_letter
            if override is not None:
                ws.column_dimensions[letter].width = override
            else:
                ws.column_dimensions[letter].width = min(
                    max(12, max_len + 2), 60
                )

        # Data sheets get the full treatment: styled navy/white header,
        # frozen top row, and an auto-filter. Summary is a key/value
        # sheet and is left plain on purpose — a filterable / styled
        # Summary header would overclaim that the sheet is a dataset.
        if ws.title in _DATA_SHEET_NAMES:
            _style_xlsx_header_row(ws)
            if ws.max_row >= 2:
                ws.freeze_panes = "A2"
                last_col_letter = ws.cell(
                    row=1, column=ws.max_column
                ).column_letter
                ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"


def write_html(model: CohortModel, out_path: Path,
               audience: str = "technical") -> None:
    esc = lambda s: (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def _table(rows: list[list[str]], headers: list[str]) -> str:
        th = "".join(f"<th>{esc(h)}</th>" for h in headers)
        trs = "".join(
            "<tr>" + "".join(f"<td>{esc(str(c))}</td>" for c in r) + "</tr>"
            for r in rows
        )
        return f'<table class="dd"><thead><tr>{th}</tr></thead><tbody>{trs}</tbody></table>'

    summary_html = "".join(
        f"<dt>{esc(k)}</dt><dd>{esc(str(v))}</dd>"
        for k, v in summary_html_pairs(model, audience)
    )

    t_layout = tables_layout(audience)
    tables_headers = [label for label, _ in t_layout]
    tables_rows = _rows_from_layout(t_layout, model.tables)
    # HTML formats `Rows` with a thousands separator; XLSX leaves it
    # numeric. Apply that one formatter difference here so the shared
    # layout stays renderer-agnostic. Customer Tables also has a `Rows`
    # column, so the lookup is unconditional (skipped if the layout
    # ever drops it).
    if "Rows" in tables_headers:
        _rows_idx = tables_headers.index("Rows")
        for row in tables_rows:
            if isinstance(row[_rows_idx], int):
                row[_rows_idx] = f"{row[_rows_idx]:,}"

    c_layout = columns_layout(audience)
    columns_headers = [label for label, _ in c_layout]
    columns_rows = _rows_from_layout(c_layout, model.columns)

    var_layout = variables_layout(audience)
    var_headers = [label for label, _ in var_layout]
    variables_rows = _rows_from_layout(var_layout, model.variables)

    sections: list[str] = [
        f'<h2>Summary</h2><dl class="summary">{summary_html}</dl>',
    ]
    if section_visible(audience, "tables"):
        sections.append("<h2>Tables</h2>" + _table(tables_rows, tables_headers))
    if section_visible(audience, "columns"):
        sections.append("<h2>Columns</h2>" + _table(columns_rows, columns_headers))
    if section_visible(audience, "variables"):
        sections.append("<h2>Variables</h2>" + _table(variables_rows, var_headers))

    body = "\n".join(sections)
    page = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<title>Data Dictionary — {esc(model.display_name)}</title>
<style>
 /* Tier 1 visual polish. Pure CSS — no DOM changes, so every test
    asserting literal <h2>Section</h2> / <th>Column</th> substrings
    still passes. */
 :root {{
   --fg:         #1f2937;
   --fg-muted:   #64748b;
   --fg-subtle:  #475569;
   --bg:         #ffffff;
   --bg-soft:    #f8fafc;
   --bg-zebra:   #f1f5f9;
   --border:     #e2e8f0;
   --border-dk:  #cbd5e1;
   --accent:     #1f3a5f;
   --accent-fg:  #ffffff;
 }}
 body {{
   font-family: system-ui, -apple-system, "Segoe UI", Roboto, Inter,
                "Helvetica Neue", Arial, sans-serif;
   margin: 32px auto; max-width: 1400px; padding: 0 24px;
   color: var(--fg); background: var(--bg);
   font-size: 14px; line-height: 1.45;
   -webkit-font-smoothing: antialiased;
 }}
 h1 {{ font-size: 1.65rem; margin: 0 0 4px; font-weight: 600;
       letter-spacing: -0.01em; color: var(--fg); }}
 h2 {{ font-size: 1.15rem; margin: 40px 0 12px; color: var(--fg);
       border-bottom: 2px solid var(--border);
       padding-bottom: 6px; font-weight: 600;
       letter-spacing: -0.005em; }}
 dl.summary {{
   display: grid;
   grid-template-columns: max-content 1fr;
   gap: 6px 24px;
   font-size: 0.9rem;
   background: var(--bg-soft);
   border: 1px solid var(--border);
   border-radius: 8px;
   padding: 16px 20px;
   margin: 12px 0 8px;
 }}
 dl.summary dt {{ font-weight: 600; color: var(--fg-subtle); }}
 dl.summary dd {{ margin: 0; color: var(--fg); }}
 table.dd {{
   border-collapse: separate;
   border-spacing: 0;
   font-size: 0.86rem;
   width: 100%;
   margin-top: 4px;
   border: 1px solid var(--border);
   border-radius: 8px;
   overflow: hidden;
 }}
 table.dd th, table.dd td {{
   border-bottom: 1px solid var(--border);
   padding: 9px 12px;
   vertical-align: top;
   text-align: left;
   /* Preserve newlines in cell text so multi-line cells (e.g. the
      sales Value Sets cell) render as separate visible lines
      instead of collapsing into a single run-on string. */
   white-space: pre-line;
 }}
 table.dd tbody tr:last-child td {{ border-bottom: none; }}
 table.dd thead th {{
   background: var(--accent);
   color: var(--accent-fg);
   font-weight: 600;
   letter-spacing: 0.01em;
   position: sticky;
   top: 0;
   z-index: 1;
   box-shadow: 0 1px 0 0 var(--border-dk);
 }}
 table.dd tbody tr:nth-child(even) td {{ background: var(--bg-zebra); }}
 table.dd tbody tr:hover td {{ background: #e2e8f0; }}
 @media print {{
   body {{ margin: 12mm; padding: 0; max-width: none; font-size: 11px; }}
   h2 {{ page-break-before: always; }}
   /* First h2 (Summary) shouldn't trigger a pre-page break — keep it
      on page 1 with the title. And push the page break AFTER the
      Summary card so the first data sheet starts on a fresh page.
      Uses `dl.summary` directly because the DOM chain is
      h1 → div → h2 → dl.summary and an adjacent-sibling selector
      starting from h1 can't reach through the intervening h2. */
   h2:first-of-type {{ page-break-before: auto; }}
   dl.summary {{ page-break-after: always; }}
   table.dd thead th {{ position: static; }}
   table.dd thead {{ display: table-header-group; }}
   table.dd {{ border-radius: 0; }}
 }}
</style></head><body>
<h1>Data Dictionary — {esc(model.display_name)}</h1>
<div style="color:#666">{esc(model.description)}</div>

{body}

</body></html>
"""
    out_path.write_text(page, encoding="utf-8")
    print(f"Wrote {out_path}", file=sys.stderr)


def write_json(model: CohortModel, out_path: Path) -> None:
    """Write a full-dump JSON sidecar.

    JSON is an internal/debug artifact — stakeholders read XLSX or
    HTML. The stakeholder-facing audiences (customer and sales) skip
    JSON entirely (see main()) so partner bundles never carry the
    full CohortModel (which still includes criteria, coding_schema,
    implemented, patient_pct, data_source even when the rendered
    sheets intentionally hide them). The internal audiences
    (technical, pharma) keep JSON for debugging.
    """
    out_path.write_text(
        json.dumps(model.to_dict(), indent=2, default=str), encoding="utf-8",
    )
    print(f"Wrote {out_path}", file=sys.stderr)


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Build a four-page data dictionary for one cohort."
    )
    parser.add_argument("--cohort", required=True,
                        help="cohort slug, e.g. mtc_aat / mtc_alzheimers")
    parser.add_argument("--audience",
                        choices=("technical", "sales", "pharma", "customer"),
                        default="technical")
    parser.add_argument("--formats", nargs="+",
                        choices=("xlsx", "html", "json"),
                        default=["xlsx", "html", "json"])
    parser.add_argument("--out-dir", default=str(OUTPUT_DIR),
                        help="output directory (default: Output/)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Skip DB connection; emit pack-only skeleton. "
                             "Useful for validating packs offline.")
    args = parser.parse_args(argv)

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if args.dry_run:
        model = build_model(args.cohort, conn=None, dry_run=True)
    else:
        psycopg = _require_psycopg()

        class _NS:
            host = None; port = None; database = None
            user = None; password = None; sslmode = None
        conn_kwargs = build_conn_kwargs(_NS())
        with psycopg.connect(**conn_kwargs) as conn:
            conn.autocommit = True
            model = build_model(args.cohort, conn, dry_run=False)

    model = filter_for_audience(model, args.audience, cohort_slug=args.cohort)

    stem = f"{model.schema_name}_dictionary"
    if args.audience != "technical":
        stem += f"_{args.audience}"
    if "xlsx" in args.formats:
        write_xlsx(model, out_dir / f"{stem}.xlsx", audience=args.audience)
    if "html" in args.formats:
        write_html(model, out_dir / f"{stem}.html", audience=args.audience)
    # JSON is an internal/debug artifact (full CohortModel dump,
    # carries criteria / coding_schema / implemented / patient_pct /
    # data_source even for trimmed audiences). Stakeholder-facing
    # audiences — customer and sales — read XLSX or HTML; suppress
    # JSON for them rather than maintain a parallel projection of
    # every layout.
    if "json" in args.formats and args.audience not in ("customer", "sales"):
        write_json(model, out_dir / f"{stem}.json")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

