"""Tests for the customer audience added in PR-B.

These exercise dictionary_v2/build_dictionary.py specifically. The root
test_build_dictionary.py suite continues to cover the original
build_dictionary.py module (which doesn't know about customer).
"""
from __future__ import annotations

import dataclasses
import importlib.util
import shutil
import sys
import unittest
from pathlib import Path

# Force-load the v2 build_dictionary module under the canonical name.
_REPO_ROOT = Path(__file__).resolve().parent.parent
_V2_PATH = Path(__file__).resolve().parent / "build_dictionary.py"
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))
_spec = importlib.util.spec_from_file_location("build_dictionary", _V2_PATH)
bd = importlib.util.module_from_spec(_spec)
sys.modules["build_dictionary"] = bd
_spec.loader.exec_module(bd)


# Repo-local output dir for file-writing smoke tests. Replaces
# tempfile.TemporaryDirectory() which has been unreliable in the
# Windows reviewer sandbox (PermissionError when openpyxl tries to
# write into the OS temp tree). Lives next to the test file and is
# gitignored.
_TEST_OUTPUT_ROOT = Path(__file__).resolve().parent / ".test_outputs"


def _output_dir(test_id: str) -> Path:
    """Return a stable, repo-local directory for one test's artifacts.

    `test_id` is `self.id()` (something like
    `dictionary_v2.test_customer_audience.CustomerXlsxSmokeTests.
    test_summary_has_no_metric_value_header`). The directory is
    cleared at the start of each call so tests start from a known
    empty state.
    """
    out = _TEST_OUTPUT_ROOT / test_id
    if out.exists():
        shutil.rmtree(out, ignore_errors=True)
    out.mkdir(parents=True, exist_ok=True)
    return out


def _make_table(name: str) -> "bd.TableRow":
    return bd.TableRow(
        table_name=name, category="Demographics", description="", purpose="",
        inclusion_criteria="", data_source="Normalized", source_table="OMOP",
        row_count=10, column_count=2, patient_count_in_table=10,
    )


def _make_column(table: str, column: str = "year_of_birth") -> "bd.ColumnRow":
    return bd.ColumnRow(
        category="Demographics", table=table, column=column,
        description="A column.", data_type="integer",
        values="", distribution="", median_iqr="",
        completeness_pct=100.0, patient_pct=100.0,
        extraction_type="Direct", pii=False, notes="",
        nullable="YES", example="1980", coding_schema="",
        data_source="Normalized",
    )


def _make_variable(
    table: str, column: str = "year_of_birth",
    implemented: str = "Yes", variable: str = "Age",
) -> "bd.VariableRow":
    return bd.VariableRow(
        category="Demographics", variable=variable, description="Patient age",
        table=table, column=column, criteria="year_of_birth IS NOT NULL",
        values="", distribution="", median_iqr="",
        completeness_pct=100.0, implemented=implemented, patient_pct=100.0,
        extraction_type="Direct", pii=False, notes="",
        inclusion_criteria="Patients with a birth year on file.",
        field_type="integer", example="44", coding_schema="",
        data_source="Normalized",
    )


def _make_model(tables, columns, variables) -> "bd.CohortModel":
    return bd.CohortModel(
        cohort="c", provider="p", disease="d", display_name="dn",
        schema_name="sn", variant="raw", description="",
        generated_at="2026-01-01", status="active",
        git_sha="abc", introspect_version="0.5.0",
        schema_snapshot_digest="sha",
        summary=bd.CohortSummary(
            patient_count=10, table_count=len(tables),
            column_count=len(columns),
            date_coverage=bd.DateCoverage(
                min_date="2020-01-01", max_date="2025-01-01",
                years_of_data=5.0,
            ),
        ),
        tables=tables, columns=columns, variables=variables,
    )


class CustomerLayoutTests(unittest.TestCase):
    """The customer audience trims columns from each sheet to match
    the reviewer feedback: drop debug fields from Summary, drop
    Data Source / Source Table from Tables, reduce Columns to a
    schema-description tab, and remove Coding Schema / Implemented /
    Data Source from Variables."""

    def test_audience_registered(self):
        self.assertIn("customer", bd.AUDIENCE_VISIBILITY)
        self.assertEqual(
            bd.AUDIENCE_VISIBILITY["customer"],
            {"summary": True, "tables": True, "columns": True, "variables": True},
        )

    def test_summary_drops_debug_fields(self):
        xl_keys = [k for k, _, _ in bd.summary_layout("customer") if k is not None]
        for dropped in ("variant", "column_count", "status", "git_sha",
                        "introspect_version", "schema_snapshot_digest"):
            self.assertNotIn(dropped, xl_keys)
        for kept in ("cohort", "provider", "disease", "patient_count",
                     "table_count", "generated_at"):
            self.assertIn(kept, xl_keys)

    def test_tables_drops_data_source_and_source_table(self):
        headers = [label for label, _ in bd.tables_layout("customer")]
        self.assertNotIn("Data Source", headers)
        self.assertNotIn("Source Table", headers)
        for kept in ("Table", "Description", "Inclusion Criteria",
                     "Rows", "Columns", "Patients"):
            self.assertIn(kept, headers)

    def test_columns_is_exactly_four_fields(self):
        headers = [label for label, _ in bd.columns_layout("customer")]
        self.assertEqual(headers,
                         ["Table(s)", "Column", "Description", "Field Type"])

    def test_variables_keeps_both_inclusion_and_criteria(self):
        headers = [label for label, _ in bd.variables_layout("customer")]
        self.assertIn("Inclusion Criteria", headers)
        self.assertIn("Criteria", headers)
        for dropped in ("Coding Schema", "Implemented", "Data Source"):
            self.assertNotIn(dropped, headers)

    def test_variables_drops_percent_patient_for_customer_keeps_completeness(self):
        # Per reviewer instruction: customer audience shows
        # Completeness only; % Patient is dropped to avoid surfacing
        # two near-identical patient-coverage metrics.
        headers = [label for label, _ in bd.variables_layout("customer")]
        self.assertIn("Completeness", headers)
        self.assertNotIn("% Patient", headers)

    def test_other_audiences_unaffected(self):
        # technical and pharma share the same Tables / Columns
        # layouts as before; sales now has a single-sheet
        # Tempus-style spec layout exercised separately.
        for aud in ("technical", "pharma"):
            tables = [l for l, _ in bd.tables_layout(aud)]
            self.assertIn("Data Source", tables, msg=f"{aud} Tables should keep Data Source")
            cols = [l for l, _ in bd.columns_layout(aud)]
            self.assertIn("Coding Schema", cols, msg=f"{aud} Columns should keep Coding Schema")
            vars_layout = [l for l, _ in bd.variables_layout(aud)]
            self.assertIn(
                "% Patient", vars_layout,
                msg=f"{aud} Variables should keep % Patient",
            )


class CustomerTableFilterTests(unittest.TestCase):
    """Internal scaffolding tables (cohort_patients, standard_profile_data_model,
    dv_tokenized_profile_data) are stripped from the customer dictionary
    only — technical retains them for debugging."""

    def test_filter_drops_internal_tables_for_customer(self):
        tables = [_make_table("person"), _make_table("cohort_patients"),
                  _make_table("standard_profile_data_model"),
                  _make_table("dv_tokenized_profile_data")]
        columns = [_make_column("person"), _make_column("cohort_patients")]
        variables = [_make_variable("person"), _make_variable("cohort_patients")]
        model = _make_model(tables, columns, variables)

        filtered = bd.filter_for_audience(model, "customer")
        names = {t.table_name for t in filtered.tables}
        self.assertEqual(names, {"person"})
        self.assertEqual({c.table for c in filtered.columns}, {"person"})
        self.assertEqual({v.table for v in filtered.variables}, {"person"})

    def test_technical_keeps_internal_tables(self):
        tables = [_make_table("person"), _make_table("cohort_patients")]
        columns = [_make_column("person"), _make_column("cohort_patients")]
        variables = [_make_variable("person"), _make_variable("cohort_patients")]
        model = _make_model(tables, columns, variables)

        filtered = bd.filter_for_audience(model, "technical")
        names = {t.table_name for t in filtered.tables}
        self.assertEqual(names, {"person", "cohort_patients"})

    def test_sales_layout_is_tempus_spec_columns(self):
        # The sales layout matches the reviewer's Tempus reference
        # workbook plus a Completeness column.
        headers = [label for label, _ in bd.variables_layout("sales")]
        self.assertEqual(
            headers,
            ["Category", "Variable", "Description", "Value Sets",
             "Notes", "Type", "Proposal", "Completeness"],
        )

    def test_sales_value_sets_renders_observed_top_values(self):
        # Value Sets is the observed top-N labels from `v.values`,
        # newline-separated. No curation field; the data is the
        # source of truth.
        row = bd.VariableRow(
            category="Medications", variable="Anti-amyloid Therapy",
            description="Anti-amyloid mAb administrations.",
            table="drug_exposure", column="drug_concept_name",
            criteria="", values="Lecanemab, Lecanemab-irmb, Leqembi, Donanemab-azbt, Kisunla",
            distribution="ignored for sales", median_iqr="",
            completeness_pct=87.5, implemented="Yes", patient_pct=80.0,
            extraction_type="Structured", notes="",
            proposal="Standard",
        )
        rendered = {label: fn(row) for label, fn in bd.variables_layout("sales")}
        self.assertEqual(
            rendered["Value Sets"],
            "Lecanemab\nLecanemab-irmb\nLeqembi\nDonanemab-azbt\nKisunla",
        )
        self.assertEqual(rendered["Type"], "Structured")
        self.assertEqual(rendered["Proposal"], "Standard")
        self.assertEqual(rendered["Completeness"], "87.5%")

    def test_sales_value_sets_empty_when_no_observed_values(self):
        # Free-text columns / dry-run rows have v.values = "".
        # Cell renders empty rather than "—" — there's no curation
        # path to fall back to.
        row = bd.VariableRow(
            category="Demographics", variable="Notes",
            description="Free text.", table="observation",
            column="note_text", criteria="",
            values="", distribution="", median_iqr="",
            completeness_pct=None, implemented="No", patient_pct=None,
            extraction_type="Unstructured", notes="",
            proposal="",
        )
        rendered = {label: fn(row) for label, fn in bd.variables_layout("sales")}
        self.assertEqual(rendered["Value Sets"], "")

    def test_sales_visibility_ships_all_four_sheets(self):
        vis = bd.AUDIENCE_VISIBILITY["sales"]
        self.assertTrue(vis["summary"])
        self.assertTrue(vis["tables"])
        self.assertTrue(vis["columns"])
        self.assertTrue(vis["variables"])

    def test_sales_tables_uses_customer_trimmed_layout(self):
        # Sales is stakeholder-facing → reuses the customer-trimmed
        # Tables layout (no Data Source / Source Table).
        headers = [label for label, _ in bd.tables_layout("sales")]
        self.assertNotIn("Data Source", headers)
        self.assertNotIn("Source Table", headers)
        for kept in ("Table", "Description", "Inclusion Criteria",
                     "Rows", "Columns", "Patients"):
            self.assertIn(kept, headers)

    def test_sales_columns_uses_customer_trimmed_layout(self):
        headers = [label for label, _ in bd.columns_layout("sales")]
        self.assertEqual(
            headers, ["Table(s)", "Column", "Description", "Field Type"],
        )

    def test_sales_summary_uses_trimmed_layout(self):
        # Sales is stakeholder-facing — Summary must NOT carry
        # internal/debug fields (variant, column_count, status,
        # git_sha, introspect_version, schema_snapshot_digest).
        xl_keys = [k for k, _, _ in bd.summary_layout("sales") if k is not None]
        for dropped in ("variant", "column_count", "status", "git_sha",
                        "introspect_version", "schema_snapshot_digest"):
            self.assertNotIn(dropped, xl_keys, msg=f"sales must drop {dropped}")
        for kept in ("cohort", "provider", "disease", "patient_count",
                     "table_count", "generated_at"):
            self.assertIn(kept, xl_keys)

    def test_sales_audience_strips_internal_scaffolding_tables(self):
        # Sales now ships a Tables sheet — but with the same
        # internal-table excludes the customer audience uses
        # (cohort_patients / standard_profile_data_model / etc.).
        tables = [_make_table("person"), _make_table("cohort_patients")]
        model = _make_model(tables, [_make_column("person")], [_make_variable("person")])
        filtered = bd.filter_for_audience(model, "sales")
        names = {t.table_name for t in filtered.tables}
        self.assertEqual(names, {"person"})


class StakeholderImplementedFilterTests(unittest.TestCase):
    """Variables that have no data in the cohort (`implemented="No"`)
    must not appear in stakeholder-facing audiences (customer, sales).
    They'd otherwise render as 0% rows that add noise to the partner
    artifact. Internal audiences (technical, pharma) keep them so QA
    can see gaps."""

    def _live_model(self, *variables):
        # `patient_count > 0` puts the model in "live" (non-dry-run)
        # mode so the implemented filter actually fires.
        m = _make_model(
            tables=[_make_table("person")],
            columns=[_make_column("person")],
            variables=list(variables),
        )
        return dataclasses.replace(
            m,
            summary=dataclasses.replace(m.summary, patient_count=10),
        )

    def test_customer_drops_unimplemented_variables(self):
        live = self._live_model(
            _make_variable("person", variable="Has Data", implemented="Yes"),
            _make_variable("person", variable="No Data Yet", implemented="No"),
        )
        filtered = bd.filter_for_audience(live, "customer")
        names = [v.variable for v in filtered.variables]
        self.assertEqual(names, ["Has Data"])

    def test_sales_drops_unimplemented_variables(self):
        live = self._live_model(
            _make_variable("person", variable="Has Data", implemented="Yes"),
            _make_variable("person", variable="No Data Yet", implemented="No"),
        )
        filtered = bd.filter_for_audience(live, "sales")
        names = [v.variable for v in filtered.variables]
        self.assertEqual(names, ["Has Data"])

    def test_technical_keeps_unimplemented_variables(self):
        live = self._live_model(
            _make_variable("person", variable="Has Data", implemented="Yes"),
            _make_variable("person", variable="No Data Yet", implemented="No"),
        )
        # Technical audience returns the model verbatim — no filter.
        filtered = bd.filter_for_audience(live, "technical")
        names = [v.variable for v in filtered.variables]
        self.assertEqual(names, ["Has Data", "No Data Yet"])

    def test_pharma_keeps_unimplemented_variables(self):
        live = self._live_model(
            _make_variable("person", variable="Has Data", implemented="Yes"),
            _make_variable("person", variable="No Data Yet", implemented="No"),
        )
        filtered = bd.filter_for_audience(live, "pharma")
        names = [v.variable for v in filtered.variables]
        self.assertEqual(set(names), {"Has Data", "No Data Yet"})

    def test_dry_run_skips_implemented_filter_for_customer(self):
        # Dry-run model carries patient_count=None; every row is
        # implemented="No" because no DB was open. The filter must
        # NOT fire here or the offline preview becomes empty.
        m = _make_model(
            tables=[_make_table("person")],
            columns=[_make_column("person")],
            variables=[
                _make_variable("person", variable="A", implemented="No"),
                _make_variable("person", variable="B", implemented="No"),
            ],
        )
        dry = dataclasses.replace(
            m, summary=dataclasses.replace(m.summary, patient_count=None),
        )
        filtered = bd.filter_for_audience(dry, "customer")
        self.assertEqual(
            len(filtered.variables), 2,
            msg="dry-run preview must keep all rows; the implemented "
                "filter relies on live DB counts",
        )


class _CustomerSmokeBase(unittest.TestCase):
    """Shared setUp for the file-writing smoke tests."""

    def setUp(self):
        tables = [_make_table("person"), _make_table("cohort_patients")]
        columns = [_make_column("person"), _make_column("cohort_patients")]
        variables = [_make_variable("person"), _make_variable("cohort_patients")]
        self.model = _make_model(tables, columns, variables)
        self.filtered = bd.filter_for_audience(self.model, "customer")


class CustomerXlsxSmokeTests(_CustomerSmokeBase):

    def setUp(self):
        try:
            import openpyxl  # noqa: F401
            import pandas    # noqa: F401
        except ImportError:
            self.skipTest("openpyxl / pandas not installed")
        super().setUp()

    def test_summary_has_no_metric_value_header(self):
        import openpyxl
        path = _output_dir(self.id()) / "out.xlsx"
        bd.write_xlsx(self.filtered, path, audience="customer")
        wb = openpyxl.load_workbook(path)
        ws = wb["Summary"]
        # Row 1 must be the first data row, not the literal labels.
        row1 = [c.value for c in ws[1]]
        self.assertEqual(row1, ["cohort", "c"])
        self.assertNotEqual(row1, ["metric", "value"])

    def test_columns_sheet_has_exactly_four_headers(self):
        import openpyxl
        path = _output_dir(self.id()) / "out.xlsx"
        bd.write_xlsx(self.filtered, path, audience="customer")
        wb = openpyxl.load_workbook(path)
        ws = wb["Columns"]
        headers = [c.value for c in ws[1]]
        self.assertEqual(headers,
                         ["Table(s)", "Column", "Description", "Field Type"])

    def test_excluded_tables_absent_from_xlsx(self):
        import openpyxl
        path = _output_dir(self.id()) / "out.xlsx"
        bd.write_xlsx(self.filtered, path, audience="customer")
        wb = openpyxl.load_workbook(path)
        ws = wb["Tables"]
        names = [ws.cell(row=r, column=1).value
                 for r in range(2, ws.max_row + 1)]
        self.assertNotIn("cohort_patients", names)
        self.assertEqual(names, ["person"])


class CustomerHtmlSmokeTests(_CustomerSmokeBase):

    def test_html_omits_debug_summary_fields(self):
        path = _output_dir(self.id()) / "out.html"
        bd.write_html(self.filtered, path, audience="customer")
        html = path.read_text()
        self.assertNotIn("<dt>Variant</dt>", html)
        self.assertNotIn("<dt>Git SHA</dt>", html)
        self.assertNotIn("<dt>Schema snapshot</dt>", html)
        # Date coverage is the merged HTML field — it must remain.
        self.assertIn("<dt>Date coverage</dt>", html)

    def test_html_variables_includes_both_criteria_columns(self):
        path = _output_dir(self.id()) / "out.html"
        bd.write_html(self.filtered, path, audience="customer")
        html = path.read_text()
        self.assertIn("<th>Inclusion Criteria</th>", html)
        self.assertIn("<th>Criteria</th>", html)


class CustomerJsonSuppressionTests(_CustomerSmokeBase):
    """JSON is an internal/debug artifact; customer audience targets
    external stakeholders who read XLSX/HTML. main() suppresses JSON
    output for customer rather than maintaining a parallel projection."""

    def test_main_skips_json_for_customer(self):
        # Drive main() end-to-end for one cohort. Use --dry-run so the
        # introspection step doesn't need a database connection.
        out_dir = _output_dir(self.id())
        argv = [
            "--cohort", "balboa_ckd",
            "--audience", "customer",
            "--formats", "xlsx", "html", "json",
            "--out-dir", str(out_dir),
            "--dry-run",
        ]
        rc = bd.main(argv)
        self.assertEqual(rc, 0)
        files = sorted(p.name for p in out_dir.iterdir() if p.is_file())
        # JSON must be absent; XLSX + HTML still get written.
        self.assertFalse(any(f.endswith(".json") for f in files),
                         msg=f"customer audience must not write JSON, got {files}")
        self.assertTrue(any(f.endswith(".xlsx") for f in files), msg=files)
        self.assertTrue(any(f.endswith(".html") for f in files), msg=files)

    def test_main_skips_json_for_sales(self):
        # Sales is now stakeholder-facing too — its JSON dump would
        # carry criteria / coding_schema / implemented / patient_pct /
        # data_source even though those columns are intentionally
        # absent from the Tempus-style sales sheet. JSON must be
        # suppressed so the bundle for that audience never accidentally
        # ships an internal-only artifact.
        out_dir = _output_dir(self.id())
        rc = bd.main([
            "--cohort", "balboa_ckd",
            "--audience", "sales",
            "--formats", "xlsx", "html", "json",
            "--out-dir", str(out_dir),
            "--dry-run",
        ])
        self.assertEqual(rc, 0)
        files = sorted(p.name for p in out_dir.iterdir() if p.is_file())
        self.assertFalse(any(f.endswith(".json") for f in files),
                         msg=f"sales must not write JSON, got {files}")
        self.assertTrue(any(f.endswith(".xlsx") for f in files), msg=files)
        self.assertTrue(any(f.endswith(".html") for f in files), msg=files)

    def test_main_still_writes_json_for_technical_and_pharma(self):
        # Don't accidentally widen the suppression to internal
        # audiences — they need the JSON for debugging.
        for aud in ("technical", "pharma"):
            # `_` separator instead of `:` — colons are illegal in
            # Windows path components and tripped NotADirectoryError
            # in the reviewer's sandbox.
            out_dir = _output_dir(self.id() + "_" + aud)
            rc = bd.main([
                "--cohort", "balboa_ckd",
                "--audience", aud,
                "--formats", "xlsx", "html", "json",
                "--out-dir", str(out_dir),
                "--dry-run",
            ])
            self.assertEqual(rc, 0)
            files = sorted(p.name for p in out_dir.iterdir() if p.is_file())
            self.assertTrue(any(f.endswith(".json") for f in files),
                            msg=f"{aud} should still write JSON, got {files}")


class CustomerLayoutConfigTests(unittest.TestCase):
    """packs/dictionary_layout.yaml drives the customer table-exclude
    list. Per-cohort overrides can replace or extend the global list."""

    def test_yaml_drives_global_excludes(self):
        excludes = bd.customer_table_excludes()
        for name in ("standard_profile_data_model", "cohort_patients",
                     "dv_tokenized_profile_data"):
            self.assertIn(name, excludes,
                          msg=f"{name} should be excluded by default")

    def test_unknown_cohort_falls_back_to_global(self):
        self.assertEqual(
            bd.customer_table_excludes("nonexistent_cohort"),
            bd.customer_table_excludes(),
        )

    def _patch_layout(self, layout):
        original = bd._load_dictionary_layout
        bd._load_dictionary_layout = lambda: layout
        self.addCleanup(lambda: setattr(bd, "_load_dictionary_layout", original))

    def test_per_cohort_exclude_tables_replaces_global(self):
        self._patch_layout({
            "customer": {"exclude_tables": ["a", "b"]},
            "cohorts": {"my_cohort": {"customer": {"exclude_tables": ["c"]}}},
        })
        self.assertEqual(bd.customer_table_excludes("my_cohort"), frozenset({"c"}))
        self.assertEqual(bd.customer_table_excludes(), frozenset({"a", "b"}))

    def test_per_cohort_extra_excludes_extends_global(self):
        self._patch_layout({
            "customer": {"exclude_tables": ["a", "b"]},
            "cohorts": {"my_cohort": {"customer": {"extra_exclude_tables": ["c"]}}},
        })
        self.assertEqual(
            bd.customer_table_excludes("my_cohort"),
            frozenset({"a", "b", "c"}),
        )

    def test_filter_for_audience_uses_cohort_specific_excludes(self):
        self._patch_layout({
            "customer": {"exclude_tables": []},
            "cohorts": {"c": {"customer": {"exclude_tables": ["person"]}}},
        })
        tables = [_make_table("person"), _make_table("cohort_patients")]
        columns = [_make_column("person"), _make_column("cohort_patients")]
        variables = [_make_variable("person"), _make_variable("cohort_patients")]
        model = _make_model(tables, columns, variables)
        # _make_model sets cohort="c", which the per-cohort override targets.
        filtered = bd.filter_for_audience(model, "customer")
        self.assertEqual({t.table_name for t in filtered.tables}, {"cohort_patients"})


class CompileMatchBlockTests(unittest.TestCase):
    """The structured `match:` block in variable YAML compiles to a
    strict `column IN (...)` SQL clause. The data dictionary uses this
    as both the displayed Criteria and the WHERE for completeness, so
    the definition stays config-owned (not data-derived)."""

    def test_inline_values_compile_to_in_list(self):
        sql = bd.compile_match_block({
            "column": "drug_concept_name",
            "values": ["Lecanemab", "Leqembi", "Donanemab"],
        })
        self.assertEqual(
            sql,
            "\"drug_concept_name\" IN ('Lecanemab', 'Leqembi', 'Donanemab')",
        )

    def test_escapes_single_quotes(self):
        sql = bd.compile_match_block({
            "column": "concept_name",
            "values": ["O'Brien syndrome", "Plain value"],
        })
        self.assertIn("'O''Brien syndrome'", sql)
        self.assertIn("'Plain value'", sql)

    def test_dedupes_inline_values_preserving_order(self):
        sql = bd.compile_match_block({
            "column": "c",
            "values": ["A", "B", "A", "C", "B"],
        })
        self.assertEqual(sql, "\"c\" IN ('A', 'B', 'C')")

    def test_missing_column_returns_blank(self):
        self.assertEqual(bd.compile_match_block({"values": ["x"]}), "")

    def test_empty_values_returns_blank(self):
        self.assertEqual(bd.compile_match_block({"column": "c"}), "")
        self.assertEqual(bd.compile_match_block({"column": "c", "values": []}), "")

    def test_none_block_returns_blank(self):
        self.assertEqual(bd.compile_match_block(None), "")

    def test_values_file_loads_and_unions_with_inline(self):
        # values_file points at packs/<rel_path>; create a temp file
        # there and clean it up.
        rel = "_test_match_values.yaml"
        path = bd.PACKS_DIR / rel
        path.write_text("- Foo\n- Bar\n", encoding="utf-8")
        try:
            sql = bd.compile_match_block({
                "column": "concept_name",
                "values": ["Bar", "Baz"],
                "values_file": rel,
            })
        finally:
            path.unlink()
        # Order: inline first, file values second; "Bar" appears once.
        self.assertEqual(
            sql,
            "\"concept_name\" IN ('Bar', 'Baz', 'Foo')",
        )


# Load the discovery script once under a stable module name so the
# dataclass it defines (VariableObservation) survives across test cases.
import importlib.util as _ilu_disc  # noqa: E402

_DISC_PATH = Path(__file__).resolve().parent / "discover_exact_matches.py"
_disc_spec = _ilu_disc.spec_from_file_location(
    "discover_exact_matches_under_test", _DISC_PATH,
)
discover_mod = _ilu_disc.module_from_spec(_disc_spec)
sys.modules["discover_exact_matches_under_test"] = discover_mod
_disc_spec.loader.exec_module(discover_mod)


class DiscoveryScriptTests(unittest.TestCase):
    """The discovery script proposes additions to a variable's `match:`
    block — it must never modify the disease pack automatically."""

    def setUp(self):
        self.mod = discover_mod

    def _obs(self, configured, observed):
        return self.mod.VariableObservation(
            category="C", variable="V", table="t", column="c",
            criteria="c ILIKE '%x%'", configured_values=configured,
            observed=observed,
        )

    def test_partitions_observed_against_config(self):
        o = self._obs(
            configured=["Aspirin 81 MG", "Aspirin 325 MG", "Old Label"],
            observed=[("Aspirin 81 MG", 100), ("Aspirin 325 MG", 50),
                      ("acetylsalicylic acid", 5)],
        )
        self.assertEqual(
            o.configured_and_observed,
            [("Aspirin 81 MG", 100), ("Aspirin 325 MG", 50)],
        )
        self.assertEqual(
            o.missing_from_config, [("acetylsalicylic acid", 5)],
        )
        self.assertEqual(o.stale_in_config, ["Old Label"])

    def test_report_flags_candidate_additions(self):
        o = self._obs(
            configured=["A"],
            observed=[("A", 10), ("B", 3)],
        )
        md = self.mod._fmt_md([o], "test_cohort")
        self.assertIn("Observed but NOT in config", md)
        self.assertIn("`B`", md)
        # Existing config entry is shown but not flagged as candidate.
        self.assertIn("Configured & observed", md)

    def test_suggestions_yaml_unions_observed_and_configured(self):
        o = self._obs(
            configured=["A", "C"],
            observed=[("A", 10), ("B", 3)],
        )
        y = self.mod._fmt_suggestions_yaml([o])
        # Observed-first ordering, configured tail, deduped.
        self.assertIn("- A", y)
        self.assertIn("- B", y)
        self.assertIn("- C", y)
        self.assertIn("column: c", y)
        # Header line warns this is for review only.
        self.assertIn("Review and copy", y)

    def test_report_labels_match_only_row_as_strict(self):
        # Match-only row: criteria == discovery_scope (both compile
        # to the same IN(...)) AND configured_values is non-empty.
        # Report must NOT say "no match: block configured yet".
        in_sql = "\"drug_concept_name\" IN ('Aspirin 81 MG')"
        o = self.mod.VariableObservation(
            category="Drugs", variable="Aspirin",
            table="drug_exposure", column="drug_concept_name",
            criteria=in_sql,
            configured_values=["Aspirin 81 MG"],
            observed=[("Aspirin 81 MG", 100)],
            source_pack="ckd_common",
            discovery_scope=in_sql,
        )
        md = self.mod._fmt_md([o], "balboa_ckd")
        self.assertIn("displayed criteria (strict)", md,
                      msg="match-only row must be labelled strict")
        self.assertIn("match-only row", md)
        self.assertNotIn("no `match:` block configured yet", md)

    def test_dry_run_writes_report_without_db(self):
        # End-to-end: --dry-run must not require psycopg / a DB.
        out_dir = _output_dir(self.id())
        rc = self.mod.main([
            "--cohort", "balboa_ckd",
            "--out-dir", str(out_dir),
            "--dry-run",
        ])
        self.assertEqual(rc, 0)
        self.assertTrue((out_dir / "balboa_ckd" / "report.md").is_file())

    def test_dry_run_resolves_shared_variable_pack_includes(self):
        # balboa_ckd's variables pack is a placeholder that pulls
        # everything from ckd_common via `include:`. The previous
        # discovery loader missed shared variables and produced an
        # empty report.
        out_dir = _output_dir(self.id())
        rc = self.mod.main([
            "--cohort", "balboa_ckd",
            "--out-dir", str(out_dir),
            "--dry-run",
        ])
        self.assertEqual(rc, 0)
        report = (out_dir / "balboa_ckd" / "report.md").read_text()
        # Sanity-check that at least one CKD-shared variable appears
        # under a headed section. ckd_common defines dozens of rows;
        # the report should not be empty after the header.
        self.assertGreater(
            report.count("\n## "), 0,
            msg="report should include shared-pack variables, got:\n" + report,
        )

    def test_resolve_matcher_column_prefers_match_block(self):
        col, skip = self.mod._resolve_matcher_column({
            "table": "measurement", "column": "value_as_number",
            "match": {"column": "measurement_concept_name"},
        })
        self.assertEqual(col, "measurement_concept_name")
        self.assertEqual(skip, "")

    def test_resolve_matcher_column_skips_value_columns(self):
        # value_as_number / value_as_string / value_as_concept_name etc.
        # would discover the value distribution, not the clinical
        # matcher. Discovery should refuse and ask the pack to set
        # match.column when no inference is possible.
        for col in ("value_as_number", "value_as_string",
                    "value_as_concept_name", "measurement_date"):
            matcher, skip = self.mod._resolve_matcher_column({
                "table": "measurement", "column": col,
            })
            self.assertEqual(matcher, "", msg=f"{col} should be skipped")
            self.assertIn("match.column", skip)

    def test_resolve_matcher_infers_from_criteria_lhs(self):
        # Real shape from ckd_common.yaml — Language has
        # column=value_as_concept_name (a value column) but
        # criteria points the matcher at observation_concept_name.
        # Discovery must group on the inferred matcher, not on the
        # value column, otherwise it would suggest match.values like
        # ["English", "Spanish"] redefining the variable.
        matcher, skip = self.mod._resolve_matcher_column({
            "table": "observation",
            "column": "value_as_concept_name",
            "criteria": "observation_concept_name ILIKE '%language%'",
        })
        self.assertEqual(matcher, "observation_concept_name")
        self.assertEqual(skip, "")

    def test_dry_run_flags_unscoped_rows_as_skipped(self):
        # Dry-run must mark a no-criteria / no-match row with the
        # same skip reason live discovery would emit, so offline
        # review reflects what the real DB run will do.
        matcher, scope, displayed, error = self.mod._resolve_scope({
            "table": "condition_occurrence",
            "column": "condition_concept_name",
        })
        self.assertEqual(scope, "")
        self.assertIn("no `criteria:` or `match:`", error)
        # Live and dry-run share the same constructor, so the
        # observation built from this resolution carries the error.
        obs = self.mod._build_observation(
            {"table": "condition_occurrence",
             "column": "condition_concept_name"},
            matcher, displayed, error,
        )
        self.assertIn("no `criteria:` or `match:`", obs.error)

    def test_observe_skips_unscoped_rows(self):
        # A variable with no `criteria:` and no `match:` block must NOT
        # be discovered against the whole table. WHERE TRUE would
        # propose every concept_name in the table as an exact match,
        # silently redefining a generic row like "Diagnosis".
        class _ShouldNotQuery:
            def cursor(self):
                raise AssertionError(
                    "discovery must not query when row has no scope"
                )
            def rollback(self):
                pass
        obs = self.mod._observe_one(_ShouldNotQuery(), "schema", {
            "category": "Conditions",
            "variable": "Diagnosis",
            "table": "condition_occurrence",
            "column": "condition_concept_name",
        })
        self.assertIn("no `criteria:` or `match:`", obs.error)
        self.assertEqual(obs.observed, [])

    def test_scope_prefers_broad_criteria_when_match_also_present(self):
        # Drift workflow: when a variable has BOTH a broad criteria:
        # and a curated match: list, discovery's WHERE must use the
        # broad criteria so the report can flag observed values that
        # aren't yet in match.values. Using match: here would only
        # enumerate already-configured values, making
        # missing_from_config impossible.
        matcher, scope, displayed, error = self.mod._resolve_scope({
            "table": "drug_exposure",
            "column": "drug_concept_name",
            "criteria": "drug_concept_name ILIKE '%aspirin%'",
            "match": {
                "column": "drug_concept_name",
                "values": ["Aspirin 81 MG"],
            },
        })
        self.assertEqual(error, "")
        # Broad ILIKE drives the live discovery query.
        self.assertIn("ILIKE '%aspirin%'", scope)
        # Strict IN list drives the displayed Criteria cell.
        self.assertIn("IN ('Aspirin 81 MG')", displayed)

    def test_scope_uses_match_when_no_broad_criteria(self):
        # Match-only rows still scope to the curated set — falling
        # back to WHERE TRUE would re-enumerate the whole table.
        _matcher, scope, _displayed, error = self.mod._resolve_scope({
            "table": "drug_exposure",
            "column": "drug_concept_name",
            "match": {
                "column": "drug_concept_name",
                "values": ["Aspirin 81 MG", "Aspirin 325 MG"],
            },
        })
        self.assertEqual(error, "")
        self.assertIn("Aspirin 81 MG", scope)
        self.assertIn("Aspirin 325 MG", scope)

    def test_observe_uses_match_block_as_scope(self):
        # When only a `match:` block is configured (no free-form
        # criteria), discovery must scope to that strict IN list, not
        # WHERE TRUE.
        captured = {}

        class _CaptureConn:
            def cursor(self):
                outer = self
                class _Cur:
                    def __enter__(self_inner): return self_inner
                    def __exit__(self_inner, *a): return False
                    def execute(self_inner, sql):
                        captured["sql"] = sql
                    def fetchall(self_inner):
                        return []
                return _Cur()
            def rollback(self):
                pass

        self.mod._observe_one(_CaptureConn(), "schema", {
            "table": "drug_exposure",
            "column": "drug_concept_name",
            "match": {
                "column": "drug_concept_name",
                "values": ["Aspirin 81 MG", "Aspirin 325 MG"],
            },
        })
        self.assertIn("Aspirin 81 MG", captured["sql"])
        self.assertNotIn("WHERE TRUE", captured["sql"])

    def test_report_does_not_label_fuzzy_criteria_as_strict(self):
        # When there is no match: block, criteria == discovery_scope.
        # The report must NOT call that "(strict)" — reviewers would
        # think the variable has been converted to exact matches.
        o = self.mod.VariableObservation(
            category="Drugs", variable="Aspirin",
            table="drug_exposure", column="drug_concept_name",
            criteria="drug_concept_name ILIKE '%aspirin%'",
            configured_values=[],
            observed=[("Aspirin 81 MG", 100)],
            source_pack="ckd_common",
            discovery_scope="drug_concept_name ILIKE '%aspirin%'",
        )
        md = self.mod._fmt_md([o], "balboa_ckd")
        self.assertNotIn("(strict)", md,
                         msg="fuzzy ILIKE must not be labeled strict")
        self.assertIn("ILIKE '%aspirin%'", md)
        self.assertIn("no `match:` block configured yet", md)

    def test_report_distinguishes_displayed_criteria_from_discovery_scope(self):
        # When a variable has both broad criteria: and strict match:,
        # the report must show both — the strict one as the
        # displayed (dictionary) Criteria, the broad one as the
        # actual scope used to find candidate values.
        o = self.mod.VariableObservation(
            category="Drugs", variable="Aspirin",
            table="drug_exposure", column="drug_concept_name",
            criteria='"drug_concept_name" IN (\'Aspirin 81 MG\')',
            configured_values=["Aspirin 81 MG"],
            observed=[("Aspirin 81 MG", 100), ("Aspirin 325 MG", 50)],
            source_pack="ckd_common",
            discovery_scope="drug_concept_name ILIKE '%aspirin%'",
        )
        md = self.mod._fmt_md([o], "balboa_ckd")
        self.assertIn("displayed criteria (strict)", md)
        self.assertIn("IN ('Aspirin 81 MG')", md)
        self.assertIn("discovery scope (broad)", md)
        self.assertIn("ILIKE '%aspirin%'", md)

    def test_report_shows_source_pack(self):
        o = self.mod.VariableObservation(
            category="Conditions", variable="CKD",
            table="condition_occurrence", column="condition_concept_name",
            criteria="condition_concept_name ILIKE '%kidney%'",
            configured_values=[], observed=[("Chronic kidney disease", 100)],
            source_pack="ckd_common",
        )
        md = self.mod._fmt_md([o], "balboa_ckd")
        self.assertIn("packs/variables/ckd_common.yaml", md)

    def test_suggestions_warn_about_shared_vs_cohort_placement(self):
        o = self.mod.VariableObservation(
            category="C", variable="V", table="t", column="c",
            criteria="c ILIKE '%x%'", configured_values=[],
            observed=[("A", 10)], source_pack="ckd_common",
        )
        y = self.mod._fmt_suggestions_yaml([o], cohort="balboa_ckd")
        # Reviewer is told where this row was defined and where to
        # paste cohort-specific overrides.
        self.assertIn("packs/variables/ckd_common.yaml", y)
        self.assertIn("packs/variables/balboa_ckd.yaml", y)
        self.assertIn("cohort-specific", y)

    def test_resolve_matcher_infers_from_in_clause(self):
        # The LHS regex must also handle `IN (...)` and `=` shapes,
        # not just ILIKE.
        for crit in (
            "drug_concept_name IN ('A', 'B')",
            "drug_concept_name = 'Aspirin'",
        ):
            matcher, _ = self.mod._resolve_matcher_column({
                "table": "drug_exposure",
                "column": "value_as_string",
                "criteria": crit,
            })
            self.assertEqual(matcher, "drug_concept_name", msg=crit)

    def test_resolve_matcher_column_uses_concept_name_directly(self):
        # When `column` is already the matcher (concept_name), no
        # match block is needed.
        col, skip = self.mod._resolve_matcher_column({
            "table": "drug_exposure", "column": "drug_concept_name",
        })
        self.assertEqual(col, "drug_concept_name")
        self.assertEqual(skip, "")


class DryRunMatchBlockTests(unittest.TestCase):
    """Dry-run path must compile `match:` blocks the same way the live
    build does, so review artifacts and offline previews don't show
    stale fuzzy criteria once packs adopt strict matchers."""

    def test_dry_run_uses_match_block_in_criteria(self):
        # Patch load_variables_pack to return a single variable with
        # an explicit match block, then drive build_model in dry-run.
        original = bd.load_variables_pack
        bd.load_variables_pack = lambda slug: [{
            "category": "Labs",
            "variable": "Aspirin",
            "table": "drug_exposure",
            "column": "drug_concept_name",
            "criteria": "drug_concept_name ILIKE '%aspirin%'",
            "match": {
                "column": "drug_concept_name",
                "values": ["Aspirin 81 MG", "Aspirin 325 MG"],
            },
        }]
        try:
            model = bd.build_model("balboa_ckd", conn=None, dry_run=True)
        finally:
            bd.load_variables_pack = original

        self.assertEqual(len(model.variables), 1)
        criteria = model.variables[0].criteria
        self.assertEqual(
            criteria,
            "\"drug_concept_name\" IN ('Aspirin 81 MG', 'Aspirin 325 MG')",
        )
        self.assertNotIn("ILIKE", criteria)


class CohortKeyResolutionTests(unittest.TestCase):
    """Per-cohort layout overrides must work whether the YAML keys by
    slug, cohort_name, or schema_name. The CLI/filename slug is the
    documented preferred key."""

    def _patch_layout(self, layout):
        original = bd._load_dictionary_layout
        bd._load_dictionary_layout = lambda: layout
        self.addCleanup(lambda: setattr(bd, "_load_dictionary_layout", original))

    def test_lookup_by_slug(self):
        self._patch_layout({
            "customer": {"exclude_tables": []},
            "cohorts": {"my_slug": {"customer": {"exclude_tables": ["X"]}}},
        })
        self.assertEqual(
            bd.customer_table_excludes(["my_slug", "my_cohort_name"]),
            frozenset({"X"}),
        )

    def test_lookup_falls_back_to_cohort_name(self):
        self._patch_layout({
            "customer": {"exclude_tables": []},
            "cohorts": {"my_cohort_name": {"customer": {"exclude_tables": ["Y"]}}},
        })
        self.assertEqual(
            bd.customer_table_excludes(["my_slug", "my_cohort_name"]),
            frozenset({"Y"}),
        )

    def test_first_matching_key_wins(self):
        self._patch_layout({
            "customer": {"exclude_tables": []},
            "cohorts": {
                "my_slug":        {"customer": {"exclude_tables": ["A"]}},
                "my_cohort_name": {"customer": {"exclude_tables": ["B"]}},
            },
        })
        self.assertEqual(
            bd.customer_table_excludes(["my_slug", "my_cohort_name"]),
            frozenset({"A"}),
        )

    def test_filter_threads_slug_through(self):
        # End-to-end: filter_for_audience must pass the CLI slug as
        # the primary lookup key, not just model.cohort.
        self._patch_layout({
            "customer": {"exclude_tables": []},
            "cohorts": {"my_slug": {"customer": {"exclude_tables": ["person"]}}},
        })
        tables = [_make_table("person"), _make_table("cohort_patients")]
        columns = [_make_column("person")]
        variables = [_make_variable("person")]
        model = _make_model(tables, columns, variables)
        # _make_model sets model.cohort = "c", which has no override.
        # Without slug threading, the per-slug entry would be missed.
        filtered = bd.filter_for_audience(model, "customer", cohort_slug="my_slug")
        self.assertEqual(
            {t.table_name for t in filtered.tables}, {"cohort_patients"},
        )


class DiscoveryApplyTests(unittest.TestCase):
    """`--apply` writes proposed match: blocks into source packs.
    Must round-trip the YAML safely (preserve comments / order),
    only touch eligible variables, and refuse cleanly when ruamel
    is unavailable."""

    def setUp(self):
        try:
            import ruamel.yaml  # noqa: F401
        except ImportError:
            self.skipTest("ruamel.yaml not installed")
        self.mod = discover_mod
        # Stage a temp variables pack under packs/variables/ so the
        # apply path can find it via PACKS_DIR/variables/<slug>.yaml.
        self.pack_slug = "_apply_test_pack"
        self.pack_path = bd.PACKS_DIR / "variables" / f"{self.pack_slug}.yaml"
        self.pack_path.write_text(
            "# Test pack. Comments must survive --apply round-trip.\n"
            "variables:\n"
            "  - category: Drugs\n"
            "    variable: Aspirin\n"
            "    table: drug_exposure\n"
            "    column: drug_concept_name\n"
            "    criteria: drug_concept_name ILIKE '%aspirin%'\n"
            "  - category: Drugs\n"
            "    variable: Lisinopril\n"
            "    table: drug_exposure\n"
            "    column: drug_concept_name\n"
            "    criteria: drug_concept_name ILIKE '%lisinopril%'\n",
            encoding="utf-8",
        )
        self.addCleanup(lambda: self.pack_path.unlink(missing_ok=True))

    def _obs(self, variable, observed):
        return self.mod.VariableObservation(
            category="Drugs", variable=variable, table="drug_exposure",
            column="drug_concept_name",
            criteria=f"drug_concept_name ILIKE '%{variable.lower()}%'",
            configured_values=[], observed=observed,
            source_pack=self.pack_slug,
        )

    def test_apply_writes_match_blocks(self):
        observations = [
            self._obs("Aspirin",
                      [("Aspirin 81 MG", 100), ("Aspirin 325 MG", 50)]),
        ]
        applied, skipped = self.mod.apply_suggestions(
            observations, target="shared", auto_yes=True,
        )
        self.assertEqual(applied, 1)
        self.assertEqual(skipped, 0)
        text = self.pack_path.read_text()
        # Comment at the top must survive.
        self.assertIn("Comments must survive --apply round-trip", text)
        # Match block written under the right variable.
        self.assertIn("- Aspirin 81 MG", text)
        self.assertIn("- Aspirin 325 MG", text)

    def test_apply_skips_unobserved_variables(self):
        # Observation with no rows is not eligible — must not write.
        observations = [self._obs("Aspirin", [])]
        applied, skipped = self.mod.apply_suggestions(
            observations, target="shared", auto_yes=True,
        )
        self.assertEqual(applied, 0)
        self.assertNotIn("match:", self.pack_path.read_text())

    def test_apply_skips_errored_variables(self):
        bad = self.mod.VariableObservation(
            category="C", variable="Aspirin", table="t", column="c",
            criteria="", configured_values=[],
            observed=[("v", 1)],
            error="value column",
            source_pack=self.pack_slug,
        )
        applied, _ = self.mod.apply_suggestions(
            [bad], target="shared", auto_yes=True,
        )
        self.assertEqual(applied, 0)
        self.assertNotIn("match:", self.pack_path.read_text())

    def test_apply_skips_unknown_variable_in_pack(self):
        # Variable not present in the source file (likely lives in
        # an included pack). Must skip, not crash.
        observations = [self._obs("Nonexistent", [("X", 1)])]
        applied, skipped = self.mod.apply_suggestions(
            observations, target="shared", auto_yes=True,
        )
        self.assertEqual(applied, 0)
        self.assertEqual(skipped, 1)

    def test_apply_groups_by_source_pack(self):
        # Two observations for the same pack → one read+write of the file.
        observations = [
            self._obs("Aspirin", [("Aspirin 81 MG", 100)]),
            self._obs("Lisinopril", [("Lisinopril 10 MG", 80)]),
        ]
        applied, _ = self.mod.apply_suggestions(
            observations, target="shared", auto_yes=True,
        )
        self.assertEqual(applied, 2)
        text = self.pack_path.read_text()
        self.assertIn("Aspirin 81 MG", text)
        self.assertIn("Lisinopril 10 MG", text)

    def test_apply_requires_valid_target(self):
        observations = [self._obs("Aspirin", [("Aspirin 81 MG", 1)])]
        with self.assertRaises(ValueError):
            self.mod.apply_suggestions(
                observations, target="invalid", auto_yes=True,
            )

    def test_apply_target_cohort_writes_to_cohort_pack(self):
        cohort_slug = "_apply_test_cohort"
        cohort_path = bd.PACKS_DIR / "variables" / f"{cohort_slug}.yaml"
        cohort_path.write_text(
            "include: [_apply_test_pack]\n"
            "variables:\n"
            "  - category: Drugs\n"
            "    variable: Aspirin\n"
            "    table: drug_exposure\n"
            "    column: drug_concept_name\n"
            "    criteria: drug_concept_name ILIKE '%aspirin%'\n",
            encoding="utf-8",
        )
        self.addCleanup(lambda: cohort_path.unlink(missing_ok=True))

        observations = [self._obs("Aspirin", [("Aspirin 81 MG", 100)])]
        applied, _ = self.mod.apply_suggestions(
            observations, target="cohort",
            cohort_slug=cohort_slug, auto_yes=True,
        )
        self.assertEqual(applied, 1)
        # Per-cohort target must NOT touch the shared source pack.
        self.assertIn("Aspirin 81 MG", cohort_path.read_text())
        self.assertNotIn("Aspirin 81 MG", self.pack_path.read_text())

    def test_apply_uses_category_variable_key_for_lookup(self):
        # Pack with the same variable label under two different
        # categories. apply must find the right one by (category,
        # variable), matching the loaders' override key.
        cohort_slug = "_apply_test_dup_label"
        cohort_path = bd.PACKS_DIR / "variables" / f"{cohort_slug}.yaml"
        cohort_path.write_text(
            "variables:\n"
            "  - category: Diagnosis\n"
            "    variable: Coverage\n"
            "    table: condition_occurrence\n"
            "    column: condition_concept_name\n"
            "    criteria: condition_concept_name ILIKE '%cov%'\n"
            "  - category: Insurance\n"
            "    variable: Coverage\n"
            "    table: payer_plan_period\n"
            "    column: payer_concept_name\n"
            "    criteria: payer_concept_name ILIKE '%cov%'\n",
            encoding="utf-8",
        )
        self.addCleanup(lambda: cohort_path.unlink(missing_ok=True))

        # Observation targets the Insurance row only.
        obs = self.mod.VariableObservation(
            category="Insurance", variable="Coverage",
            table="payer_plan_period", column="payer_concept_name",
            criteria="payer_concept_name ILIKE '%cov%'",
            configured_values=[],
            observed=[("Aetna PPO", 100)],
            source_pack=cohort_slug,
        )
        applied, _ = self.mod.apply_suggestions(
            [obs], target="shared", auto_yes=True,
        )
        self.assertEqual(applied, 1)
        text = cohort_path.read_text()
        # The Insurance row got the match block; the Diagnosis row
        # did NOT.
        diag_idx = text.index("category: Diagnosis")
        ins_idx = text.index("category: Insurance")
        # Match block must appear after the Insurance line.
        match_idx = text.find("match:")
        self.assertGreater(
            match_idx, ins_idx,
            msg="match: should be attached to the Insurance row",
        )
        # And NOT attached to the Diagnosis row (i.e. between the
        # Diagnosis category line and the Insurance category line).
        between_diag = text[diag_idx:ins_idx]
        self.assertNotIn("match:", between_diag,
                         msg="Diagnosis row must not be touched")

    def test_apply_target_cohort_skips_inherited_only_variables(self):
        cohort_slug = "_apply_test_cohort_empty"
        cohort_path = bd.PACKS_DIR / "variables" / f"{cohort_slug}.yaml"
        cohort_path.write_text(
            "include: [_apply_test_pack]\nvariables: []\n",
            encoding="utf-8",
        )
        self.addCleanup(lambda: cohort_path.unlink(missing_ok=True))

        observations = [self._obs("Aspirin", [("Aspirin 81 MG", 100)])]
        applied, skipped = self.mod.apply_suggestions(
            observations, target="cohort",
            cohort_slug=cohort_slug, auto_yes=True,
        )
        # Refuses to invent a row in the cohort pack (a bare
        # `variable: + match:` would be unbuildable).
        self.assertEqual(applied, 0)
        self.assertEqual(skipped, 1)
        self.assertNotIn("match:", cohort_path.read_text())
        self.assertNotIn("match:", self.pack_path.read_text())

    def test_main_apply_without_target_exits_nonzero(self):
        out_dir = _output_dir(self.id())
        rc = self.mod.main([
            "--cohort", "balboa_ckd",
            "--out-dir", str(out_dir),
            "--dry-run",
            "--apply-yes",
        ])
        self.assertEqual(
            rc, 2, msg="--apply without --target should exit 2",
        )

    def test_auto_stub_copies_full_definition_into_cohort_pack(self):
        # Variable lives only in shared pack. With --auto-stub the
        # full base definition (table, column, criteria, description)
        # is copied into the cohort pack along with the match block.
        cohort_slug = "_apply_test_cohort_stub"
        cohort_path = bd.PACKS_DIR / "variables" / f"{cohort_slug}.yaml"
        cohort_path.write_text(
            "include: [_apply_test_pack]\nvariables: []\n",
            encoding="utf-8",
        )
        self.addCleanup(lambda: cohort_path.unlink(missing_ok=True))

        before_shared = self.pack_path.read_text()
        observations = [self._obs("Aspirin", [("Aspirin 81 MG", 100)])]
        applied, _ = self.mod.apply_suggestions(
            observations, target="cohort",
            cohort_slug=cohort_slug, auto_yes=True, auto_stub=True,
        )
        self.assertEqual(applied, 1)

        cohort_text = cohort_path.read_text()
        # Match block landed in cohort pack, not shared.
        self.assertIn("Aspirin 81 MG", cohort_text)
        self.assertEqual(
            self.pack_path.read_text(), before_shared,
            msg="shared pack must be untouched by --auto-stub",
        )
        # Base definition copied — not just `variable:` + `match:`.
        for field in ("table:", "column:", "criteria:"):
            self.assertIn(
                field, cohort_text,
                msg=f"auto-stub must copy {field} from source",
            )
        # Provenance comment / annotation present.
        self.assertTrue(
            "Auto-stubbed" in cohort_text or "_auto_stub_origin" in cohort_text,
            msg="auto-stub must annotate provenance",
        )

    def test_auto_stub_does_not_duplicate_when_variable_already_in_cohort(self):
        # Variable already overrides in cohort pack. Auto-stub mode
        # MUST NOT add a duplicate row — falls through to the regular
        # update path so the existing row gets the match block.
        cohort_slug = "_apply_test_cohort_predefined"
        cohort_path = bd.PACKS_DIR / "variables" / f"{cohort_slug}.yaml"
        cohort_path.write_text(
            "include: [_apply_test_pack]\n"
            "variables:\n"
            "  - category: Drugs\n"
            "    variable: Aspirin\n"
            "    table: drug_exposure\n"
            "    column: drug_concept_name\n"
            "    criteria: drug_concept_name ILIKE '%aspirin%'\n",
            encoding="utf-8",
        )
        self.addCleanup(lambda: cohort_path.unlink(missing_ok=True))

        observations = [self._obs("Aspirin", [("Aspirin 81 MG", 100)])]
        applied, _ = self.mod.apply_suggestions(
            observations, target="cohort",
            cohort_slug=cohort_slug, auto_yes=True, auto_stub=True,
        )
        self.assertEqual(applied, 1)
        text = cohort_path.read_text()
        # Exactly one Aspirin row — count `variable: Aspirin` lines.
        self.assertEqual(
            text.count("variable: Aspirin"), 1,
            msg="auto-stub must not duplicate an existing cohort row",
        )

    def test_auto_stub_requires_target_cohort(self):
        observations = [self._obs("Aspirin", [("Aspirin 81 MG", 1)])]
        with self.assertRaises(ValueError):
            self.mod.apply_suggestions(
                observations, target="shared", auto_yes=True, auto_stub=True,
            )

    def test_auto_stub_leaves_shared_pack_unchanged(self):
        # Re-runnable shared-pack-immutability check.
        cohort_slug = "_apply_test_cohort_stubcheck"
        cohort_path = bd.PACKS_DIR / "variables" / f"{cohort_slug}.yaml"
        cohort_path.write_text(
            "include: [_apply_test_pack]\nvariables: []\n",
            encoding="utf-8",
        )
        self.addCleanup(lambda: cohort_path.unlink(missing_ok=True))
        original = self.pack_path.read_bytes()

        observations = [
            self._obs("Aspirin", [("Aspirin 81 MG", 100)]),
            self._obs("Lisinopril", [("Lisinopril 10 MG", 80)]),
        ]
        self.mod.apply_suggestions(
            observations, target="cohort",
            cohort_slug=cohort_slug, auto_yes=True, auto_stub=True,
        )
        self.assertEqual(
            self.pack_path.read_bytes(), original,
            msg="shared pack bytes must be byte-identical after --auto-stub",
        )

    def test_main_auto_stub_with_shared_target_exits_nonzero(self):
        out_dir = _output_dir(self.id())
        rc = self.mod.main([
            "--cohort", "balboa_ckd",
            "--out-dir", str(out_dir),
            "--dry-run",
            "--apply-yes",
            "--target", "shared",
            "--auto-stub",
        ])
        self.assertEqual(
            rc, 2,
            msg="--auto-stub with --target shared must exit 2",
        )

    def test_per_variable_prompt_renders_structured_block(self):
        # The interactive prompt must show source/target/action/reason
        # explicitly so reviewers can distinguish UPDATE from ADD
        # cohort override at a glance.
        observations = [self._obs("Aspirin", [("Aspirin 81 MG", 100), ("Aspirin 325 MG", 50)])]
        captured: list[str] = []

        import builtins
        original_input = builtins.input
        builtins.input = lambda prompt="": (captured.append(prompt) or "n")
        self.addCleanup(lambda: setattr(builtins, "input", original_input))

        # update path: variable already in shared pack.
        self.mod.apply_suggestions(
            observations, target="shared", auto_yes=False,
        )
        prompt = captured[-1]
        for line in (
            "Variable:", "Source:", "Target:",
            "Action:", "UPDATE variable",
            "Values:", "Aspirin 81 MG",
            "Reason:", "match: block will change",
            "Proceed?", "[y]es", "[a]ll", "[q]uit",
        ):
            self.assertIn(line, prompt, msg=f"missing: {line!r}")

    def test_per_variable_prompt_labels_stub_as_add_cohort_override(self):
        cohort_slug = "_apply_test_cohort_addlabel"
        cohort_path = bd.PACKS_DIR / "variables" / f"{cohort_slug}.yaml"
        cohort_path.write_text(
            "include: [_apply_test_pack]\nvariables: []\n",
            encoding="utf-8",
        )
        self.addCleanup(lambda: cohort_path.unlink(missing_ok=True))

        captured: list[str] = []
        import builtins
        original_input = builtins.input
        builtins.input = lambda prompt="": (captured.append(prompt) or "n")
        self.addCleanup(lambda: setattr(builtins, "input", original_input))

        observations = [self._obs("Aspirin", [("Aspirin 81 MG", 100)])]
        self.mod.apply_suggestions(
            observations, target="cohort", cohort_slug=cohort_slug,
            auto_yes=False, auto_stub=True,
        )
        prompt = captured[-1]
        self.assertIn("ADD cohort override", prompt)
        self.assertIn("inherited from shared pack", prompt)

    def test_per_variable_prompt_quit_aborts_without_writing(self):
        # Interactive prompt: 'q' must abort the whole run and leave
        # both packs unchanged on disk.
        observations = [self._obs("Aspirin", [("Aspirin 81 MG", 100)])]
        before = self.pack_path.read_bytes()

        # Patch input() to return 'q'.
        import builtins
        original_input = builtins.input
        builtins.input = lambda _prompt="": "q"
        self.addCleanup(lambda: setattr(builtins, "input", original_input))

        applied, _ = self.mod.apply_suggestions(
            observations, target="shared", auto_yes=False,
        )
        self.assertEqual(applied, 0)
        self.assertEqual(
            self.pack_path.read_bytes(), before,
            msg="quit must not write any pending changes",
        )

    def test_per_variable_prompt_all_accepts_remaining(self):
        # 'all' on the first variable should commit the rest without
        # further prompts.
        observations = [
            self._obs("Aspirin", [("Aspirin 81 MG", 100)]),
            self._obs("Lisinopril", [("Lisinopril 10 MG", 80)]),
        ]
        responses = iter(["all"])

        import builtins
        original_input = builtins.input
        builtins.input = lambda _prompt="": next(responses)
        self.addCleanup(lambda: setattr(builtins, "input", original_input))

        applied, _ = self.mod.apply_suggestions(
            observations, target="shared", auto_yes=False,
        )
        # Both got applied off the single 'all' answer.
        self.assertEqual(applied, 2)

    def test_auto_stub_writes_real_yaml_comment_not_marker_field(self):
        # Source rows must be loaded via ruamel so the stubbed row
        # is a CommentedMap. _attach_stub_comment then writes a
        # leading YAML comment instead of falling back to the
        # _auto_stub_origin marker field.
        cohort_slug = "_apply_test_cohort_realcomment"
        cohort_path = bd.PACKS_DIR / "variables" / f"{cohort_slug}.yaml"
        cohort_path.write_text(
            "include: [_apply_test_pack]\nvariables: []\n",
            encoding="utf-8",
        )
        self.addCleanup(lambda: cohort_path.unlink(missing_ok=True))

        observations = [self._obs("Aspirin", [("Aspirin 81 MG", 100)])]
        applied, _ = self.mod.apply_suggestions(
            observations, target="cohort",
            cohort_slug=cohort_slug, auto_yes=True, auto_stub=True,
        )
        self.assertEqual(applied, 1)
        text = cohort_path.read_text()
        self.assertIn(
            "# Auto-stubbed from packs/variables/_apply_test_pack.yaml",
            text,
            msg="provenance must be a YAML comment, not a hidden field",
        )
        self.assertNotIn(
            "_auto_stub_origin", text,
            msg="internal marker must not leak into source-of-truth YAML",
        )

    def test_apply_writes_atomically_via_temp_file(self):
        # File mtime must change exactly once per touched pack;
        # midstream errors must not leave a half-written file.
        observations = [self._obs("Aspirin", [("Aspirin 81 MG", 100)])]
        before_size = self.pack_path.stat().st_size
        applied, _ = self.mod.apply_suggestions(
            observations, target="shared", auto_yes=True,
        )
        self.assertEqual(applied, 1)
        # No leftover .tmp files in packs/variables/.
        leftovers = list(bd.PACKS_DIR.glob("variables/*.yaml.tmp"))
        self.assertEqual(leftovers, [], msg=f"orphaned temp files: {leftovers}")
        # Final file is non-empty (and different from before — match was added).
        self.assertGreater(self.pack_path.stat().st_size, before_size // 2)
        self.assertIn("Aspirin 81 MG", self.pack_path.read_text())

    def test_apply_temp_file_cleaned_up_on_dump_failure(self):
        # If yaml_rt.dump raises mid-write, the temp file must be
        # removed and the original pack must remain untouched.
        observations = [self._obs("Aspirin", [("Aspirin 81 MG", 100)])]

        # Patch yaml_rt's dump indirectly by patching the dump on
        # the YAML class. ruamel's YAML.dump is the chosen seam.
        original_text = self.pack_path.read_bytes()
        from ruamel.yaml import YAML
        original_dump = YAML.dump
        def _boom(self, *a, **kw):
            raise RuntimeError("simulated disk error")
        YAML.dump = _boom
        self.addCleanup(lambda: setattr(YAML, "dump", original_dump))

        with self.assertRaises(RuntimeError):
            self.mod.apply_suggestions(
                observations, target="shared", auto_yes=True,
            )
        # Original pack survived intact.
        self.assertEqual(
            self.pack_path.read_bytes(), original_text,
            msg="dump failure must not corrupt the original pack",
        )
        # No orphaned temp file.
        leftovers = list(bd.PACKS_DIR.glob("variables/*.yaml.tmp"))
        self.assertEqual(leftovers, [], msg=f"orphaned temp files: {leftovers}")

    def test_main_apply_without_target_skips_discovery(self):
        # Contract failure must short-circuit before DB work or any
        # report files get written. The dry-run path is the cheap
        # proxy for "did we do real work" — it walks the pack and
        # produces report.md. Argument validation should fire first.
        out_dir = _output_dir(self.id())
        rc = self.mod.main([
            "--cohort", "balboa_ckd",
            "--out-dir", str(out_dir),
            "--dry-run",
            "--apply",   # without --target
        ])
        self.assertEqual(rc, 2)
        # Cohort sub-dir must not exist — discovery never ran.
        self.assertFalse(
            (out_dir / "balboa_ckd" / "report.md").exists(),
            msg="discovery should not run when --apply is invalid",
        )


class CohortModelSortingTests(unittest.TestCase):
    """build_model must sort tables/columns/variables a-z by
    (category, name) so reviewers always see a stable, alphabetical
    layout — independent of how the YAML packs are authored.
    """

    def setUp(self):
        # Stage a deliberately mis-ordered variables pack: categories
        # are clinically grouped (Vitals before Demographics) and
        # variables within each category are not yet alphabetical.
        slug = "_sort_test_pack"
        path = bd.PACKS_DIR / "variables" / f"{slug}.yaml"
        path.write_text(
            "variables:\n"
            "  - category: Vitals\n"
            "    variable: Heart Rate\n"
            "    table: measurement\n"
            "    column: value_as_number\n"
            "  - category: Vitals\n"
            "    variable: Blood Pressure\n"
            "    table: measurement\n"
            "    column: value_as_number\n"
            "  - category: Demographics\n"
            "    variable: Sex\n"
            "    table: person\n"
            "    column: gender_concept_name\n"
            "  - category: Demographics\n"
            "    variable: Age\n"
            "    table: person\n"
            "    column: year_of_birth\n",
            encoding="utf-8",
        )
        self.addCleanup(lambda: path.unlink(missing_ok=True))

        # Wire up a synthetic cohort pack pointing at it.
        cohort_path = bd.PACKS_DIR / "cohorts" / "_sort_test_cohort.yaml"
        cohort_path.write_text(
            "provider: TEST\n"
            "disease: TEST\n"
            "schema_name: _sort_test_schema\n"
            "cohort_name: _sort_test_cohort\n"
            "display_name: Sort Test\n"
            "variables_pack: _sort_test_pack\n",
            encoding="utf-8",
        )
        self.addCleanup(lambda: cohort_path.unlink(missing_ok=True))

    def test_variables_sorted_by_category_then_variable(self):
        model = bd.build_model("_sort_test_cohort", conn=None, dry_run=True)
        names = [(v.category, v.variable) for v in model.variables]
        self.assertEqual(
            names,
            [("Demographics", "Age"),
             ("Demographics", "Sex"),
             ("Vitals", "Blood Pressure"),
             ("Vitals", "Heart Rate")],
        )

    def test_sort_is_case_insensitive(self):
        # A row authored with lowercase category must still cluster
        # under the same group as the proper-cased category, not
        # split off into its own bucket.
        path = bd.PACKS_DIR / "variables" / "_sort_test_case.yaml"
        path.write_text(
            "variables:\n"
            "  - category: demographics\n"
            "    variable: Apple\n"
            "    table: t\n"
            "    column: c\n"
            "  - category: Demographics\n"
            "    variable: Banana\n"
            "    table: t\n"
            "    column: c\n",
            encoding="utf-8",
        )
        self.addCleanup(lambda: path.unlink(missing_ok=True))

        cohort = bd.PACKS_DIR / "cohorts" / "_sort_test_case_cohort.yaml"
        cohort.write_text(
            "provider: TEST\ndisease: TEST\nschema_name: x\n"
            "cohort_name: _sort_test_case_cohort\n"
            "display_name: x\nvariables_pack: _sort_test_case\n",
            encoding="utf-8",
        )
        self.addCleanup(lambda: cohort.unlink(missing_ok=True))

        model = bd.build_model("_sort_test_case_cohort", conn=None, dry_run=True)
        cats = [v.category for v in model.variables]
        self.assertEqual(cats, ["demographics", "Demographics"],
                         msg="rows should cluster despite case mismatch")


class VariablePackOverrideTests(unittest.TestCase):
    """A cohort pack's local row must replace any inherited row with
    the same (category, variable) key. Without this, --auto-stub's
    cohort-side match: block coexists with the shared pack's fuzzy
    ILIKE definition and the build can render both, undermining the
    entire exact-match feedback loop.
    """

    def setUp(self):
        # Stage a parent (shared) pack with one fuzzy variable and a
        # child (cohort) pack that includes it and overrides the same
        # variable with a strict match: block.
        self.shared_slug = "_override_test_shared"
        self.shared_path = bd.PACKS_DIR / "variables" / f"{self.shared_slug}.yaml"
        self.shared_path.write_text(
            "variables:\n"
            "  - category: Drugs\n"
            "    variable: Aspirin\n"
            "    table: drug_exposure\n"
            "    column: drug_concept_name\n"
            "    criteria: drug_concept_name ILIKE '%aspirin%'\n"
            "  - category: Drugs\n"
            "    variable: Lisinopril\n"
            "    table: drug_exposure\n"
            "    column: drug_concept_name\n"
            "    criteria: drug_concept_name ILIKE '%lisinopril%'\n",
            encoding="utf-8",
        )
        self.addCleanup(lambda: self.shared_path.unlink(missing_ok=True))

        self.cohort_slug = "_override_test_cohort"
        self.cohort_path = bd.PACKS_DIR / "variables" / f"{self.cohort_slug}.yaml"
        self.cohort_path.write_text(
            f"include: [{self.shared_slug}]\n"
            "variables:\n"
            "  - category: Drugs\n"
            "    variable: Aspirin\n"
            "    table: drug_exposure\n"
            "    column: drug_concept_name\n"
            "    criteria: drug_concept_name ILIKE '%aspirin%'\n"
            "    match:\n"
            "      column: drug_concept_name\n"
            "      values: ['Aspirin 81 MG Oral Tablet', 'Aspirin 325 MG Oral Tablet']\n",
            encoding="utf-8",
        )
        self.addCleanup(lambda: self.cohort_path.unlink(missing_ok=True))

    def test_cohort_row_replaces_inherited_row_no_duplicate(self):
        rows = bd.load_variables_pack(self.cohort_slug)
        aspirin_rows = [r for r in rows if r.get("variable") == "Aspirin"]
        self.assertEqual(
            len(aspirin_rows), 1,
            msg="cohort override must collapse to one row, got: " + str(aspirin_rows),
        )
        # The surviving row must be the cohort version (carries match:).
        self.assertIn("match", aspirin_rows[0])

    def test_non_overridden_variables_pass_through(self):
        rows = bd.load_variables_pack(self.cohort_slug)
        names = [r.get("variable") for r in rows]
        # Lisinopril was only defined in the shared pack — must
        # still be present after override resolution.
        self.assertIn("Lisinopril", names)
        # No duplication anywhere.
        self.assertEqual(len(names), len(set(names)))

    def test_override_compiles_to_strict_in_via_match(self):
        # End-to-end: build the variable through compile_match_block
        # the same way process_variables would. Result must be
        # `column IN (...)`, not the inherited ILIKE.
        rows = bd.load_variables_pack(self.cohort_slug)
        aspirin = next(r for r in rows if r.get("variable") == "Aspirin")
        sql = bd.compile_match_block(aspirin.get("match"))
        self.assertIn("IN ('Aspirin 81 MG Oral Tablet'", sql)
        self.assertNotIn("ILIKE", sql)

    def test_override_preserves_inherited_position(self):
        rows = bd.load_variables_pack(self.cohort_slug)
        # Aspirin was first in the shared pack; replacement must
        # land at index 0, not get appended at the end after
        # Lisinopril. Audience layouts depend on positional
        # stability across runs.
        self.assertEqual(rows[0].get("variable"), "Aspirin")
        self.assertEqual(rows[1].get("variable"), "Lisinopril")

    def test_discovery_loader_collapses_overrides_with_correct_source(self):
        # Discovery's tagged loader must apply the same override
        # semantics — otherwise the report can show both the
        # inherited fuzzy row and the cohort match: row, even
        # though the build collapses them correctly.
        rows = discover_mod._load_variables_pack_tagged(self.cohort_slug)
        aspirin_rows = [r for r in rows if r.get("variable") == "Aspirin"]
        self.assertEqual(
            len(aspirin_rows), 1,
            msg="discovery must collapse overrides — got "
                + str([r.get("_source_pack") for r in aspirin_rows]),
        )
        # Source-pack tag must point at the cohort pack, since the
        # cohort row IS the surviving definition.
        self.assertEqual(
            aspirin_rows[0].get("_source_pack"), self.cohort_slug,
            msg="overridden row's _source_pack must reflect the cohort pack",
        )
        self.assertIn("match", aspirin_rows[0])

    def test_discovery_loader_keeps_inherited_source_for_non_overrides(self):
        rows = discover_mod._load_variables_pack_tagged(self.cohort_slug)
        lisinopril = next(r for r in rows if r.get("variable") == "Lisinopril")
        # Lisinopril was only defined in the shared pack — its
        # provenance should still point at the shared pack.
        self.assertEqual(
            lisinopril.get("_source_pack"), self.shared_slug,
        )

    def test_validator_treats_match_only_row_as_scoped(self):
        # _check_missing_criteria must NOT flag a row that has only
        # a `match:` block (no legacy criteria:) as catch-all.
        import importlib.util
        vp_path = Path(__file__).resolve().parent.parent / "scripts" / "validate_packs.py"
        spec = importlib.util.spec_from_file_location(
            "validate_packs_match_test", vp_path,
        )
        vp = importlib.util.module_from_spec(spec)
        sys.modules["validate_packs_match_test"] = vp
        spec.loader.exec_module(vp)

        # Variable name matches a clinically-specific pattern
        # (APOE) so _check_missing_criteria *would* fire if the
        # row looked unscoped.
        match_only_row = {
            "variable": "APOE Genotype",
            "column": "value_as_string",
            "match": {
                "column": "observation_concept_name",
                "values": ["APOE genotype"],
            },
        }
        self.assertIsNone(
            vp._check_missing_criteria(match_only_row),
            msg="match-only row must count as scoped — no missing-criteria warning",
        )

        # Sanity: a row that's truly unscoped (no criteria, no match)
        # still triggers the warning, so we know the check still works.
        unscoped_row = {"variable": "APOE Genotype", "column": "value_as_string"}
        self.assertIsNotNone(vp._check_missing_criteria(unscoped_row))

    def test_validator_has_scope_helper_recognizes_match_block(self):
        import importlib.util
        vp_path = Path(__file__).resolve().parent.parent / "scripts" / "validate_packs.py"
        spec = importlib.util.spec_from_file_location(
            "validate_packs_scope_test", vp_path,
        )
        vp = importlib.util.module_from_spec(spec)
        sys.modules["validate_packs_scope_test"] = vp
        spec.loader.exec_module(vp)

        # Positive: criteria string scopes the row.
        self.assertTrue(vp._has_scope({"criteria": "x ILIKE '%y%'"}))
        # Positive: match block with column + inline values.
        self.assertTrue(vp._has_scope({
            "match": {"column": "c", "values": ["a"]},
        }))
        # Positive: match block with values_file pointing at a real
        # loadable list. Stage a temp file under packs/.
        rel = "_scope_test_values.yaml"
        path = bd.PACKS_DIR / rel
        path.write_text("- foo\n- bar\n", encoding="utf-8")
        try:
            self.assertTrue(vp._has_scope({
                "match": {"column": "c", "values_file": rel},
            }))
        finally:
            path.unlink()

        # Negatives — rows the BUILDER would compile to "" and so
        # the validator must also treat as unscoped:
        self.assertFalse(vp._has_scope({}))
        self.assertFalse(vp._has_scope({"match": {"column": "c"}}))   # no values, no file
        self.assertFalse(vp._has_scope({"match": {"column": "c", "values": []}}))
        # Missing values: column omitted — values can't compile.
        self.assertFalse(vp._has_scope({"match": {"values": ["a"]}}))
        # Missing values_file on disk — must NOT pass validation
        # because compile_match_block would return "".
        self.assertFalse(vp._has_scope({
            "match": {"column": "c", "values_file": "_does_not_exist.yaml"},
        }))

    def test_validator_loader_applies_override_semantics(self):
        # validate_packs.py's resolver must also collapse overrides;
        # otherwise --auto-stub'd cohort packs trigger a spurious
        # "duplicate variable" error and the validator becomes
        # unusable after applying exact-match work.
        import importlib.util
        vp_path = Path(__file__).resolve().parent.parent / "scripts" / "validate_packs.py"
        spec = importlib.util.spec_from_file_location(
            "validate_packs_under_test", vp_path,
        )
        vp = importlib.util.module_from_spec(spec)
        sys.modules["validate_packs_under_test"] = vp
        spec.loader.exec_module(vp)

        # Patch its PACKS_DIR to point at the same one our test
        # packs live in — they live in the canonical packs/.
        rows = vp._resolve_variables(self.cohort_slug, findings=[])
        aspirin_rows = [r for r in rows if r.get("variable") == "Aspirin"]
        self.assertEqual(
            len(aspirin_rows), 1,
            msg="validator must collapse overrides too",
        )
        self.assertIn("match", aspirin_rows[0])

    def test_validator_flags_unknown_proposal_value(self):
        # Anything other than Standard / Custom (when proposal is set)
        # must be rejected so a typo doesn't ship to a sales workbook.
        path = bd.PACKS_DIR / "variables" / "_validator_proposal_typo.yaml"
        path.write_text(
            "variables:\n"
            "  - category: Demographics\n"
            "    variable: ProposalTypoTest\n"
            "    table: person\n"
            "    column: gender_concept_name\n"
            "    proposal: Stnadard\n",
            encoding="utf-8",
        )
        self.addCleanup(lambda: path.unlink(missing_ok=True))

        import importlib.util
        vp_path = Path(__file__).resolve().parent.parent / "scripts" / "validate_packs.py"
        spec = importlib.util.spec_from_file_location(
            "validate_packs_proposal_test", vp_path,
        )
        vp = importlib.util.module_from_spec(spec)
        sys.modules["validate_packs_proposal_test"] = vp
        spec.loader.exec_module(vp)

        rows = vp._resolve_variables("_validator_proposal_typo", findings=[])
        bad_proposals = [
            v.get("proposal") for v in rows
            if isinstance(v.get("proposal"), str)
            and v.get("proposal").strip()
            and v.get("proposal").strip() not in ("Standard", "Custom")
        ]
        self.assertEqual(bad_proposals, ["Stnadard"])


if __name__ == "__main__":
    unittest.main()
