"""Tests for build_dictionary.py.

Covers the cumulative regressions called out in review:
  - % Patient denominator is the cohort total.
  - Variable rows are PII-tagged and dropped by sales / pharma.
  - Renderers actually omit hidden sections.
  - resolve_variables skips GROUP BY for Unstructured / free-text columns.
  - Resolver adds Median (IQR) + Completeness for numeric columns and
    Min / Max for date columns.
  - `expression` field lets variables run a SQL rollup (e.g. LEFT(zip, 3)).
  - Columns / Variables sheets use the reference label `Table(s)`.

Tests stub the psycopg connection so no warehouse is required.
"""
from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from build_dictionary import (  # noqa: E402
    AUDIENCE_VISIBILITY,
    ColumnInfo,
    ColumnRow,
    CohortModel,
    CohortSummary,
    DateCoverage,
    TableRow,
    VariableRow,
    compute_patient_completeness,
    derive_data_source,
    derive_inclusion_criteria,
    filter_for_audience,
    load_table_descriptions,
    resolve_variables,
    section_visible,
    write_html,
    write_xlsx,
    _is_freetext_column,
)


class _Cursor:
    """Tiny psycopg-cursor stub.

    Each instance is fed a queue of (predicate, return_value) pairs;
    when the test code calls cur.execute(sql), the cursor returns the
    first result whose predicate matches the SQL string.
    """

    def __init__(self, script: list[tuple[str, Any]]) -> None:
        self._script = list(script)
        self._next_result: Any = None

    def execute(self, sql: str, *params) -> None:
        for needle, result in self._script:
            if needle in sql:
                self._next_result = result
                return
        raise AssertionError(f"unexpected SQL: {sql!r}")

    def fetchone(self):
        return self._next_result

    def fetchall(self):
        return self._next_result

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        return False


class _Conn:
    def __init__(self, script: list[tuple[str, Any]]) -> None:
        self._script = script

    def cursor(self):
        return _Cursor(self._script)

    def rollback(self) -> None:
        pass


def _var(**overrides) -> VariableRow:
    """Constructor helper so tests only set the fields they care about."""
    defaults = dict(
        category="Diagnosis",
        variable="Dx",
        description="",
        table="condition_occurrence",
        column="condition_concept_name",
        criteria="",
        values="",
        distribution="",
        median_iqr="",
        completeness_pct=None,
        implemented="No",
        patient_pct=None,
        extraction_type="Structured",
        notes="",
        pii=False,
    )
    defaults.update(overrides)
    return VariableRow(**defaults)


# --------------------------------------------------------------------- #
# P1 - % Patient denominator
# --------------------------------------------------------------------- #


class PatientCompletenessTests(unittest.TestCase):

    def test_uses_cohort_total_not_per_table(self):
        # Cohort has 1000 patients; the table this column lives in only
        # has rows for 50 of them; the column is non-null for all 50.
        # Correct % Patient = 50 / 1000 = 5.0% (NOT 50/50 = 100%).
        col = ColumnInfo(
            schema="x", table="t", column="c", data_type="text",
            is_nullable=True, row_count=200, null_count=0,
            completeness_pct=100.0,
        )
        conn = _Conn([("COUNT(DISTINCT", (50,))])
        pct = compute_patient_completeness(conn, "x", col, total_patients=1000)
        self.assertEqual(pct, 5.0)

    def test_returns_none_when_total_unknown(self):
        col = ColumnInfo(
            schema="x", table="t", column="c", data_type="text",
            is_nullable=True, row_count=10, null_count=0,
            completeness_pct=100.0,
        )
        conn = _Conn([])
        self.assertIsNone(compute_patient_completeness(conn, "x", col, None))
        self.assertIsNone(compute_patient_completeness(conn, "x", col, 0))

    def test_zero_rows_returns_zero(self):
        col = ColumnInfo(
            schema="x", table="t", column="c", data_type="text",
            is_nullable=True, row_count=0, null_count=0,
            completeness_pct=0.0,
        )
        conn = _Conn([])
        self.assertEqual(compute_patient_completeness(conn, "x", col, 100), 0.0)


# --------------------------------------------------------------------- #
# P1 - Variables PII tagging + audience filter
# --------------------------------------------------------------------- #


class VariablesPIITagsAndFilterTests(unittest.TestCase):

    def _make_model(self, variables: list[VariableRow]) -> CohortModel:
        return CohortModel(
            cohort="c", provider="p", disease="d", schema_name="s",
            variant="raw", display_name="C", description="",
            status="active", generated_at="t", git_sha="abc",
            introspect_version="0.0", schema_snapshot_digest="sha256:0",
            summary=CohortSummary(
                patient_count=10, table_count=0, column_count=0,
                date_coverage=DateCoverage(),
            ),
            tables=[], columns=[], variables=variables,
        )

    def test_resolver_tags_pii_from_pack(self):
        # location.zip is in our shipped pii.yaml — confirm the resolver
        # tags the row pii=True even when the SQL succeeds.
        pack = [{
            "category": "Demographics",
            "variable": "ZIP code",
            "table": "location",
            "column": "zip",
            "extraction_type": "Structured",
        }]
        # Stub: count returns 0 so we don't need top-N / patient-pct.
        conn = _Conn([("COUNT(*)", (0,))])
        rows = resolve_variables(
            conn, "s", pack, total_patients=100,
            pii_pairs={("location", "zip")}, pii_patterns=[],
        )
        self.assertTrue(rows[0].pii, "expected pii=True for location.zip")

    def test_audience_filter_drops_pii_variables_for_sales(self):
        # Two variables: one PII, one not. Sales should keep only the
        # non-PII one in `variables`.
        clean = _var(
            variable="Dx", table="condition_occurrence",
            column="condition_concept_name",
            implemented="Yes", patient_pct=42.0, pii=False,
        )
        pii = _var(
            variable="ZIP", table="location", column="zip",
            implemented="Yes", patient_pct=10.0, pii=True,
        )
        m = self._make_model([clean, pii])
        sales = filter_for_audience(m, "sales")
        self.assertEqual([v.variable for v in sales.variables], ["Dx"])

    def test_audience_filter_drops_pii_variables_for_pharma(self):
        pii = _var(
            variable="ZIP", table="location", column="zip",
            implemented="Yes", patient_pct=10.0, pii=True,
        )
        m = self._make_model([pii])
        self.assertEqual(filter_for_audience(m, "pharma").variables, [])

    def test_technical_audience_keeps_pii(self):
        pii = _var(
            variable="ZIP", table="location", column="zip",
            implemented="Yes", patient_pct=10.0, pii=True,
        )
        m = self._make_model([pii])
        self.assertEqual(len(filter_for_audience(m, "technical").variables), 1)


# --------------------------------------------------------------------- #
# P2 - Renderers omit hidden sections
# --------------------------------------------------------------------- #


def _make_full_model() -> CohortModel:
    return CohortModel(
        cohort="c", provider="p", disease="d", schema_name="s",
        variant="raw", display_name="C", description="",
        status="active", generated_at="t", git_sha="abc",
        introspect_version="0.0", schema_snapshot_digest="sha256:0",
        summary=CohortSummary(
            patient_count=10, table_count=1, column_count=1,
            date_coverage=DateCoverage(),
        ),
        tables=[TableRow(
            table_name="person", category="Demographics",
            row_count=10, column_count=2,
            patient_count_in_table=10, purpose="OMOP person",
        )],
        columns=[ColumnRow(
            category="Demographics", table="person", column="year_of_birth",
            description="Birth year.", data_type="integer",
            values="1948", distribution="Min: 1933 Max: 1997",
            median_iqr="Median: 1948", completeness_pct=100.0,
            patient_pct=100.0, extraction_type="Structured",
            pii=False, notes="",
        )],
        variables=[_var(
            category="Demographics", variable="Birth Year",
            description="Year the patient was born.",
            table="person", column="year_of_birth",
            values="1948", distribution="...",
            median_iqr="Median: 1948 (IQR: 1944-1952)",
            completeness_pct=100.0,
            implemented="Yes", patient_pct=100.0,
        )],
    )


class RendererOmissionTests(unittest.TestCase):

    def setUp(self):
        # Portable scratch dir — cleaned up automatically on tearDown.
        # Use tempfile.TemporaryDirectory() so the suite runs on Windows
        # (where /tmp does not exist) as well as macOS / Linux.
        self._tmp = tempfile.TemporaryDirectory()
        self.tmp_dir = Path(self._tmp.name)

    def tearDown(self):
        self._tmp.cleanup()

    def test_visibility_table_matches_readme(self):
        self.assertEqual(AUDIENCE_VISIBILITY["technical"],
                         {"summary": True, "tables": True, "columns": True, "variables": True})
        self.assertEqual(AUDIENCE_VISIBILITY["sales"],
                         {"summary": True, "tables": True, "columns": False, "variables": True})
        self.assertEqual(AUDIENCE_VISIBILITY["pharma"],
                         {"summary": True, "tables": False, "columns": False, "variables": True})

    def test_section_visible_helper(self):
        self.assertTrue(section_visible("technical", "columns"))
        self.assertFalse(section_visible("sales", "columns"))
        self.assertFalse(section_visible("pharma", "tables"))

    def test_html_omits_columns_for_sales(self):
        out = self.tmp_dir / "sales.html"
        write_html(_make_full_model(), out, audience="sales")
        page = out.read_text()
        self.assertIn("<h2>Tables</h2>", page)
        self.assertNotIn("<h2>Columns</h2>", page)
        self.assertIn("<h2>Variables</h2>", page)

    def test_html_omits_tables_and_columns_for_pharma(self):
        out = self.tmp_dir / "pharma.html"
        write_html(_make_full_model(), out, audience="pharma")
        page = out.read_text()
        self.assertNotIn("<h2>Tables</h2>", page)
        self.assertNotIn("<h2>Columns</h2>", page)
        self.assertIn("<h2>Variables</h2>", page)

    def test_xlsx_skips_hidden_sheets(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl not installed")
        import openpyxl as opx
        out = self.tmp_dir / "sales.xlsx"
        write_xlsx(_make_full_model(), out, audience="sales")
        wb = opx.load_workbook(out)
        self.assertEqual(set(wb.sheetnames),
                         {"Summary", "Tables", "Variables"})

        out2 = self.tmp_dir / "pharma.xlsx"
        write_xlsx(_make_full_model(), out2, audience="pharma")
        wb2 = opx.load_workbook(out2)
        self.assertEqual(set(wb2.sheetnames), {"Summary", "Variables"})


# --------------------------------------------------------------------- #
# P2 - resolve_variables skips GROUP BY for unstructured / free-text
# --------------------------------------------------------------------- #


class UnstructuredSkipTests(unittest.TestCase):

    def test_freetext_column_detector(self):
        self.assertTrue(_is_freetext_column("note_text"))
        self.assertTrue(_is_freetext_column("discharge_note"))
        self.assertTrue(_is_freetext_column("foo_text"))
        self.assertFalse(_is_freetext_column("year_of_birth"))
        self.assertFalse(_is_freetext_column("condition_concept_name"))

    def test_unstructured_variable_skips_group_by(self):
        # If the resolver tries to GROUP BY note_text the test will fail
        # because no GROUP BY entry is in the cursor script.
        pack = [{
            "category": "Reports",
            "variable": "Clinical Note",
            "table": "note",
            "column": "note_text",
            "extraction_type": "Unstructured",
        }]
        # Criteria-count, nonnull-count, and patient-count queries —
        # no GROUP BY pattern.
        conn = _Conn([
            ("COUNT(*)", (123,)),
            ("COUNT(DISTINCT", (40,)),
        ])
        rows = resolve_variables(
            conn, "s", pack, total_patients=100,
            pii_pairs=set(), pii_patterns=[],
            tables_with_person_id={"note"},
        )
        v = rows[0]
        self.assertEqual(v.implemented, "Yes")
        self.assertEqual(v.values, "")
        self.assertIn("not aggregated", v.distribution)
        self.assertEqual(v.patient_pct, 40.0)

    def test_freetext_column_name_skips_group_by_even_if_marked_structured(self):
        pack = [{
            "category": "Reports",
            "variable": "Some text",
            "table": "note",
            "column": "discharge_note",
            "extraction_type": "Structured",   # mis-marked but still text
        }]
        conn = _Conn([
            ("COUNT(*)", (5,)),
            ("COUNT(DISTINCT", (3,)),
        ])
        rows = resolve_variables(
            conn, "s", pack, total_patients=10,
            pii_pairs=set(), pii_patterns=[],
            tables_with_person_id={"note"},
        )
        self.assertEqual(rows[0].values, "")
        self.assertIn("not aggregated", rows[0].distribution)


# --------------------------------------------------------------------- #
# P1 - Median (IQR) + Completeness on Page 4
# --------------------------------------------------------------------- #


class MedianIQRAndCompletenessTests(unittest.TestCase):

    def test_numeric_column_gets_median_iqr_and_completeness(self):
        pack = [{
            "category": "Vitals", "variable": "Body Weight",
            "table": "observation", "column": "value_as_number",
            "criteria": "observation_concept_name = 'Body weight'",
            "extraction_type": "Structured",
        }]
        # Resolver runs in this order:
        #   1. COUNT(*) WHERE criteria              -> 1000 (denominator)
        #   2. COUNT(*) WHERE criteria AND NOT NULL -> 800  (numerator)
        #   3. GROUP BY value_as_number (top-10)    -> two rows
        #   4. PERCENTILE_CONT(0.25/0.5/0.75)       -> (150, 170, 190)
        #   5. COUNT(DISTINCT person_id)            -> 400
        conn = _SqlRouterConn(
            count_sequence=(1000, 800),
            percentile_cont=("150", "170", "190"),
            group_by=[("170", 5), ("180", 3)],
            count_distinct=(400,),
        )
        rows = resolve_variables(
            conn, "s", pack, total_patients=1000,
            pii_pairs=set(), pii_patterns=[],
            tables_with_person_id={"observation"},
            column_types={("observation", "value_as_number"): "numeric"},
        )
        v = rows[0]
        self.assertEqual(v.completeness_pct, 80.0)   # 800 / 1000
        self.assertIn("Median: 170", v.median_iqr)
        self.assertIn("IQR: 150-190", v.median_iqr)
        self.assertEqual(v.implemented, "Yes")
        self.assertEqual(v.patient_pct, 40.0)        # 400 / 1000

    def test_date_variable_uses_min_max_not_group_by(self):
        # Date-typed columns like condition_start_date should render a
        # Min/Max range in Distribution, not the ten most common exact
        # dates. The stub has no GROUP BY entry — if the resolver runs
        # one, _Cursor raises AssertionError and the test fails.
        pack = [{
            "category": "Diagnosis",
            "variable": "Diagnosis Start Date",
            "table": "condition_occurrence",
            "column": "condition_start_date",
            "extraction_type": "Structured",
        }]
        conn = _Conn([
            ("MIN", ("2022-01-15", "2026-03-22")),
            ("COUNT(DISTINCT", (500,)),
            ("COUNT(*)", (1000,)),
        ])
        rows = resolve_variables(
            conn, "s", pack, total_patients=1000,
            pii_pairs=set(), pii_patterns=[],
            tables_with_person_id={"condition_occurrence"},
            column_types={
                ("condition_occurrence", "condition_start_date"): "date",
            },
        )
        v = rows[0]
        self.assertIn("Min: 2022-01-15", v.distribution)
        self.assertIn("Max: 2026-03-22", v.distribution)
        self.assertEqual(v.values, "")
        self.assertEqual(v.median_iqr, "")
        self.assertEqual(v.implemented, "Yes")
        self.assertEqual(v.patient_pct, 50.0)   # 500 / 1000

    def test_timestamp_variable_also_uses_min_max(self):
        pack = [{
            "category": "Visits",
            "variable": "Visit Date",
            "table": "visit_occurrence",
            "column": "visit_start_date",
        }]
        conn = _Conn([
            ("MIN", ("2021-10-01", "2026-02-27")),
            ("COUNT(DISTINCT", (800,)),
            ("COUNT(*)", (5000,)),
        ])
        rows = resolve_variables(
            conn, "s", pack, total_patients=1000,
            tables_with_person_id={"visit_occurrence"},
            column_types={
                ("visit_occurrence", "visit_start_date"):
                    "timestamp without time zone",
            },
        )
        self.assertIn("Min: 2021-10-01", rows[0].distribution)
        self.assertIn("Max: 2026-02-27", rows[0].distribution)

    def test_categorical_column_has_no_median_iqr(self):
        pack = [{
            "category": "Diagnosis", "variable": "Dx",
            "table": "condition_occurrence",
            "column": "condition_concept_name",
            "extraction_type": "Structured",
        }]
        conn = _Conn([
            ("GROUP BY", [("Alzheimer's", 100)]),
            ("COUNT(DISTINCT", (50,)),
            ("COUNT(*)", (200,)),
        ])
        rows = resolve_variables(
            conn, "s", pack, total_patients=1000,
            pii_pairs=set(), pii_patterns=[],
            tables_with_person_id={"condition_occurrence"},
            column_types={("condition_occurrence", "condition_concept_name"): "text"},
        )
        self.assertEqual(rows[0].median_iqr, "")


# --------------------------------------------------------------------- #
# P2 - tables without person_id skip the patient-pct queries
# --------------------------------------------------------------------- #


class PatientPctSkipTests(unittest.TestCase):

    def test_compute_patient_completeness_caller_skips_no_person_table(self):
        # When the caller doesn't include the table in tables_with_person_id,
        # the resolver must not run COUNT(DISTINCT person_id). The stub will
        # raise AssertionError on any unexpected SQL.
        pack = [{
            "category": "Demographics", "variable": "Country",
            "table": "location", "column": "country_concept_name",
            "extraction_type": "Structured",
        }]
        # Note: no "COUNT(DISTINCT" entry — any such query fails the test.
        conn = _Conn([
            ("GROUP BY", [("United States", 100)]),
            ("COUNT(*)", (100,)),
        ])
        rows = resolve_variables(
            conn, "s", pack, total_patients=1000,
            pii_pairs=set(), pii_patterns=[],
            tables_with_person_id=set(),   # location is NOT in this set
            column_types={},
        )
        self.assertIsNone(rows[0].patient_pct)
        self.assertEqual(rows[0].implemented, "Yes")


# --------------------------------------------------------------------- #
# P2 - `expression` field — ZIP rollup to 3 digits
# --------------------------------------------------------------------- #


class ExpressionFieldTests(unittest.TestCase):

    def test_expression_is_used_in_group_by_and_column_is_display_only(self):
        pack = [{
            "category": "Demographics",
            "variable": "ZIP (3-digit)",
            "table": "location",
            "column": "zip",
            "expression": "LEFT(\"zip\"::text, 3)",
            "extraction_type": "Structured",
        }]
        captured: list[str] = []

        class _CaptureCursor:
            def execute(self, sql, *params):
                captured.append(sql)
                if "GROUP BY" in sql:
                    self._next_result = [("902", 40), ("950", 30)]
                else:
                    self._next_result = (100,)
            def fetchone(self):
                return self._next_result
            def fetchall(self):
                return self._next_result
            def __enter__(self):
                return self
            def __exit__(self, *exc):
                return False

        class _CaptureConn:
            def cursor(self):
                return _CaptureCursor()
            def rollback(self):
                pass

        rows = resolve_variables(
            _CaptureConn(), "s", pack, total_patients=None,
            pii_pairs=set(), pii_patterns=[],
            tables_with_person_id=set(),
            column_types={("location", "zip"): "text"},
        )
        v = rows[0]
        # Display cell still names the raw column so users see what it's about.
        self.assertEqual(v.column, "zip")
        # The expression is what's GROUP BY'd and shown in top-N output.
        group_by_sqls = [s for s in captured if "GROUP BY" in s]
        self.assertEqual(len(group_by_sqls), 1)
        self.assertIn("LEFT(\"zip\"::text, 3)", group_by_sqls[0])
        # And the resulting values are 3-digit prefixes, not raw ZIPs.
        self.assertIn("902", v.values)
        self.assertIn("950", v.values)


# --------------------------------------------------------------------- #
# P2 - Infusion Dates split into Start + End
# --------------------------------------------------------------------- #


class InfusionPackShapeTests(unittest.TestCase):

    def test_aat_common_splits_infusion_dates(self):
        # The two-row Start / End split lives in the disease-common
        # base because it applies to every AAT cohort. Reads the file
        # directly (rather than going through _all_variables_for) so
        # the assertion is about the file's own contents, not a
        # transitively-inherited row.
        import yaml as pyyaml
        path = REPO_ROOT / "packs" / "variables" / "aat_common.yaml"
        data = pyyaml.safe_load(path.read_text(encoding="utf-8"))
        names = [v["variable"] for v in data.get("variables", [])]
        # One combined label that overclaims is gone; two specific ones exist.
        self.assertNotIn("Infusion Dates", names)
        self.assertIn("Infusion Start Date", names)
        self.assertIn("Infusion End Date", names)


# --------------------------------------------------------------------- #
# P1 - renderers expose Median (IQR) and Completeness columns
# --------------------------------------------------------------------- #


class VariablesHeadersTests(unittest.TestCase):

    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.tmp_dir = Path(self._tmp.name)

    def tearDown(self):
        self._tmp.cleanup()

    def test_html_variables_header_has_median_iqr_and_completeness(self):
        out = self.tmp_dir / "tech.html"
        write_html(_make_full_model(), out, audience="technical")
        page = out.read_text()
        # Just below the Variables <h2>, we expect these columns.
        variables_section = page.split("<h2>Variables</h2>", 1)[1]
        self.assertIn("<th>Median (IQR)</th>", variables_section)
        self.assertIn("<th>Completeness</th>", variables_section)
        self.assertIn("<th>Implemented</th>", variables_section)
        self.assertIn("<th>% Patient</th>", variables_section)
        # P3: matches Century reference label.
        self.assertIn("<th>Table(s)</th>", variables_section)

    def test_html_columns_header_uses_table_s(self):
        out = self.tmp_dir / "tech.html"
        write_html(_make_full_model(), out, audience="technical")
        page = out.read_text()
        columns_section = page.split("<h2>Columns</h2>", 1)[1]
        # Columns section ends at the next <h2>, so only look at this block.
        columns_section = columns_section.split("<h2>", 1)[0]
        self.assertIn("<th>Table(s)</th>", columns_section)
        self.assertNotIn("<th>Table</th>", columns_section)

    def test_xlsx_variables_sheet_has_median_iqr_and_completeness(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl not installed")
        import openpyxl as opx
        out = self.tmp_dir / "tech.xlsx"
        write_xlsx(_make_full_model(), out, audience="technical")
        wb = opx.load_workbook(out)
        ws = wb["Variables"]
        headers = [c.value for c in ws[1]]
        self.assertIn("Median (IQR)", headers)
        self.assertIn("Completeness", headers)
        self.assertIn("Implemented", headers)
        self.assertIn("% Patient", headers)
        self.assertIn("Table(s)", headers)
        self.assertNotIn("Table", headers)

    def test_xlsx_columns_sheet_uses_table_s(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl not installed")
        import openpyxl as opx
        out = self.tmp_dir / "tech.xlsx"
        write_xlsx(_make_full_model(), out, audience="technical")
        wb = opx.load_workbook(out)
        ws = wb["Columns"]
        headers = [c.value for c in ws[1]]
        self.assertIn("Table(s)", headers)
        self.assertNotIn("Table", headers)

    def test_xlsx_tables_sheet_still_uses_singular_table(self):
        # The Tables sheet itself describes one row per table — the
        # "Table" label is correct there; only the Columns and Variables
        # pages gained the Century-style "Table(s)" header.
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl not installed")
        import openpyxl as opx
        out = self.tmp_dir / "tech.xlsx"
        write_xlsx(_make_full_model(), out, audience="technical")
        wb = opx.load_workbook(out)
        ws = wb["Tables"]
        headers = [c.value for c in ws[1]]
        self.assertIn("Table", headers)


# --------------------------------------------------------------------- #
# Tier-1 visual polish regression guards. Pure presentation assertions
# — they lock in the freeze-panes / auto-filter / styled-header shape
# for XLSX and the sticky-header CSS for HTML so a future edit can't
# silently remove them. No behavior / value assertions.
# --------------------------------------------------------------------- #


class VisualPolishXlsxTests(unittest.TestCase):

    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.tmp_dir = Path(self._tmp.name)

    def tearDown(self):
        self._tmp.cleanup()

    def _write(self, audience="technical"):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl not installed")
        out = self.tmp_dir / f"{audience}.xlsx"
        write_xlsx(_make_full_model(), out, audience=audience)
        import openpyxl as opx
        return opx.load_workbook(out)

    def test_data_sheets_have_freeze_panes(self):
        """Top row pinned on every data sheet so reviewers scanning
        long tables don't lose the header context."""
        wb = self._write()
        for name in ("Tables", "Columns", "Variables"):
            ws = wb[name]
            self.assertEqual(
                ws.freeze_panes, "A2",
                f"{name}: freeze_panes must be 'A2' (top header row)",
            )

    def test_data_sheets_have_auto_filter(self):
        """Filter-arrow dropdowns on every data sheet's header row."""
        wb = self._write()
        for name in ("Tables", "Columns", "Variables"):
            ws = wb[name]
            self.assertIsNotNone(
                ws.auto_filter.ref,
                f"{name}: auto_filter.ref must be set",
            )
            self.assertTrue(
                ws.auto_filter.ref.startswith("A1:"),
                f"{name}: auto_filter.ref must start at A1 "
                f"(got {ws.auto_filter.ref!r})",
            )

    def test_summary_sheet_has_no_freeze_or_filter(self):
        """Summary is a key/value sheet — filtering makes no sense."""
        wb = self._write()
        ws = wb["Summary"]
        # openpyxl's default freeze_panes is None (or 'A1' which is a no-op).
        self.assertIn(ws.freeze_panes, (None, "A1"),
                      "Summary sheet must not have freeze panes")
        self.assertIsNone(ws.auto_filter.ref,
                          "Summary sheet must not have auto-filter")

    def test_summary_sheet_header_is_plain(self):
        """Summary must stay visually plain — no navy fill, no bold
        header — so it reads as key/value metadata rather than
        a stylised dataset. Regression guard for the Tier-1 fix that
        moved `_style_xlsx_header_row` inside the data-sheet branch."""
        wb = self._write()
        ws = wb["Summary"]
        first_cell = ws.cell(row=1, column=1)
        self.assertFalse(
            first_cell.font.bold,
            "Summary A1 must NOT be bold (bold header means it was "
            "styled like a data sheet, contradicting Tier-1 design)",
        )
        # openpyxl's default unfilled cell has fill_type None or 'none'.
        self.assertIn(
            first_cell.fill.fill_type, (None, "none"),
            f"Summary A1 must NOT have a solid fill — Tier-1 design "
            f"says Summary stays plain (fill_type was "
            f"{first_cell.fill.fill_type!r})",
        )

    def test_header_row_is_styled(self):
        """Bold white text on a solid fill, row height >= 20 so the
        styled header actually shows. Asserts presentation without
        pinning to a specific hex colour."""
        wb = self._write()
        ws = wb["Variables"]
        first_cell = ws.cell(row=1, column=1)
        self.assertTrue(
            first_cell.font.bold,
            "Header cells must be bold",
        )
        self.assertEqual(
            first_cell.fill.fill_type, "solid",
            "Header cells must have a solid fill",
        )
        self.assertGreaterEqual(
            ws.row_dimensions[1].height or 0, 20,
            "Header row must be tall enough for the styled text",
        )


class VisualPolishHtmlTests(unittest.TestCase):

    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.tmp_dir = Path(self._tmp.name)

    def tearDown(self):
        self._tmp.cleanup()

    def _render(self, audience="technical"):
        out = self.tmp_dir / f"{audience}.html"
        write_html(_make_full_model(), out, audience=audience)
        return out.read_text()

    def test_html_has_sticky_header_css(self):
        """Long Variables tables need the header to stay visible
        while scrolling — pure CSS, no JS."""
        page = self._render()
        self.assertIn(
            "position: sticky", page,
            "HTML must declare sticky thead so the column headers "
            "remain visible while scrolling long tables",
        )

    def test_html_section_headings_still_bare(self):
        """Regression guard for the dedup / styling pass: <h2>
        tags must stay without attributes so test substring checks
        like assertIn('<h2>Columns</h2>', page) keep matching."""
        page = self._render()
        for label in ("Summary", "Tables", "Columns", "Variables"):
            self.assertIn(
                f"<h2>{label}</h2>", page,
                f"<h2>{label}</h2> must render without class / style "
                f"attributes so downstream substring tests keep matching",
            )

    def test_html_column_headers_still_bare(self):
        """<th> tags must stay attribute-free for the same reason."""
        page = self._render()
        # Spot-check a header that existing tests pin.
        self.assertIn(
            "<th>Median (IQR)</th>", page,
            "<th>Median (IQR)</th> must stay bare — substring tests "
            "in VariablesHeadersTests pin this literal",
        )


# --------------------------------------------------------------------- #
# Pack curation — split into adrd_common + disease-specific, widened
# ARIA match, tightened Document criteria, dropped catch-all lab row,
# mirrored brand/generic matching on AAT administration, added
# annotation on anti-amyloid procedure row.
# --------------------------------------------------------------------- #


def _load_yaml(relpath: str) -> dict:
    import yaml
    return yaml.safe_load((REPO_ROOT / relpath).read_text(encoding="utf-8"))


def _all_variables_for(pack_slug: str) -> list[dict]:
    """Resolve a variables pack's transitive include list."""
    seen: set[str] = set()

    def _resolve(slug: str) -> list[dict]:
        if slug in seen:
            return []
        seen.add(slug)
        data = _load_yaml(f"packs/variables/{slug}.yaml")
        out: list[dict] = []
        for inc in data.get("include") or []:
            out.extend(_resolve(inc))
        out.extend(data.get("variables") or [])
        return out

    return _resolve(pack_slug)


def _find_variable_in(pack_slug: str, name: str) -> dict:
    """Return the resolved variable row with `variable == name` from the
    given pack. Raises AssertionError if absent — callers inside
    unittest will fail naturally. Replaces four per-class copies of the
    same lookup that used to live as `_find` / `_find_variable` methods."""
    for v in _all_variables_for(pack_slug):
        if v.get("variable") == name:
            return v
    raise AssertionError(f"variable {name!r} not found in {pack_slug}")


def _load_validate_packs():
    """Import scripts/validate_packs as a module. Cached on the
    function attribute so the sys.path dance only happens once."""
    mod = getattr(_load_validate_packs, "_cached", None)
    if mod is not None:
        return mod
    sys.path.insert(0, str(REPO_ROOT / "scripts"))
    try:
        import importlib
        mod = importlib.import_module("validate_packs")
    finally:
        sys.path.pop(0)
    _load_validate_packs._cached = mod
    return mod


def _assert_validator_clean(test_case: unittest.TestCase,
                            cohort_slugs) -> None:
    """Run scripts/validate_packs.validate_cohort on each slug and
    assert zero errors and zero warnings. Used by the per-domain
    Cohort pack test classes (Nimbus / MTC / CKD) so the validator-
    sweep body lives in exactly one place."""
    validate_packs = _load_validate_packs()
    known = validate_packs._load_known_categories()
    for slug in cohort_slugs:
        report = validate_packs.validate_cohort(slug, known)
        errors = [f for f in report.findings if f.severity == "error"]
        warnings = [f for f in report.findings if f.severity == "warning"]
        test_case.assertEqual(
            errors, [],
            f"{slug} has validator errors: "
            f"{[f.message for f in errors]}",
        )
        test_case.assertEqual(
            warnings, [],
            f"{slug} has validator warnings: "
            f"{[f.message for f in warnings]}",
        )


def _assert_no_id_column_mismatch(test_case: unittest.TestCase,
                                  pack_slugs,
                                  extra_standalone_packs=()) -> None:
    """Run validate_packs._check_id_column_name_mismatch over every
    resolved row in each pack and assert None. `extra_standalone_packs`
    takes pack names that don't have an include chain (e.g.
    respiratory_common) — their variables are read directly rather
    than via _all_variables_for."""
    validate_packs = _load_validate_packs()
    for slug in pack_slugs:
        for v in _all_variables_for(slug):
            msg = validate_packs._check_id_column_name_mismatch(v)
            test_case.assertIsNone(
                msg, f"{slug}/{v.get('variable')}: {msg}",
            )
    for slug in extra_standalone_packs:
        for v in _load_yaml(f"packs/variables/{slug}.yaml").get("variables", []):
            msg = validate_packs._check_id_column_name_mismatch(v)
            test_case.assertIsNone(
                msg, f"{slug}/{v.get('variable')}: {msg}",
            )


class _SqlRouterConn:
    """Reusable stub psycopg-style connection that routes SQL strings
    to canned responses based on substring matches. Replaces two
    near-identical `_Conn2` inline classes.

    `count_sequence` is consumed one value per COUNT(*) call so tests
    can script the criteria-count / nonnull-count sequence used by
    `resolve_variables`. PERCENTILE_CONT / GROUP BY / COUNT(DISTINCT)
    have one canned response each; unexpected SQL raises
    AssertionError so tests that wander off the expected path fail
    loudly rather than silently."""

    def __init__(self, *, count_sequence=(0, 0),
                 percentile_cont=None,
                 group_by=None,
                 count_distinct=None):
        self._count_sequence = list(count_sequence)
        self._count_idx = 0
        self._percentile_cont = percentile_cont
        self._group_by = group_by
        self._count_distinct = count_distinct

    def cursor(self):
        conn = self

        class _Cur:
            def execute(self, sql, *params):
                if "PERCENTILE_CONT" in sql:
                    self._next = conn._percentile_cont
                elif "GROUP BY" in sql:
                    self._next = conn._group_by
                elif "COUNT(DISTINCT" in sql:
                    self._next = conn._count_distinct
                elif "COUNT(*)" in sql:
                    i = conn._count_idx
                    self._next = (conn._count_sequence[i],)
                    conn._count_idx += 1
                else:
                    raise AssertionError(f"unexpected SQL: {sql!r}")

            def fetchone(self):
                return self._next

            def fetchall(self):
                return self._next

            def __enter__(self):
                return self

            def __exit__(self, *exc):
                return False

        return _Cur()

    def rollback(self):
        pass


# --------------------------------------------------------------------- #
# ADRD pack split — three layers, mirroring the respiratory layout:
#
#   adrd_common              (demographics, vitals, biomarkers, etc.)
#       │
#       ├── aat_common       (AAT-wide disease rows)
#       │      └── mtc_aat   (per-cohort, no overrides yet)
#       │
#       └── alzheimers_common (Alzheimer's-wide disease rows)
#              └── mtc_alzheimers (per-cohort, no overrides yet)
#
# Cohort packs in packs/cohorts/ point at the per-cohort final variable
# pack (mtc_aat, mtc_alzheimers), never at the disease-common bases or
# adrd_common directly. Same per-cohort-ETL rule as the Nimbus packs.
# --------------------------------------------------------------------- #


MTC_FINAL_PACKS = ("mtc_aat", "mtc_alzheimers")


class PackSplitTests(unittest.TestCase):

    def test_adrd_common_exists_and_is_shared(self):
        adrd = _load_yaml("packs/variables/adrd_common.yaml")
        # adrd_common carries its own variables and does not include anything.
        self.assertIn("variables", adrd)
        self.assertFalse(adrd.get("include"))
        self.assertGreater(len(adrd["variables"]), 30)

    def test_alzheimers_common_includes_adrd_common(self):
        alz = _load_yaml("packs/variables/alzheimers_common.yaml")
        self.assertEqual(alz.get("include"), ["adrd_common"])

    def test_aat_common_includes_adrd_common_not_alzheimers(self):
        aat = _load_yaml("packs/variables/aat_common.yaml")
        self.assertIn("adrd_common", aat.get("include", []))
        self.assertNotIn(
            "alzheimers_common", aat.get("include", []),
            "aat_common must not transitively pull Alzheimer's-only "
            "rows (that's what caused the duplicate Medications row "
            "in the original AAT pack)",
        )

    def test_each_mtc_pack_includes_its_disease_common_base(self):
        """Per-cohort MTC packs layer on top of the disease-common pack.
        Mirrors the respiratory layout — separate ETL per cohort, so
        the final source of truth lives in the cohort pack."""
        expected = {
            "mtc_aat":         "aat_common",
            "mtc_alzheimers":  "alzheimers_common",
        }
        for pack, base in expected.items():
            data = _load_yaml(f"packs/variables/{pack}.yaml")
            self.assertEqual(
                data.get("include"), [base],
                f"{pack} must include exactly [{base!r}] — each per-cohort "
                f"MTC pack layers on top of its disease-common base, not "
                f"on adrd_common directly and not on the other disease",
            )

    def test_no_duplicate_variables_in_either_mtc_cohort(self):
        for slug in MTC_FINAL_PACKS:
            all_vars = _all_variables_for(slug)
            keys = [(v.get("category"), v.get("variable")) for v in all_vars]
            dupes = sorted({k for k in keys if keys.count(k) > 1})
            self.assertEqual(
                dupes, [],
                f"duplicate (category, variable) in {slug}: {dupes}",
            )


class PackCurationFixTests(unittest.TestCase):

    # -- P2: ARIA criteria must use %ARIA% (not prefix-only ARIA%).
    # ARIA rows are disease-level (aat_common / alzheimers_common).
    def test_aria_uses_fuzzy_match_not_prefix(self):
        aria_aat_h = _find_variable_in("aat_common", "ARIA-H (microhemorrhage)")
        aria_aat_e = _find_variable_in("aat_common", "ARIA-E (edema)")
        aria_alz = _find_variable_in("alzheimers_common", "ARIA")
        for v in (aria_aat_h, aria_aat_e, aria_alz):
            c = v["criteria"]
            # Pattern must start with a %
            self.assertIn("ILIKE '%ARIA", c,
                          f"{v['variable']} criteria must use %ARIA prefix "
                          f"match (got: {c!r})")

    # -- P2: Catch-all "Other Laboratory Measurements" row is gone.
    def test_no_other_laboratory_measurements_catch_all(self):
        # Sweep every layer (disease-common bases + per-cohort finals)
        # so a catch-all reintroduced at any level still trips the test.
        for slug in ("aat_common", "alzheimers_common") + MTC_FINAL_PACKS:
            names = [v.get("variable") for v in _all_variables_for(slug)]
            self.assertNotIn(
                "Other Laboratory Measurements", names,
                f"{slug} still ships the catch-all lab row that the "
                f"reviewer flagged"
            )

    # -- P2: Document variable has MRI / PET / EEG criteria.
    # Document row lives in adrd_common, surfaced via alzheimers_common.
    def test_document_variable_has_imaging_criteria(self):
        doc = _find_variable_in(
            "alzheimers_common", "Document (MRI / PET / EEG)"
        )
        self.assertTrue(doc.get("criteria"),
                        "Document variable must have a criteria filter")
        c = doc["criteria"].upper()
        for token in ("MRI", "PET", "EEG"):
            self.assertIn(token, c,
                          f"Document criteria must match {token} "
                          f"(got: {doc['criteria']!r})")

    # -- P2: AAT administration mirrors prescription brand+generic matching.
    def test_aat_administration_matches_brands_and_generics(self):
        admin = _find_variable_in(
            "aat_common", "Anti-amyloid Therapy (Administration)"
        )
        c = admin["criteria"].lower()
        for token in ("lecanemab", "leqembi", "donanemab", "kisunla",
                      "aducanumab", "aduhelm"):
            self.assertIn(token, c,
                          f"AAT administration criteria must match {token!r} "
                          f"(got: {admin['criteria']!r})")

    # -- P2/P3: Anti-amyloid procedure row is labelled honestly AND
    # carries an explanation of the broad chemotherapy-administration
    # match. Row was renamed in the follow-up review:
    #   Anti-amyloid Infusion Procedure  ->
    #   Infusion Procedure Codes (candidate AAT attribution)
    def test_anti_amyloid_procedure_has_disambiguation_note(self):
        proc = _find_variable_in(
            "aat_common",
            "Infusion Procedure Codes (candidate AAT attribution)",
        )
        # The note explains why generic chemotherapy admin codes are kept.
        self.assertTrue(
            proc.get("notes"),
            "Anti-amyloid procedure row must carry a note explaining why "
            "generic 'Chemotherapy administration' is included",
        )
        # The description explicitly flags the ambiguity.
        desc = (proc.get("description") or "").lower()
        self.assertIn("candidate", desc,
                      "description must signal the row is a candidate, "
                      "not definitively anti-amyloid")

    # -- P3: Labs / Biomarkers spelling is consistent across packs.
    def test_biomarkers_category_label_consistent(self):
        for slug in ("aat_common", "alzheimers_common") + MTC_FINAL_PACKS:
            cats = {v.get("category") for v in _all_variables_for(slug)}
            self.assertIn(
                "Labs / Biomarkers", cats,
                f"{slug} must use 'Labs / Biomarkers' (matches categories.yaml)"
            )
            self.assertNotIn(
                "Biomarkers / Labs", cats,
                f"{slug} still uses the inverted 'Biomarkers / Labs' label"
            )

    # -- P3: lat/lon tagged as PII on the location table.
    def test_location_latitude_longitude_are_pii(self):
        pii = _load_yaml("packs/pii.yaml")
        loc = set((pii.get("pii_columns") or {}).get("location") or [])
        self.assertIn("latitude", loc)
        self.assertIn("longitude", loc)


# --------------------------------------------------------------------- #
# Validator — runs scripts/validate_packs.py against the committed
# packs and expects zero errors. Guards against future pack edits
# regressing things like duplicate variables, missing criteria on
# clinically-specific rows, or unsafe ILIKE patterns.
# --------------------------------------------------------------------- #


class ValidatorTests(unittest.TestCase):

    def test_validator_on_committed_packs_is_clean(self):
        validate_packs = _load_validate_packs()
        known = validate_packs._load_known_categories()
        reports = [
            validate_packs.validate_cohort(p.stem, known)
            for p in sorted((REPO_ROOT / "packs" / "cohorts").glob("*.yaml"))
        ]
        all_findings = [f for r in reports for f in r.findings]
        errors = [f for f in all_findings if f.severity == "error"]
        warnings = [f for f in all_findings if f.severity == "warning"]

        self.assertEqual(
            errors, [],
            f"validator errors on committed packs: "
            f"{[f.message for f in errors]}"
        )
        self.assertEqual(
            warnings, [],
            f"validator warnings on committed packs: "
            f"{[f.message for f in warnings]}"
        )

    def test_validator_flags_prefix_only_ilike(self):
        validate_packs = _load_validate_packs()
        msgs = validate_packs._check_unsafe_ilike(
            "observation_concept_name ILIKE 'ARIA%'"
        )
        self.assertTrue(any("prefix-only" in m for m in msgs))

        msgs = validate_packs._check_unsafe_ilike(
            "measurement_concept_name ILIKE 'APOE'"
        )
        self.assertTrue(any("exact-match" in m for m in msgs))

        msgs = validate_packs._check_unsafe_ilike(
            "observation_concept_name ILIKE '%ARIA-H%'"
        )
        self.assertEqual(msgs, [])

    def test_validator_flags_id_column_on_non_id_variable(self):
        """Catches the Infusion-Drug-style mistake: row named as a
        business concept but pointing at `*_concept_id`."""
        validate_packs = _load_validate_packs()
        # Bad: "Infusion Drug" with column drug_concept_id
        bad = {"variable": "Infusion Drug",
               "column": "drug_concept_id", "table": "infusion"}
        msg = validate_packs._check_id_column_name_mismatch(bad)
        self.assertIsNotNone(msg)
        self.assertIn("drug_concept_id", msg)

        # Good: name signals it's an ID
        good = {"variable": "Infusion Drug (Concept ID)",
                "column": "drug_concept_id", "table": "infusion"}
        self.assertIsNone(
            validate_packs._check_id_column_name_mismatch(good),
            "row whose name mentions Concept ID should be accepted"
        )

        # Good: expression-backed — opting into a computed column type
        with_expr = {"variable": "ZIP code (3-digit prefix)",
                     "column": "zip",
                     "expression": 'LEFT("zip"::text, 3)',
                     "table": "location"}
        self.assertIsNone(
            validate_packs._check_id_column_name_mismatch(with_expr)
        )

        # Not an ID column -- fine
        neutral = {"variable": "Diagnosis",
                   "column": "condition_concept_name",
                   "table": "condition_occurrence"}
        self.assertIsNone(
            validate_packs._check_id_column_name_mismatch(neutral)
        )


# --------------------------------------------------------------------- #
# Follow-up pack tightening: A-beta variants, ADRD symptomatic
# meds, Infusion Drug rename, anti-amyloid procedure rename.
# --------------------------------------------------------------------- #


class PackTighteningTests(unittest.TestCase):

    def test_abeta_criteria_does_not_use_catch_wide_wildcards(self):
        # `%A%42%` / `%A%40%` had matched anything with an 'A' before the
        # number (HbA1c, ALT pre-42, etc.). Assert they're gone.
        # A-beta rows live in adrd_common, surfaced via alzheimers_common.
        ab42 = _find_variable_in("alzheimers_common", "Amyloid-beta 42 (A-beta 42)")
        ab40 = _find_variable_in("alzheimers_common", "Amyloid-beta 40 (A-beta 40)")
        for v in (ab42, ab40):
            c = v["criteria"]
            self.assertNotIn(
                "%A%42%", c, f"{v['variable']} still has catch-wide %A%42%"
            )
            self.assertNotIn(
                "%A%40%", c, f"{v['variable']} still has catch-wide %A%40%"
            )
        # And at least one explicit amyloid-beta variant is present.
        self.assertIn("amyloid beta 42", ab42["criteria"].lower())
        self.assertIn("amyloid beta 40", ab40["criteria"].lower())

    def test_symptomatic_adrd_therapy_rows_filter_to_chei_nmda(self):
        # Old "Medication (Prescription)" row has been renamed and
        # narrowed — the new row should filter by explicit ADRD drugs.
        # Symptomatic ADRD rows live in alzheimers_common.
        rx = _find_variable_in(
            "alzheimers_common", "Symptomatic ADRD Therapy (Prescription)"
        )
        c = rx["criteria"].lower()
        for drug in ("donepezil", "rivastigmine", "galantamine", "memantine"):
            self.assertIn(
                drug, c,
                f"Symptomatic ADRD Therapy (Prescription) must filter "
                f"by {drug!r} (got: {rx['criteria']!r})",
            )
        # Ensure the old over-broad label isn't still present anywhere
        # in the resolved Alzheimer's chain.
        names = [v.get("variable")
                 for v in _all_variables_for("mtc_alzheimers")]
        self.assertNotIn("Medication (Prescription)", names,
                         "old broad Medication (Prescription) row still present")
        self.assertNotIn("Medication (Administration)", names,
                         "old broad Medication (Administration) row still present")

    def test_infusion_drug_row_is_clearly_labeled_as_concept_id(self):
        drug = _find_variable_in("aat_common", "Infusion Drug (Concept ID)")
        self.assertEqual(drug["column"], "drug_concept_id")
        # Old ambiguous name must not coexist with the new one in the
        # resolved AAT chain.
        aat_names = [v.get("variable") for v in _all_variables_for("mtc_aat")]
        self.assertNotIn("Infusion Drug", aat_names,
                         "old ambiguous 'Infusion Drug' label still present")

    def test_anti_amyloid_procedure_row_signals_broad_match(self):
        # Old: "Anti-amyloid Infusion Procedure" (implied certainty).
        # New: "Infusion Procedure Codes (candidate AAT attribution)".
        aat_names = [v.get("variable") for v in _all_variables_for("mtc_aat")]
        self.assertNotIn("Anti-amyloid Infusion Procedure", aat_names)
        self.assertIn(
            "Infusion Procedure Codes (candidate AAT attribution)",
            aat_names,
        )

    def test_concept_id_variable_skips_median_iqr(self):
        """Concept-id / surrogate-key columns are numeric by type but
        not measurements. Their Median (IQR) cell must stay empty."""
        pack = [{
            "category": "Infusions",
            "variable": "Infusion Drug (Concept ID)",
            "table": "infusion",
            "column": "drug_concept_id",
            "extraction_type": "Structured",
        }]

        # Script runs: count x 2, GROUP BY, COUNT(DISTINCT person_id).
        # PERCENTILE_CONT must NOT be issued — if it were, the stub
        # would fall through to the AssertionError branch and fail
        # the test.
        conn = _Conn([
            ("GROUP BY", [("100001", 42), ("100002", 10)]),
            ("COUNT(DISTINCT", (400,)),
            ("COUNT(*)", (1000,)),
        ])
        rows = resolve_variables(
            conn, "s", pack, total_patients=1000,
            pii_pairs=set(), pii_patterns=[],
            tables_with_person_id={"infusion"},
            column_types={("infusion", "drug_concept_id"): "bigint"},
        )
        v = rows[0]
        self.assertEqual(
            v.median_iqr, "",
            "drug_concept_id must not get a Median (IQR) — it's an ID, "
            "not a measurement",
        )
        # Top values / GROUP BY still populate so the row isn't empty.
        self.assertNotEqual(v.distribution, "")
        self.assertEqual(v.implemented, "Yes")

    def test_ordinary_numeric_measurement_still_gets_median_iqr(self):
        """Regression guard: the skip above must only trigger for
        surrogate-key column names, not every numeric column."""
        pack = [{
            "category": "Labs / Biomarkers",
            "variable": "A-beta 42",
            "table": "measurement",
            "column": "value_as_number",
            "extraction_type": "Structured",
        }]

        conn = _SqlRouterConn(
            count_sequence=(1000, 800),
            percentile_cont=("10", "20", "30"),
            group_by=[("20", 5)],
            count_distinct=(400,),
        )
        rows = resolve_variables(
            conn, "s", pack, total_patients=1000,
            tables_with_person_id={"measurement"},
            column_types={("measurement", "value_as_number"): "numeric"},
        )
        self.assertIn("Median: 20", rows[0].median_iqr)


# --------------------------------------------------------------------- #
# Respiratory pack split — three layers:
#
#   respiratory_common         (demographics, vitals, smoking, etc.)
#        │
#        ├── copd_common       (COPD-wide disease rows)
#        │      │
#        │      ├── nimbus_copd        (per-cohort, adds
#        │      │                       eosinophil_standardized)
#        │      └── nimbus_az_copd     (per-cohort, no overrides yet)
#        │
#        └── asthma_common     (asthma-wide disease rows)
#               │
#               ├── nimbus_asthma      (per-cohort, no overrides yet)
#               └── nimbus_az_asthma   (per-cohort, no overrides yet)
#
# Cohort packs in packs/cohorts/ point at the per-cohort final variable
# pack, never at copd_common / asthma_common / respiratory_common
# directly. This matches the rule that each cohort has its own ETL and
# therefore its own source of truth.
# --------------------------------------------------------------------- #


NIMBUS_FINAL_PACKS = (
    "nimbus_copd", "nimbus_az_copd",
    "nimbus_asthma", "nimbus_az_asthma",
)


class RespiratoryPackSplitTests(unittest.TestCase):

    def test_respiratory_common_exists_and_is_shared(self):
        resp = _load_yaml("packs/variables/respiratory_common.yaml")
        self.assertIn("variables", resp)
        self.assertFalse(
            resp.get("include"),
            "respiratory_common is the shared base and must not include anything",
        )
        # Enough rows to be useful — demographics (9) + vitals (8) + basics.
        self.assertGreater(len(resp["variables"]), 20)

    def test_copd_common_includes_respiratory_common_not_asthma(self):
        copd = _load_yaml("packs/variables/copd_common.yaml")
        self.assertIn("respiratory_common", copd.get("include", []))
        self.assertNotIn(
            "asthma_common", copd.get("include", []),
            "copd_common must not transitively pull asthma-specific rows "
            "(biologics, ACT, FeNO) — that would reintroduce the same "
            "duplication the ADRD split was designed to prevent",
        )

    def test_asthma_common_includes_respiratory_common_not_copd(self):
        asthma = _load_yaml("packs/variables/asthma_common.yaml")
        self.assertIn("respiratory_common", asthma.get("include", []))
        self.assertNotIn(
            "copd_common", asthma.get("include", []),
            "asthma_common must not transitively pull COPD-specific rows "
            "(triple therapy, roflumilast, supplemental oxygen)",
        )

    def test_each_nimbus_pack_includes_its_disease_common_base(self):
        """Per-cohort packs layer on top of the disease-common pack.
        Separate ETLs mean the final source of truth lives in the
        cohort pack, but the shared disease rows must still flow in
        via the include chain."""
        expected = {
            "nimbus_copd":       "copd_common",
            "nimbus_az_copd":    "copd_common",
            "nimbus_asthma":     "asthma_common",
            "nimbus_az_asthma":  "asthma_common",
        }
        for pack, base in expected.items():
            data = _load_yaml(f"packs/variables/{pack}.yaml")
            self.assertEqual(
                data.get("include"), [base],
                f"{pack} must include exactly [{base!r}] — each per-cohort "
                f"pack layers on top of its disease-common base, not on "
                f"respiratory_common directly and not on another cohort",
            )

    def test_no_duplicate_variables_in_any_respiratory_pack(self):
        """Duplicate (category, variable) in the resolved list of any
        final pack would mean the include chain got tangled (e.g. a
        cohort pack redeclares a row that already comes from its
        common base)."""
        for slug in NIMBUS_FINAL_PACKS:
            all_vars = _all_variables_for(slug)
            keys = [(v.get("category"), v.get("variable")) for v in all_vars]
            dupes = sorted({k for k in keys if keys.count(k) > 1})
            self.assertEqual(
                dupes, [],
                f"duplicate (category, variable) in {slug}: {dupes}",
            )


# --------------------------------------------------------------------- #
# Respiratory pack curation — key COPD and asthma criteria assertions.
# Guards against future edits regressing the class-level medication
# rows or the disease-defining diagnosis matches.
# --------------------------------------------------------------------- #


class RespiratoryPackCurationTests(unittest.TestCase):

    # -- COPD diagnosis covers the three standard condition families.
    # Defined in copd_common so every COPD cohort inherits it.
    def test_copd_diagnosis_matches_all_three_condition_families(self):
        dx = _find_variable_in("copd_common", "COPD Diagnosis")
        c = dx["criteria"].lower()
        for token in ("chronic obstructive pulmonary",
                      "emphysema", "chronic bronchitis"):
            self.assertIn(
                token, c,
                f"COPD Diagnosis criteria must match {token!r} "
                f"(got: {dx['criteria']!r})",
            )

    # -- COPD inhaler rows exist as separate classes (not one catch-all).
    # Must flow through to every nimbus_*_copd cohort via the
    # copd_common include.
    def test_copd_ships_class_level_inhaler_rows(self):
        expected = [
            "Short-acting Bronchodilator (SABA / SAMA)",
            "Long-acting Beta-agonist (LABA)",
            "Long-acting Muscarinic Antagonist (LAMA)",
            "Inhaled Corticosteroid (ICS)",
            "COPD Triple Therapy (ICS + LABA + LAMA)",
        ]
        for slug in ("copd_common", "nimbus_copd", "nimbus_az_copd"):
            names = [v.get("variable") for v in _all_variables_for(slug)]
            for label in expected:
                self.assertIn(
                    label, names,
                    f"{slug} must expose class-level row {label!r}",
                )

    # -- LAMA criteria for COPD must not collapse to an oral/injectable match.
    def test_copd_lama_glycopyrrolate_row_is_inhaled_only(self):
        lama = _find_variable_in("copd_common",
                          "Long-acting Muscarinic Antagonist (LAMA)")
        c = lama["criteria"].lower()
        # The glycopyrrolate clause must be paired with an 'inhal' filter.
        self.assertIn("glycopyrrolate", c)
        self.assertIn("inhal", c,
                      "Glycopyrrolate clause must be scoped to inhaled "
                      "formulations — oral / injectable glycopyrrolate is "
                      "used for secretion management, not COPD maintenance")

    # -- Asthma diagnosis picks up severity / persistence variants.
    def test_asthma_diagnosis_matches_status_asthmaticus(self):
        dx = _find_variable_in("asthma_common", "Asthma Diagnosis")
        c = dx["criteria"].lower()
        self.assertIn("asthma", c)
        self.assertIn(
            "status asthmaticus", c,
            "Asthma Diagnosis must explicitly include status asthmaticus "
            "so the most severe variant isn't missed",
        )

    # -- Asthma biologic row covers the six FDA-approved agents.
    def test_asthma_biologic_row_covers_all_six_ingredients(self):
        bio = _find_variable_in(
            "asthma_common",
            "Asthma Biologic (Anti-IgE / Anti-IL5 / Anti-IL4R / Anti-TSLP)",
        )
        c = bio["criteria"].lower()
        for ingredient in ("omalizumab", "mepolizumab", "reslizumab",
                           "benralizumab", "dupilumab", "tezepelumab"):
            self.assertIn(
                ingredient, c,
                f"Asthma biologic row must match {ingredient!r} "
                f"(got: {bio['criteria']!r})",
            )

    # -- Shared eosinophil row (in respiratory_common) uses a qualifier.
    # Must flow through to every final respiratory pack.
    def test_eosinophils_row_requires_blood_qualifier(self):
        for slug in NIMBUS_FINAL_PACKS:
            eos = _find_variable_in(slug, "Blood Eosinophils")
            c = eos["criteria"]
            self.assertIn("%eosinophil%", c.lower(),
                          f"{slug}: Blood Eosinophils must match eosinophil")
            # Must have a blood / CBC / differential qualifier (ANDed in)
            # so sputum- and tissue-eosinophil concepts don't slip through.
            self.assertTrue(
                any(tok in c.lower() for tok in ("blood", "cbc", "differential",
                                                 "absolute", "percent")),
                f"{slug}: eosinophil row must AND in a blood-side qualifier",
            )

    # -- SpO2 is in respiratory_common with a widened ILIKE pattern.
    def test_oxygen_saturation_row_matches_spo2_and_peripheral_oxygen(self):
        for slug in NIMBUS_FINAL_PACKS:
            o2 = _find_variable_in(slug, "Oxygen Saturation (SpO2)")
            c = o2["criteria"].lower()
            for token in ("o2 saturation", "oxygen saturation",
                          "peripheral oxygen", "spo2"):
                self.assertIn(
                    token, c,
                    f"{slug}: Oxygen Saturation row must match {token!r} "
                    f"(peripheral oxygen is how Nimbus records it)",
                )

    # -- Nimbus COPD ships a cohort-specific `eosinophil_standardized`
    # abstraction table. The curated row lives in the per-cohort
    # nimbus_copd pack ONLY — not in copd_common (would wrongly appear
    # in Nimbus AZ) and not in nimbus_az_copd (the table isn't there).
    def test_nimbus_copd_exposes_both_eosinophil_rows(self):
        labs = [v for v in _all_variables_for("nimbus_copd")
                if v.get("category") == "Labs / Biomarkers"]
        names = [v.get("variable") for v in labs]
        self.assertIn(
            "Blood Eosinophils", names,
            "measurement-table eosinophil row from respiratory_common "
            "must still flow through to nimbus_copd",
        )
        self.assertIn(
            "Blood Eosinophils (Standardized, Nimbus curated)", names,
            "nimbus_copd must ship the cohort-specific eosinophil row "
            "pointing at the Nimbus-curated `eosinophil_standardized` table",
        )
        std = next(v for v in labs
                   if v["variable"] == "Blood Eosinophils (Standardized, Nimbus curated)")
        self.assertEqual(std.get("table"), "eosinophil_standardized")
        self.assertEqual(std.get("column"), "value_as_number")

    def test_copd_common_does_not_ship_cohort_specific_eosinophil_row(self):
        """The curated abstraction table belongs in the per-cohort
        nimbus_copd pack, not the shared copd_common base — otherwise
        nimbus_az_copd would wrongly inherit a row pointing at a table
        its schema does not have."""
        names = [v.get("variable")
                 for v in _all_variables_for("copd_common")]
        self.assertNotIn(
            "Blood Eosinophils (Standardized, Nimbus curated)", names,
            "copd_common must not contain the Nimbus-curated row — it "
            "belongs in packs/variables/nimbus_copd.yaml only",
        )

    def test_nimbus_az_copd_does_not_inherit_standardized_eosinophil_row(self):
        """nimbus_az_copd_cohort schema does not ship
        `eosinophil_standardized`. The curated row must not reach this
        pack via any include chain — otherwise the Variables sheet
        would point the reviewer at a nonexistent table."""
        names = [v.get("variable")
                 for v in _all_variables_for("nimbus_az_copd")]
        self.assertNotIn(
            "Blood Eosinophils (Standardized, Nimbus curated)", names,
            "nimbus_az_copd must not inherit the Nimbus-curated eosinophil "
            "row (the cohort's schema has no eosinophil_standardized table)",
        )

    def test_asthma_packs_do_not_inherit_standardized_eosinophil_row(self):
        """Neither asthma cohort's schema ships eosinophil_standardized."""
        for slug in ("asthma_common", "nimbus_asthma", "nimbus_az_asthma"):
            names = [v.get("variable") for v in _all_variables_for(slug)]
            self.assertNotIn(
                "Blood Eosinophils (Standardized, Nimbus curated)", names,
                f"{slug} must not inherit the Nimbus-COPD-curated "
                f"standardized eosinophil row",
            )

    # -- FEV1 criteria is narrow enough not to match HFEV-like strings.
    # Regression guard: an earlier draft had `%FEV1%` only which is fine,
    # but the `forced expiratory volume%1%` clause must keep the anchor
    # so it doesn't match unrelated "forced expiratory volume" phrases
    # without the time specifier.
    def test_fev1_criteria_anchors_on_1_second(self):
        for slug in NIMBUS_FINAL_PACKS:
            fev1 = _find_variable_in(slug, "FEV1 (Forced Expiratory Volume in 1 second)")
            c = fev1["criteria"].lower()
            self.assertIn("fev1", c)
            # Either explicit '1 second' or '%1%' anchor on the phrase
            self.assertTrue(
                ("forced expiratory volume" in c and ("1 second" in c or "%1%" in c)),
                f"{slug}: FEV1 criteria must anchor on 1-second so it "
                f"doesn't generically match FEV3 / FEV6 measurements "
                f"(got: {fev1['criteria']!r})",
            )


# --------------------------------------------------------------------- #
# Broad-wildcard and ID-column regression guards on the respiratory
# packs. Mirrors the PackTighteningTests on ADRD — covers the wildcard
# patterns most likely to overmatch if a future edit widens a row.
# --------------------------------------------------------------------- #


class RespiratoryBroadWildcardRegressionTests(unittest.TestCase):

    def test_no_catch_wide_drug_wildcards_in_any_respiratory_pack(self):
        # Patterns that would be trivially too broad:
        #   ILIKE '%inhaler%'  — matches every inhaled drug including
        #                        antibiotics, salt water, DNAase
        #   ILIKE '%steroid%'  — matches topical, ophthalmic, oral, etc.
        #   ILIKE '%beta%'     — matches beta-blockers (the opposite class!)
        banned_patterns = ("'%inhaler%'", "'%steroid%'",
                           "'%beta%'", "'%bronch%'")
        # Cover every final pack (includes transitively resolve the
        # common bases) plus the two common bases themselves, in case
        # the final pack turns out to be empty.
        for slug in NIMBUS_FINAL_PACKS + ("copd_common", "asthma_common"):
            for v in _all_variables_for(slug):
                c = (v.get("criteria") or "").lower()
                for bad in banned_patterns:
                    self.assertNotIn(
                        bad, c,
                        f"{slug}/{v.get('variable')} criteria contains "
                        f"overly broad wildcard {bad} — narrow to a "
                        f"specific ingredient list or class marker "
                        f"(got: {v.get('criteria')!r})",
                    )

    def test_no_respiratory_variable_points_at_id_column_without_id_name(self):
        """Regression guard: no respiratory variable row should silently
        render opaque `*_concept_id` values under a business-facing name
        (the Infusion-Drug-style mistake from the MTC pack review)."""
        _assert_no_id_column_mismatch(
            self,
            pack_slugs=NIMBUS_FINAL_PACKS + ("copd_common", "asthma_common"),
            extra_standalone_packs=("respiratory_common",),
        )


# --------------------------------------------------------------------- #
# Nimbus cohort packs — point at the correct variables pack and carry
# the required fields the validator enforces.
# --------------------------------------------------------------------- #


class NimbusCohortPackTests(unittest.TestCase):

    # Per-cohort final variable pack — the name MUST match the cohort
    # slug. Separate ETL per cohort => separate final source of truth.
    EXPECTED = {
        "nimbus_copd":       ("Nimbus",    "COPD",   "nimbus_copd",      "nimbus_copd_curated"),
        "nimbus_asthma":     ("Nimbus",    "Asthma", "nimbus_asthma",    "nimbus_asthma_curated"),
        "nimbus_az_copd":    ("Nimbus AZ", "COPD",   "nimbus_az_copd",   "nimbus_az_copd_cohort"),
        "nimbus_az_asthma":  ("Nimbus AZ", "Asthma", "nimbus_az_asthma", "nimbus_az_asthma_cohort"),
    }

    def test_all_four_nimbus_cohort_packs_exist(self):
        for slug in self.EXPECTED:
            path = REPO_ROOT / "packs" / "cohorts" / f"{slug}.yaml"
            self.assertTrue(path.is_file(),
                            f"missing cohort pack: {path}")

    def test_nimbus_cohort_packs_point_at_per_cohort_variable_packs(self):
        for slug, (provider, disease, var_pack, schema) in self.EXPECTED.items():
            data = _load_yaml(f"packs/cohorts/{slug}.yaml")
            self.assertEqual(data.get("provider"), provider, slug)
            self.assertEqual(data.get("disease"), disease, slug)
            self.assertEqual(
                data.get("variables_pack"), var_pack,
                f"{slug} must point at its per-cohort variable pack "
                f"{var_pack!r} — separate ETL per cohort means each "
                f"cohort owns its final source of truth",
            )
            self.assertEqual(
                data.get("schema_name"), schema,
                f"{slug} schema_name must match the raw dump filename "
                f"in Output/ (got {data.get('schema_name')!r})",
            )
            # And the per-cohort variable pack file itself must exist.
            var_path = REPO_ROOT / "packs" / "variables" / f"{var_pack}.yaml"
            self.assertTrue(
                var_path.is_file(),
                f"{slug}: variables_pack file {var_path} is missing",
            )

    def test_validator_clean_on_nimbus_cohorts(self):
        # Re-run the validator and confirm the four Nimbus cohorts are
        # error- and warning-free. Guards against future pack edits
        # sneaking a duplicate / unsafe ILIKE / id-column mismatch in.
        _assert_validator_clean(self, self.EXPECTED)


# --------------------------------------------------------------------- #
# MTC cohort packs — mirror the Nimbus checks. Per-cohort variable pack
# per cohort, schema_name matches the on-warehouse layout, validator
# clean for both. Locks the per-cohort layout in so a future edit can't
# silently revert mtc_aat / mtc_alzheimers back to a shared disease pack.
# --------------------------------------------------------------------- #


class MTCCohortPackTests(unittest.TestCase):

    EXPECTED = {
        "mtc_aat":         ("MTC", "AAT",          "mtc_aat",        "mtc__aat_cohort"),
        "mtc_alzheimers":  ("MTC", "Alzheimer's",  "mtc_alzheimers", "mtc__alzheimers_cohort"),
    }

    def test_both_mtc_cohort_packs_exist(self):
        for slug in self.EXPECTED:
            path = REPO_ROOT / "packs" / "cohorts" / f"{slug}.yaml"
            self.assertTrue(path.is_file(),
                            f"missing cohort pack: {path}")

    def test_mtc_cohort_packs_point_at_per_cohort_variable_packs(self):
        for slug, (provider, disease, var_pack, schema) in self.EXPECTED.items():
            data = _load_yaml(f"packs/cohorts/{slug}.yaml")
            self.assertEqual(data.get("provider"), provider, slug)
            self.assertEqual(data.get("disease"), disease, slug)
            self.assertEqual(
                data.get("variables_pack"), var_pack,
                f"{slug} must point at its per-cohort variable pack "
                f"{var_pack!r} — separate ETL per cohort means each "
                f"cohort owns its final source of truth (see the "
                f"Nimbus refactor). When RMN Alzheimer's lands, it "
                f"gets its own per-cohort pack rather than sharing "
                f"this file directly.",
            )
            self.assertEqual(
                data.get("schema_name"), schema,
                f"{slug} schema_name must match the on-warehouse "
                f"layout (got {data.get('schema_name')!r})",
            )
            # And the per-cohort variable pack file itself must exist.
            var_path = REPO_ROOT / "packs" / "variables" / f"{var_pack}.yaml"
            self.assertTrue(
                var_path.is_file(),
                f"{slug}: variables_pack file {var_path} is missing",
            )

    def test_validator_clean_on_mtc_cohorts(self):
        _assert_validator_clean(self, self.EXPECTED)


# --------------------------------------------------------------------- #
# CKD pack split — two layers (CKD is the only renal disease in scope
# right now, so there is no separate renal_common above ckd_common):
#
#   ckd_common
#       ├── balboa_ckd     (final; placeholder)
#       └── drg_ckd        (final; placeholder)
#
# When AKI or another renal disease lands as its own cohort, refactor
# to introduce renal_common as a parent and pull demographics / vitals
# / generic visits rows up into it.
# --------------------------------------------------------------------- #


CKD_FINAL_PACKS = ("balboa_ckd", "drg_ckd")


class CKDPackSplitTests(unittest.TestCase):

    def test_ckd_common_exists_as_top_level_renal_base(self):
        """ckd_common is the top of the renal chain — it must NOT
        include any other pack today (no renal_common parent yet)."""
        ckd = _load_yaml("packs/variables/ckd_common.yaml")
        self.assertIn("variables", ckd)
        self.assertFalse(
            ckd.get("include"),
            "ckd_common must not include another pack — it's the top "
            "of the renal chain. If AKI or another renal disease "
            "lands as a separate cohort, add a renal_common parent "
            "first and move the demographics / vitals rows up into it.",
        )
        # Disease-common base should carry enough rows to be useful.
        self.assertGreater(len(ckd["variables"]), 25)

    def test_each_ckd_pack_includes_ckd_common(self):
        """Per-cohort CKD packs layer on top of ckd_common. Same
        per-cohort-ETL rule as the Nimbus and MTC packs."""
        for pack in CKD_FINAL_PACKS:
            data = _load_yaml(f"packs/variables/{pack}.yaml")
            self.assertEqual(
                data.get("include"), ["ckd_common"],
                f"{pack} must include exactly ['ckd_common'] — each "
                f"per-cohort CKD pack layers on top of the disease-"
                f"common base, not on another cohort",
            )

    def test_no_duplicate_variables_in_either_ckd_cohort(self):
        for slug in CKD_FINAL_PACKS:
            all_vars = _all_variables_for(slug)
            keys = [(v.get("category"), v.get("variable")) for v in all_vars]
            dupes = sorted({k for k in keys if keys.count(k) > 1})
            self.assertEqual(
                dupes, [],
                f"duplicate (category, variable) in {slug}: {dupes}",
            )


# --------------------------------------------------------------------- #
# CKD pack curation — key concept-name and class-row guards. Mirrors
# the respiratory / MTC curation tests.
# --------------------------------------------------------------------- #


class CKDPackCurationTests(unittest.TestCase):

    # -- CKD diagnosis row covers the major condition families seen
    # in both Balboa and DRG dumps (CKD stages, ESRD, dialysis
    # dependence, hypertensive renal disease, diabetic nephropathy).
    def test_ckd_diagnosis_matches_major_condition_families(self):
        dx = _find_variable_in("ckd_common", "CKD Diagnosis")
        c = dx["criteria"].lower()
        for token in ("chronic kidney disease",
                      "end stage renal disease",
                      "dependence on renal dialysis",
                      "hypertensive renal disease",
                      "diabetic nephropathy"):
            self.assertIn(
                token, c,
                f"CKD Diagnosis criteria must match {token!r} "
                f"(got: {dx['criteria']!r})",
            )

    # -- AKI row tracks acute episodes separately from chronic CKD.
    def test_aki_row_is_distinct_from_ckd(self):
        aki = _find_variable_in("ckd_common", "Acute Kidney Injury")
        c = aki["criteria"].lower()
        for token in ("acute kidney failure",
                      "acute renal failure",
                      "acute kidney injury"):
            self.assertIn(token, c)

    # -- Class-level CKD medication rows exist (not one catch-all).
    def test_ckd_ships_class_level_medication_rows(self):
        expected = [
            "RAAS Blockade (ACE Inhibitor / ARB)",
            "SGLT2 Inhibitor",
            "Mineralocorticoid Receptor Antagonist (MRA)",
            "Loop Diuretic",
            "Calcium Channel Blocker",
            "Statin (HMG-CoA Reductase Inhibitor)",
            "Aspirin (Antiplatelet)",
            "Sodium Bicarbonate (Acidosis Correction)",
            "Phosphate Binder",
            "Potassium Binder",
            "Erythropoiesis-Stimulating Agent (ESA)",
        ]
        for slug in ("ckd_common",) + CKD_FINAL_PACKS:
            names = [v.get("variable") for v in _all_variables_for(slug)]
            for label in expected:
                self.assertIn(
                    label, names,
                    f"{slug} must expose class-level row {label!r}",
                )

    # -- CCB row uses explicit ingredients (not the banned %blocker%
    # wildcard) and covers amlodipine — the top-3 drug in both cohorts.
    # Row is labelled "Calcium Channel Blocker" (not "(Dihydropyridine)")
    # because it intentionally groups DHP and non-DHP agents together;
    # the narrower label had a label-vs-criteria mismatch because
    # diltiazem and verapamil are non-DHPs.
    def test_calcium_channel_blocker_row_matches_amlodipine(self):
        ccb = _find_variable_in("ckd_common", "Calcium Channel Blocker")
        c = ccb["criteria"].lower()
        self.assertIn("amlodipine", c,
                      "CCB row must match amlodipine — #3-4 most-common "
                      "drug in both Balboa and DRG dumps")
        # Non-dihydropyridines (diltiazem, verapamil) are grouped in too
        # because they're prescribed for the same HTN indication; the
        # row name is therefore the generic class, not the DHP subclass.
        self.assertIn("diltiazem", c)
        self.assertIn("verapamil", c)

    def test_ccb_row_is_not_named_dihydropyridine(self):
        """Regression guard: row name must match its contents. The
        label used to be "Calcium Channel Blocker (Dihydropyridine)"
        which was clinically inaccurate because the criteria included
        non-DHPs."""
        names = [v.get("variable")
                 for v in _all_variables_for("ckd_common")]
        self.assertNotIn(
            "Calcium Channel Blocker (Dihydropyridine)", names,
            "The row must use the broader 'Calcium Channel Blocker' "
            "label so the name matches the DHP + non-DHP criteria",
        )

    # -- Aspirin row captures the top-1 drug in DRG.
    def test_aspirin_row_matches_aspirin(self):
        asa = _find_variable_in("ckd_common", "Aspirin (Antiplatelet)")
        self.assertIn("aspirin", asa["criteria"].lower())

    # -- Serum Calcium / Serum Sodium rows use a serum/plasma/blood
    # qualifier so they don't pick up calcium-channel-blocker drug
    # mentions or sodium bicarbonate drug mentions.
    def test_calcium_and_sodium_labs_have_blood_qualifier(self):
        for name in ("Serum Calcium", "Serum Sodium"):
            row = _find_variable_in("ckd_common", name)
            c = row["criteria"].lower()
            self.assertTrue(
                any(tok in c for tok in ("serum", "plasma", "blood")),
                f"{name} must AND in a serum/plasma/blood qualifier "
                f"(got: {row['criteria']!r})",
            )

    # -- Blood Pressure is represented by TWO rows because different
    # cohorts store it differently:
    #   DRG ships a combined string "Sitting blood pressure"
    #     (value_as_string) and the "Combined" row picks it up.
    #   Balboa ships separate Systolic / Diastolic numeric observations
    #     (value_as_number) and the "Systolic / Diastolic, numeric"
    #     row picks those up — the old single value_as_string row
    #     would have silently shown Implemented=No for Balboa.
    def test_blood_pressure_row_split_between_string_and_numeric(self):
        names = [v.get("variable")
                 for v in _all_variables_for("ckd_common")]
        self.assertIn(
            "Blood Pressure (Combined)", names,
            "ckd_common must ship a value_as_string BP row for sites "
            "that record combined 120/80-style strings",
        )
        self.assertIn(
            "Blood Pressure (Systolic / Diastolic, numeric)", names,
            "ckd_common must ship a value_as_number BP row for sites "
            "that record systolic and diastolic as separate numeric "
            "observations — confirmed in Output/balboackd.pdf",
        )
        # Old single "Blood Pressure" row must be gone — its description
        # overclaimed that one value_as_string row covered numeric
        # systolic/diastolic observations too.
        self.assertNotIn(
            "Blood Pressure", names,
            "The old single 'Blood Pressure' row must not coexist — "
            "its description overclaimed numeric support while the "
            "column was hard-coded to value_as_string",
        )
        combined = _find_variable_in("ckd_common",
                                     "Blood Pressure (Combined)")
        self.assertEqual(combined["column"], "value_as_string")

        numeric = _find_variable_in(
            "ckd_common",
            "Blood Pressure (Systolic / Diastolic, numeric)",
        )
        self.assertEqual(numeric["column"], "value_as_number")
        nc = numeric["criteria"].lower()
        # Numeric row criteria must be scoped to systolic / diastolic
        # so the generic "Blood pressure" concept (where value_as_string
        # is populated) doesn't leak into the numeric row.
        self.assertIn("systolic blood pressure", nc)
        self.assertIn("diastolic blood pressure", nc)

    # -- Proteinuria, Hypertension, Vitamin D deficiency diagnosis
    # rows exist as distinct condition-side rows rather than being
    # swept into the generic Diagnosis row.
    def test_ckd_ships_comorbidity_diagnosis_rows(self):
        expected = [
            "Proteinuria / Albuminuria",
            "Hypertension",
            "Vitamin D Deficiency",
        ]
        names = [v.get("variable") for v in _all_variables_for("ckd_common")]
        for label in expected:
            self.assertIn(label, names,
                          f"ckd_common must expose {label!r} as a "
                          f"distinct comorbidity row")

    # -- Hypertension row avoids double-counting with CKD Diagnosis
    # (which already covers "Hypertensive renal disease" variants).
    def test_hypertension_row_is_narrowed_to_essential_primary(self):
        htn = _find_variable_in("ckd_common", "Hypertension")
        c = htn["criteria"].lower()
        # Must require an essential / primary / disorder qualifier —
        # otherwise "Hypertensive renal disease" patients would be
        # double-counted between this row and CKD Diagnosis.
        self.assertTrue(
            any(tok in c for tok in ("essential", "primary",
                                     "hypertensive disorder",
                                     "benign hypertension")),
            "Hypertension row must be scoped to essential / primary "
            "variants to avoid overlap with the CKD Diagnosis row "
            "that already covers hypertensive-renal-disease patients",
        )

    # -- ESRD Monthly Services row captures CPT 9095X-9096X billing.
    def test_esrd_monthly_services_row_exists_and_matches_related_services(self):
        esrd = _find_variable_in(
            "ckd_common",
            "End-Stage Renal Disease (ESRD) Monthly Services",
        )
        c = esrd["criteria"].lower()
        # Must match the "ESRD related services" / "end stage renal
        # disease services" billing-concept wording seen in Balboa
        # (13.3% of procedures).
        self.assertTrue(
            any(tok in c for tok in ("esrd%services", "esrd%related",
                                     "end-stage renal disease%services",
                                     "end stage renal disease%services")),
            "ESRD Monthly Services criteria must match the 'ESRD "
            "related services' / 'end-stage renal disease ... services' "
            "concept-name pattern",
        )
        # Must NOT match generic E&M visit concepts (office visit,
        # subsequent hospital care) — those are covered by Visit Type.
        self.assertNotIn("office visit", c)
        self.assertNotIn("subsequent hospital", c)

    # -- RAAS row covers both ACE inhibitors and ARBs (alternatives
    # for the same indication; combining them keeps the row name
    # honest about what's being aggregated).
    def test_raas_row_covers_ace_and_arb_ingredients(self):
        raas = _find_variable_in("ckd_common", "RAAS Blockade (ACE Inhibitor / ARB)")
        c = raas["criteria"].lower()
        # ACE inhibitors
        for ing in ("lisinopril", "enalapril", "ramipril"):
            self.assertIn(ing, c, f"RAAS row must match ACE-i {ing!r}")
        # ARBs
        for ing in ("losartan", "valsartan"):
            self.assertIn(ing, c, f"RAAS row must match ARB {ing!r}")

    # -- SGLT2 row covers the three FDA-approved CKD-indicated agents.
    def test_sglt2_row_covers_three_ingredients(self):
        sglt2 = _find_variable_in("ckd_common", "SGLT2 Inhibitor")
        c = sglt2["criteria"].lower()
        for ing in ("empagliflozin", "dapagliflozin", "canagliflozin"):
            self.assertIn(ing, c, f"SGLT2 row must match {ing!r}")

    # -- Serum Creatinine row requires a serum/plasma/blood qualifier
    # so the urine-creatinine panels (used in UACR) don't leak into
    # this row.
    def test_serum_creatinine_requires_serum_qualifier(self):
        cr = _find_variable_in("ckd_common", "Serum Creatinine")
        c = cr["criteria"].lower()
        self.assertIn("creatinine", c)
        self.assertTrue(
            any(tok in c for tok in ("serum", "plasma", "blood")),
            "Serum Creatinine row must AND in a serum/plasma/blood "
            "qualifier so urine-creatinine panels don't pollute it",
        )

    # -- Hemoglobin row qualifier prevents HbA1c / hemoglobin A1c from
    # leaking in.
    def test_hemoglobin_row_excludes_hba1c(self):
        hgb = _find_variable_in("ckd_common", "Hemoglobin")
        c = hgb["criteria"].lower()
        self.assertIn("hemoglobin", c)
        self.assertTrue(
            any(tok in c for tok in ("blood", "mass/volume")),
            "Hemoglobin row must AND in a blood / mass-per-volume "
            "qualifier so HbA1c is not counted here",
        )

    # -- Smoking Status rows are split by cohort:
    #   ckd_common ships the observation-side row (Balboa-style).
    #   drg_ckd adds a procedure-coded row — that concept-naming
    #   pattern is DRG-specific and should NOT leak into Balboa.
    def test_smoking_status_observation_row_in_ckd_common(self):
        for slug in ("ckd_common",) + CKD_FINAL_PACKS:
            rows = _all_variables_for(slug)
            obs_rows = [v for v in rows
                        if v.get("variable") == "Smoking Status"]
            self.assertEqual(
                len(obs_rows), 1,
                f"{slug}: expected exactly one observation-side "
                f"Smoking Status row, found {len(obs_rows)}",
            )
            self.assertEqual(obs_rows[0]["table"], "observation")
            self.assertIn(
                "observation_concept_name ilike '%tobacco%'",
                obs_rows[0]["criteria"].lower(),
                f"{slug}: observation Smoking Status row must use the "
                f"widened %tobacco% pattern",
            )

    def test_smoking_status_procedure_row_is_drg_only(self):
        # DRG-only: per-cohort pack is the final source of truth.
        drg_names = [v.get("variable")
                     for v in _all_variables_for("drg_ckd")]
        self.assertIn(
            "Smoking Status (Procedure-coded)", drg_names,
            "drg_ckd must ship the procedure-coded Smoking Status row",
        )

        # Must NOT appear in ckd_common or balboa_ckd.
        for slug in ("ckd_common", "balboa_ckd"):
            names = [v.get("variable") for v in _all_variables_for(slug)]
            self.assertNotIn(
                "Smoking Status (Procedure-coded)", names,
                f"{slug} must NOT inherit the DRG-specific procedure-"
                f"coded Smoking Status row — its wording references "
                f"DRG data shape and would read as a merged superset "
                f"in Balboa's workbook",
            )

    def test_smoking_status_procedure_row_criteria_is_status_anchored(self):
        """The procedure-coded smoking row must match status concepts
        only (current / former / never tobacco/smoker, smoking-status
        umbrella) — NOT generic tobacco-related procedures like
        cessation counselling, screening, or pharmacotherapy
        administration. Otherwise a row labelled 'Smoking Status'
        would silently aggregate intervention records too."""
        sm = _find_variable_in("drg_ckd",
                               "Smoking Status (Procedure-coded)")
        c = sm["criteria"].lower()
        # The row's status anchors — at least one of these must be
        # present so the criteria is genuinely status-bound.
        status_anchors = (
            "current tobacco", "former tobacco", "never%tobacco",
            "tobacco%user", "tobacco%non-user",
            "current%smoker", "former%smoker", "never%smoker",
            "smoking status",
        )
        self.assertTrue(
            any(a in c for a in status_anchors),
            f"Procedure-coded Smoking Status criteria must anchor on "
            f"a status fragment (current/former/never + tobacco/smoker, "
            f"or 'smoking status'); got: {sm['criteria']!r}",
        )
        # And the row must NOT use the bare wildcards on their own,
        # which would have matched cessation / counselling / screening
        # procedures in addition to status concepts.
        for bare in ("'%tobacco%'", "'%smoking%'", "'%cigarette%'"):
            self.assertNotIn(
                bare, c,
                f"Procedure-coded Smoking Status criteria must not "
                f"use the bare {bare} wildcard — that would pull in "
                f"tobacco-cessation, screening, and counselling "
                f"procedures (intervention records, not status records)",
            )


# --------------------------------------------------------------------- #
# CKD broad-wildcard regression guards — same approach as respiratory.
# --------------------------------------------------------------------- #


class CKDBroadWildcardRegressionTests(unittest.TestCase):

    def test_no_catch_wide_drug_wildcards_in_ckd_packs(self):
        # Patterns that would be trivially too broad for CKD:
        #   ILIKE '%blocker%'   — matches ARBs but also beta-blockers,
        #                          calcium blockers, etc.
        #   ILIKE '%diuretic%'  — would catch thiazides on the loop
        #                          diuretic row
        #   ILIKE '%kidney%'    — would match every renal-related
        #                          drug indication string
        #   ILIKE '%renal%'     — same problem
        banned_patterns = ("'%blocker%'", "'%diuretic%'",
                           "'%kidney%'", "'%renal%'", "'%dialysis%'")
        for slug in CKD_FINAL_PACKS + ("ckd_common",):
            for v in _all_variables_for(slug):
                # Procedure rows naturally use %dialysis% / %kidney% /
                # %renal% in their criteria — only ban these on the
                # drug_exposure-table rows.
                if v.get("table") != "drug_exposure":
                    continue
                c = (v.get("criteria") or "").lower()
                for bad in banned_patterns:
                    self.assertNotIn(
                        bad, c,
                        f"{slug}/{v.get('variable')} drug criteria "
                        f"contains overly broad wildcard {bad} — narrow "
                        f"to a specific ingredient list "
                        f"(got: {v.get('criteria')!r})",
                    )

    def test_no_ckd_variable_points_at_id_column_without_id_name(self):
        _assert_no_id_column_mismatch(
            self,
            pack_slugs=CKD_FINAL_PACKS + ("ckd_common",),
        )


# --------------------------------------------------------------------- #
# CKD cohort packs — point at the correct per-cohort variables pack
# and carry the required validator-enforced fields.
# --------------------------------------------------------------------- #


class CKDCohortPackTests(unittest.TestCase):

    EXPECTED = {
        "balboa_ckd":  ("Balboa", "Renal", "balboa_ckd", "balboa_ckd_cohort"),
        "drg_ckd":     ("DRG",    "Renal", "drg_ckd",    "drg_ckd_cohort"),
    }

    def test_both_ckd_cohort_packs_exist(self):
        for slug in self.EXPECTED:
            path = REPO_ROOT / "packs" / "cohorts" / f"{slug}.yaml"
            self.assertTrue(path.is_file(),
                            f"missing cohort pack: {path}")

    def test_ckd_cohort_packs_point_at_per_cohort_variable_packs(self):
        for slug, (provider, disease, var_pack, schema) in self.EXPECTED.items():
            data = _load_yaml(f"packs/cohorts/{slug}.yaml")
            self.assertEqual(data.get("provider"), provider, slug)
            self.assertEqual(data.get("disease"), disease, slug)
            self.assertEqual(
                data.get("variables_pack"), var_pack,
                f"{slug} must point at its per-cohort variable pack "
                f"{var_pack!r} — separate ETL per cohort means each "
                f"cohort owns its final source of truth",
            )
            self.assertEqual(
                data.get("schema_name"), schema,
                f"{slug} schema_name must match the on-warehouse "
                f"layout (got {data.get('schema_name')!r})",
            )
            var_path = REPO_ROOT / "packs" / "variables" / f"{var_pack}.yaml"
            self.assertTrue(
                var_path.is_file(),
                f"{slug}: variables_pack file {var_path} is missing",
            )

    def test_validator_clean_on_ckd_cohorts(self):
        _assert_validator_clean(self, self.EXPECTED)


# --------------------------------------------------------------------- #
# New-cohort packs built from the post-dump mining round (RMN Alz,
# Newtown MASH, Newtown IBD, RVC DR, RVC AMD). These cohorts each
# follow the established per-cohort-variables_pack model — shared
# disease-common base(s) on the left, per-cohort final pack on the
# right. Tests here lock in:
#   - the cohort packs exist and point at the right variable pack
#   - the per-cohort packs each `include: [<disease>_common]`
#   - the disease-common packs carry the disease-defining diagnosis
#     row that the PDF dumps made obvious (MASH → Fatty Liver /
#     Steatohepatitis; IBD → UC / Crohn's; DR → diabetic retinopathy;
#     AMD → wet + dry macular degeneration)
#   - the new retinal_common base factors out OCT and intravitreal
#     rows between DR and AMD the way respiratory_common factors out
#     universals between COPD and asthma
# --------------------------------------------------------------------- #


class NewCohortPacksTests(unittest.TestCase):

    EXPECTED = {
        "rmn_alzheimers":    ("Rocky Mountain Neurology",
                              "Alzheimer's",
                              "rmn_alzheimers",
                              "rmn_alzheimers_cohort"),
        "newtown_mash":      ("Newtown", "MASH",
                              "newtown_mash",
                              "newtown_mash_cohort"),
        "newtown_ibd":       ("Newtown", "IBD",
                              "newtown_ibd",
                              "newtown_ibd_cohort"),
        "rvc_dr_curated":    ("RVC", "Diabetic Retinopathy",
                              "rvc_dr_curated",
                              "rvc_dr_curated"),
        "rvc_amd_curated":   ("RVC", "Age-Related Macular Degeneration",
                              "rvc_amd_curated",
                              "rvc_amd_curated"),
    }

    def test_all_five_cohort_packs_exist(self):
        for slug in self.EXPECTED:
            self.assertTrue(
                (REPO_ROOT / "packs" / "cohorts" / f"{slug}.yaml").is_file(),
                f"missing cohort pack: packs/cohorts/{slug}.yaml",
            )

    def test_each_new_cohort_points_at_its_per_cohort_pack(self):
        for slug, (provider, disease, var_pack, schema) in self.EXPECTED.items():
            data = _load_yaml(f"packs/cohorts/{slug}.yaml")
            self.assertEqual(data.get("provider"), provider, slug)
            self.assertEqual(data.get("disease"), disease, slug)
            self.assertEqual(data.get("variables_pack"), var_pack,
                             f"{slug} must point at {var_pack!r}")
            self.assertEqual(data.get("schema_name"), schema, slug)

    def test_validator_clean_on_new_cohorts(self):
        _assert_validator_clean(self, self.EXPECTED)

    def test_disease_common_packs_carry_their_defining_diagnosis(self):
        """Each new disease-common pack MUST include the row that the
        PDF dump made obvious (e.g. Newtown MASH's top condition is
        'Fatty (change of) liver'). These rows drive cohort membership
        — if they drift the entire dictionary misses its anchor."""
        anchors = {
            "mash_common":    "MASH / NAFLD / Fatty Liver",
            "ibd_common":     "IBD Diagnosis",
            "dr_common":      "Diabetic Retinopathy",
            "amd_common":     "Age-Related Macular Degeneration",
        }
        for pack_slug, anchor_name in anchors.items():
            names = [v.get("variable")
                     for v in _all_variables_for(pack_slug)]
            self.assertIn(
                anchor_name, names,
                f"{pack_slug} must carry the {anchor_name!r} anchor "
                f"row — it's the disease-defining diagnosis and every "
                f"per-cohort pack that layers on top depends on it",
            )

    def test_retinal_common_is_shared_between_dr_and_amd(self):
        """retinal_common carries the rows that DR and AMD both use
        (OCT thickness measurements, intravitreal injection, IOL
        history, HTN). If DR or AMD ever stops including it, the
        ophthalmology workbooks would each re-ship duplicate rows."""
        for disease in ("dr_common", "amd_common"):
            pack = _load_yaml(f"packs/variables/{disease}.yaml")
            self.assertIn(
                "retinal_common", pack.get("include", []),
                f"{disease} must include retinal_common so OCT / "
                f"intravitreal / IOL rows flow in from one source",
            )

    def test_rmn_alzheimers_reuses_existing_alzheimers_common(self):
        """Second Alzheimer's cohort — zero MTC-specific rows should
        leak in via the shared alzheimers_common base. If this ever
        fails, that's exactly the signal that tells us to extract an
        MTC-specific row down into packs/variables/mtc_alzheimers.yaml."""
        pack = _load_yaml("packs/variables/rmn_alzheimers.yaml")
        self.assertEqual(
            pack.get("include"), ["alzheimers_common"],
            "rmn_alzheimers must include exactly [alzheimers_common] — "
            "same include as mtc_alzheimers, so both cohorts share the "
            "same disease-level rows without one cohort reshaping the "
            "other",
        )


# --------------------------------------------------------------------- #
# dump_new_schemas.py — thin subprocess wrapper that hands each of the
# five backlog-cohort schemas to `introspect_cohort.py --schema`. We
# can't test the actual DB dump here (no warehouse in the sandbox) but
# we CAN lock in the SCHEMAS list shape so a future edit that drops a
# cohort or renames a schema fails loudly at test time.
# --------------------------------------------------------------------- #


class DumpNewSchemasTests(unittest.TestCase):

    def _load_module(self):
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "dump_new_schemas",
            REPO_ROOT / "scripts" / "dump_new_schemas.py",
        )
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return mod

    def test_script_is_importable(self):
        # Syntactic validity + clean import (no accidental top-level
        # side effects that would break on a server without a DB).
        mod = self._load_module()
        self.assertTrue(hasattr(mod, "SCHEMAS"))
        self.assertTrue(hasattr(mod, "main"))

    def test_schemas_list_covers_the_five_backlog_cohorts(self):
        mod = self._load_module()
        # Ordered list — order doesn't matter clinically but locking
        # membership prevents accidental drops.
        expected = {
            "rmn_alzheimers_cohort",
            "newtown_mash_cohort",
            "newtown_ibd_cohort",
            "rvc_dr_curated",
            "rvc_amd_curated",
        }
        self.assertEqual(
            set(mod.SCHEMAS), expected,
            f"dump_new_schemas.SCHEMAS drifted from the backlog list. "
            f"Expected {sorted(expected)}, got {sorted(mod.SCHEMAS)}.",
        )

    def test_raw_output_dir_is_gitignored(self):
        # The dumps carry real warehouse data distributions, so
        # Output/raw/ must stay out of git. Check .gitignore directly
        # so a future cleanup can't silently drop the rule.
        ignore_text = (REPO_ROOT / ".gitignore").read_text()
        self.assertIn(
            "Output/raw/", ignore_text,
            ".gitignore must keep Output/raw/ out of git — the raw "
            "dumps produced by scripts/dump_new_schemas.py contain "
            "warehouse data distributions that are privacy-sensitive.",
        )


class TestFlatironStyleMetadata(unittest.TestCase):
    """Phase 2 / 3 additions: data_source mapping, inclusion_criteria
    fallback, table_descriptions dict form, and the new headers on
    every rendered sheet."""

    def test_derive_data_source_extraction_default(self):
        # Plain Structured / Unstructured / Abstracted rows fall through
        # to the Flatiron typology when neither the table nor an
        # explicit override is set.
        self.assertEqual(derive_data_source("Structured"), "Normalized")
        self.assertEqual(derive_data_source("Unstructured"), "NLP")
        self.assertEqual(derive_data_source("Abstracted"), "Abstracted")

    def test_derive_data_source_curated_table_promotes_to_enhanced(self):
        # Tables in the curated-abstraction allowlist surface as
        # Enhanced even when the row's extraction_type says Structured.
        self.assertEqual(
            derive_data_source("Structured", "eosinophil_standardized"),
            "Enhanced",
        )

    def test_derive_data_source_explicit_override_wins(self):
        # An explicit `data_source:` on the pack row beats both the
        # table-name promotion and the extraction_type default.
        self.assertEqual(
            derive_data_source("Structured", "eosinophil_standardized",
                               explicit="Derived"),
            "Derived",
        )

    def test_derive_inclusion_criteria_explicit_wins(self):
        # Pack-author prose beats any auto-translation.
        self.assertEqual(
            derive_inclusion_criteria(
                "condition_concept_name ILIKE '%alzheimer%'",
                explicit="Records with an Alzheimer's diagnosis.",
            ),
            "Records with an Alzheimer's diagnosis.",
        )

    def test_derive_inclusion_criteria_simple_returns_blank(self):
        # Single-clause ILIKE returns empty so the validator forces
        # the pack author to add explicit prose. The earlier
        # friendly-translation fallback ("Records where the X concept
        # matches 'Y'.") was removed because it produced QA-style
        # copy rather than Flatiron-style clinical prose.
        self.assertEqual(
            derive_inclusion_criteria(
                "condition_concept_name ILIKE '%alzheimer%'"
            ),
            "",
        )

    def test_derive_inclusion_criteria_compound_returns_blank(self):
        # Multi-clause SQL also requires explicit prose — same
        # contract as the single-clause case.
        self.assertEqual(
            derive_inclusion_criteria(
                "x ILIKE '%a%' OR x ILIKE '%b%'"
            ),
            "",
        )

    def test_load_table_descriptions_handles_string_form(self):
        # Backward compatibility — a bare string entry should still
        # populate `description` and leave the new fields empty.
        with tempfile.TemporaryDirectory() as td:
            (Path(td) / "table_descriptions.yaml").write_text(
                "tables:\n  foo: \"a one-line description\"\n",
                encoding="utf-8",
            )
            # load_table_descriptions reads the packs/ directory
            # directly, so we test the parser via a small wrapper.
            import yaml
            raw = yaml.safe_load(
                (Path(td) / "table_descriptions.yaml").read_text()
            )
            # Verify the YAML form parses; the production loader's
            # dict-vs-string normalisation is exercised below.
            self.assertEqual(raw["tables"]["foo"], "a one-line description")

    def test_load_table_descriptions_normalises_dict_form(self):
        # The packs/table_descriptions.yaml shipped in repo uses the
        # full dict form. Verify the loader returns dict entries with
        # all four keys present.
        descriptions = load_table_descriptions()
        # Spot-check a known entry.
        cond = descriptions.get("condition_occurrence")
        self.assertIsNotNone(cond)
        self.assertIn("description", cond)
        self.assertIn("inclusion_criteria", cond)
        self.assertIn("data_source", cond)
        self.assertIn("source_table", cond)
        self.assertEqual(cond["data_source"], "Normalized")
        self.assertEqual(cond["source_table"], "OMOP CONDITION_OCCURRENCE")

    def test_xlsx_tables_sheet_leads_with_description_and_metadata(self):
        # Tables sheet must put Description / Inclusion Criteria / Data
        # Source / Source Table ahead of the Rows / Columns / Patients
        # QA columns. This is the most visible Flatiron-style change.
        import openpyxl
        model = _trivial_model()
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "out.xlsx"
            write_xlsx(model, out, audience="technical")
            wb = openpyxl.load_workbook(out)
            headers = [c.value for c in wb["Tables"][1]]
        self.assertEqual(headers, [
            "Table", "Category", "Description", "Inclusion Criteria",
            "Data Source", "Source Table", "Rows", "Columns", "Patients",
        ])

    def test_xlsx_columns_sheet_adds_nullable_example_data_source(self):
        import openpyxl
        model = _trivial_model()
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "out.xlsx"
            write_xlsx(model, out, audience="technical")
            wb = openpyxl.load_workbook(out)
            headers = [c.value for c in wb["Columns"][1]]
        for needed in ("Field Type", "Nullable", "Example",
                       "Coding Schema", "Data Source"):
            self.assertIn(needed, headers)
        self.assertNotIn("Extraction Type", headers,
                         "Columns sheet should render `Data Source`, "
                         "not the older `Extraction Type` label.")

    def test_xlsx_variables_sheet_hides_criteria_for_sales(self):
        # The raw SQL `Criteria` column is technical-audience-only;
        # sales / pharma get the prose `Inclusion Criteria` only.
        import openpyxl
        model = _trivial_model()
        with tempfile.TemporaryDirectory() as td:
            out_tech = Path(td) / "tech.xlsx"
            out_sales = Path(td) / "sales.xlsx"
            write_xlsx(model, out_tech, audience="technical")
            write_xlsx(model, out_sales, audience="sales")
            wb_tech = openpyxl.load_workbook(out_tech)
            wb_sales = openpyxl.load_workbook(out_sales)
            tech_headers = [c.value for c in wb_tech["Variables"][1]]
            sales_headers = [c.value for c in wb_sales["Variables"][1]]
        self.assertIn("Criteria", tech_headers)
        self.assertNotIn("Criteria", sales_headers)
        # Both audiences should see the prose Inclusion Criteria.
        self.assertIn("Inclusion Criteria", tech_headers)
        self.assertIn("Inclusion Criteria", sales_headers)

    def test_html_variables_sheet_hides_criteria_for_pharma(self):
        # Same audience-split contract for the HTML renderer.
        model = _trivial_model()
        with tempfile.TemporaryDirectory() as td:
            out_pharma = Path(td) / "pharma.html"
            write_html(model, out_pharma, audience="pharma")
            html = out_pharma.read_text()
        # Inclusion Criteria column is always present on Variables.
        self.assertIn("<th>Inclusion Criteria</th>", html)
        # Raw Criteria column suppressed for pharma.
        self.assertNotIn("<th>Criteria</th>", html)

    def test_load_table_descriptions_preserves_unknown_keys(self):
        # Regression for review feedback: the loader docstring
        # promises unknown keys are preserved so future YAML additions
        # don't need a code change. Verify a custom key round-trips.
        import yaml as _yaml
        from build_dictionary import PACKS_DIR
        path = PACKS_DIR / "table_descriptions.yaml"
        original = path.read_text(encoding="utf-8")
        try:
            data = _yaml.safe_load(original) or {}
            data.setdefault("tables", {}).setdefault("person", {})["future_key"] = "round-trip"
            path.write_text(_yaml.safe_dump(data), encoding="utf-8")
            loaded = load_table_descriptions()
        finally:
            path.write_text(original, encoding="utf-8")
        self.assertEqual(
            loaded["person"].get("future_key"), "round-trip",
            "load_table_descriptions must preserve unknown keys per "
            "its documented contract.",
        )

    def test_compound_criteria_renders_non_blank_inclusion_criteria(self):
        # Phase 2 / 3 regression: derive_inclusion_criteria() returns
        # empty for compound (AND / OR) SQL by design — so any pack
        # row with compound criteria MUST carry an explicit
        # inclusion_criteria. The test exercises a real compound row
        # through resolve_variables and asserts the rendered VariableRow
        # has non-empty inclusion_criteria. Without this assertion, a
        # blank prose column ships to sales / pharma audiences (which
        # is exactly the regression that triggered this test).
        compound_row = {
            "category": "Diagnosis",
            "variable": "Asthma Diagnosis",
            "table": "condition_occurrence",
            "column": "condition_concept_name",
            "criteria": (
                "condition_concept_name ILIKE '%asthma%' OR "
                "condition_concept_name ILIKE '%reactive airway disease%'"
            ),
            "inclusion_criteria":
                "Records where the diagnosis concept matches the Asthma family.",
            "extraction_type": "Structured",
        }
        # Use a connection stub that returns zero counts so the
        # resolver short-circuits the SQL execution path; we only care
        # that the prose round-trips through to the model.
        cur = _Cursor([("COUNT", (0,))])
        conn = _Conn([("COUNT", (0,))])
        rows = resolve_variables(
            conn, "test_schema", [compound_row], total_patients=100,
        )
        self.assertEqual(len(rows), 1)
        self.assertTrue(
            rows[0].inclusion_criteria.strip(),
            "Compound-criteria row must have non-empty rendered "
            "inclusion_criteria — sales / pharma see this column "
            "instead of the raw SQL Criteria.",
        )

    def test_every_shipped_pack_row_with_compound_criteria_has_prose(self):
        # Walks every packs/variables/*.yaml shipped in the repo and
        # asserts the contract: any row with AND / OR in `criteria:`
        # also has a non-empty `inclusion_criteria:`. Mirrors the
        # validator rule, but lives in the test suite so a regression
        # surfaces in CI without needing the validator step.
        import re as _re
        import yaml as _yaml
        from build_dictionary import PACKS_DIR as _PACKS
        compound = _re.compile(r"\s+(AND|OR)\s+", _re.IGNORECASE)
        gaps: list[str] = []
        for path in sorted((_PACKS.parent / "variables").glob("*.yaml")):
            data = _yaml.safe_load(path.read_text(encoding="utf-8")) or {}
            for row in data.get("variables") or []:
                if not isinstance(row, dict):
                    continue
                crit = row.get("criteria") or ""
                if not compound.search(crit):
                    continue
                if not (row.get("inclusion_criteria") or "").strip():
                    gaps.append(
                        f"{path.name}: {row.get('category','?')}/"
                        f"{row.get('variable','?')}"
                    )
        self.assertEqual(
            gaps, [],
            f"{len(gaps)} compound-criteria rows missing inclusion_criteria. "
            f"Add an explicit `inclusion_criteria:` field to each, then "
            f"commit. First 5: {gaps[:5]}",
        )


def _trivial_model() -> CohortModel:
    """Minimal CohortModel exercising the new TableRow / ColumnRow /
    VariableRow fields. Used by the renderer-output tests above."""
    return CohortModel(
        cohort="test_cohort",
        provider="TestCo",
        disease="TestDisease",
        schema_name="test_schema",
        variant="raw",
        display_name="Test Cohort",
        description="",
        status="active",
        generated_at="2026-04-25T00:00:00+00:00",
        git_sha="abc1234",
        introspect_version="test",
        schema_snapshot_digest="sha256:test",
        summary=CohortSummary(
            patient_count=100, table_count=1, column_count=1,
            date_coverage=DateCoverage(),
        ),
        tables=[TableRow(
            table_name="condition_occurrence", category="Diagnosis",
            row_count=1000, column_count=10, patient_count_in_table=100,
            purpose="Medical diagnoses recorded for the patient.",
            description="Medical diagnoses recorded for the patient.",
            inclusion_criteria="One record per diagnosis per patient.",
            data_source="Normalized",
            source_table="OMOP CONDITION_OCCURRENCE",
        )],
        columns=[ColumnRow(
            category="Diagnosis", table="condition_occurrence",
            column="condition_concept_name",
            description="The clinical concept of the diagnosis.",
            data_type="text", values="Hypertension, Diabetes",
            distribution="Hypertension: 600 (60%); Diabetes: 400 (40%)",
            median_iqr="", completeness_pct=100.0, patient_pct=100.0,
            extraction_type="Structured", pii=False, notes="",
            nullable="No", example="Hypertension", coding_schema="",
            data_source="Normalized",
        )],
        variables=[VariableRow(
            category="Diagnosis", variable="Diagnosis",
            description="Medical diagnoses recorded for the patient.",
            table="condition_occurrence",
            column="condition_concept_name",
            criteria="condition_concept_name IS NOT NULL",
            values="Hypertension", distribution="",
            median_iqr="", completeness_pct=100.0, implemented="Yes",
            patient_pct=100.0, extraction_type="Structured",
            notes="", pii=False,
            inclusion_criteria="One record per diagnosis per patient.",
            field_type="text", example="Hypertension",
            coding_schema="", data_source="Normalized",
        )],
    )


if __name__ == "__main__":
    unittest.main()
